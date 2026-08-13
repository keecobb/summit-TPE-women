"""Automated sanity check for WomensSummitTPE.xlsx -- catches the SHAPE of
corruption this project has already hit twice by hand (the Team 299/300
shared-ESPN-ID collision, and the backfill_pgs_team_300.py bug that dumped
~340 opposing players onto Team 300's roster), plus a handful of other
structural checks that would have caught either one immediately instead of
waiting for someone to eyeball a roster and notice it looks wrong.
 
This is READ-ONLY -- it never writes to the workbook. Run it:
  - after any scrape/backfill script touches the workbook
  - after (or as part of) build_cache.py, before pointing a live site at
    the resulting cache
  - any time something "feels off" and you want a fast structural check
    before digging in by hand
 
Exit code is 1 if any ERROR-level finding exists (so this can gate a
pipeline -- e.g. `check_data_health.py && build_cache.py`), 0 otherwise
(WARNINGs alone don't block anything, they're just flagged for a human).
 
Usage:
    python check_data_health.py --path WomensSummitTPE.xlsx
    python check_data_health.py --path WomensSummitTPE.xlsx --max-roster 25 --min-roster 6
"""
 
import argparse
import sys
from collections import Counter, defaultdict
 
# Real D1 rosters run roughly 10-18 (starters + bench + a couple walk-ons).
# The Players sheet is a CURRENT-roster snapshot, not a multi-season log, so
# this stays tight -- the 360-player bug blew right past any reasonable
# number here. Kept generous on the high end to avoid false alarms for a
# team with unusual walk-on/transfer churn.
DEFAULT_MAX_ROSTER = 26
DEFAULT_MIN_ROSTER = 6
# Single-game outlier thresholds -- flagged as WARNING (rare real games do
# happen), not ERROR. FG/FT made > attempted, or team playing itself, are
# ERROR-level since those are never legitimate.
MAX_PLAUSIBLE_MINUTES = 55
MAX_PLAUSIBLE_POINTS = 65
MAX_PLAUSIBLE_REBOUNDS = 35
 
 
def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping
 
 
class Findings:
    def __init__(self):
        self.items = []  # (severity, category, message)
 
    def error(self, category, message):
        self.items.append(("ERROR", category, message))
 
    def warn(self, category, message):
        self.items.append(("WARNING", category, message))
 
    @property
    def errors(self):
        return [i for i in self.items if i[0] == "ERROR"]
 
    @property
    def warnings(self):
        return [i for i in self.items if i[0] == "WARNING"]
 
 
def load_sheets(path):
    import openpyxl
    print(f"Opening {path} read-only ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
 
    teams_ws = wb["Teams"]
    th = header_map(teams_ws)
    teams = {}
    for row in teams_ws.iter_rows(min_row=2, values_only=True):
        tid = row[th["Team ID"] - 1]
        if tid is None:
            continue
        teams[tid] = dict(
            name=row[th["Team"] - 1], division=row[th["Division"] - 1],
            conference=row[th["Conference"] - 1],
            espn_id=row[th["ESPN Team ID"] - 1] if "ESPN Team ID" in th else None,
        )
 
    games_ws = wb["Games"]
    gh = header_map(games_ws)
    games = []
    for row in games_ws.iter_rows(min_row=2, values_only=True):
        gid = row[gh["Game ID"] - 1]
        if gid is None:
            continue
        games.append(dict(
            game_id=gid, season=row[gh["Season"] - 1],
            home_id=row[gh["Home Team ID"] - 1], away_id=row[gh["Away Team ID"] - 1],
            home_name=row[gh["Home Team"] - 1], away_name=row[gh["Away Team"] - 1],
        ))
 
    players_ws = wb["Players"]
    plh = header_map(players_ws)
    players = []
    for row in players_ws.iter_rows(min_row=2, values_only=True):
        pid = row[plh["Player ID"] - 1]
        if pid is None:
            continue
        name = f"{row[plh['First Name'] - 1] or ''} {row[plh['Last Name'] - 1] or ''}".strip()
        players.append(dict(player_id=pid, name=name, team_id=row[plh["Team ID"] - 1]))
 
    pgs_ws = wb["PlayerGameStats"]
    pgh = header_map(pgs_ws)
    pgs_rows = []
    for row in pgs_ws.iter_rows(min_row=2, values_only=True):
        pid = row[pgh["Player ID"] - 1]
        if pid is None:
            continue
        pgs_rows.append(dict(
            player_id=pid, team_id=row[pgh["Team ID"] - 1], opp_team_id=row[pgh["Opponent Team ID"] - 1],
            game_id=row[pgh["Game ID"] - 1], season=row[pgh["Season"] - 1],
            minutes=row[pgh["Min"] - 1], points=row[pgh["Points"] - 1], rebounds=row[pgh["Rebound"] - 1],
            fgm=row[pgh["FG Made"] - 1], fga=row[pgh["FG Attempt"] - 1],
            fg3m=row[pgh["3FG M"] - 1], fg3a=row[pgh["3FG A"] - 1],
            ftm=row[pgh["FT M"] - 1], fta=row[pgh["FT A"] - 1],
        ))
 
    wb.close()
    return teams, games, players, pgs_rows
 
 
# ---------- individual checks ----------
 
def check_duplicate_espn_ids(teams, f):
    """The exact shape of the Team 299/300 bug: two internal Team IDs
    sharing one ESPN Team ID, so anything keyed off that ESPN ID writes
    both teams' data to whichever team resolves first."""
    by_espn = defaultdict(list)
    for tid, info in teams.items():
        if info["espn_id"]:
            by_espn[str(info["espn_id"])].append((tid, info["name"]))
    for espn_id, owners in by_espn.items():
        if len(owners) > 1:
            names = ", ".join(f"{name!r} (Team {tid})" for tid, name in owners)
            f.error("duplicate_espn_id",
                     f"ESPN Team ID {espn_id} is shared by {len(owners)} teams: {names}. "
                     f"Any scrape/backfill keyed off this ESPN ID will write both teams' data "
                     f"to whichever one resolves first.")
 
 
def check_missing_espn_ids(teams, f):
    missing = [(tid, info["name"]) for tid, info in teams.items() if not info["espn_id"]]
    if missing:
        names = ", ".join(f"{name!r} (Team {tid})" for tid, name in missing[:10])
        more = f" (+{len(missing) - 10} more)" if len(missing) > 10 else ""
        f.warn("missing_espn_id", f"{len(missing)} teams have no ESPN Team ID set, so they can't be "
                                   f"scraped: {names}{more}")
 
 
def check_roster_sizes(teams, players, f, max_roster, min_roster):
    """The exact shape of the backfill_pgs_team_300.py bug: one team's
    Players-sheet roster blows way past what a real D1 team looks like."""
    counts = Counter(p["team_id"] for p in players)
    for tid, count in counts.items():
        team = teams.get(tid)
        name = team["name"] if team else f"<unknown team {tid}>"
        if count > max_roster:
            f.error("roster_too_large",
                     f"Team {tid} ({name!r}) has {count} rows on the Players sheet -- more than "
                     f"{max_roster}, a real roster shouldn't be this large. This is the exact "
                     f"pattern of the backfill_pgs_team_300.py bug (opposing players misfiled onto "
                     f"one team). Run diagnose_team_300_roster.py --team-id {tid} to inspect.")
        elif count < min_roster:
            f.warn("roster_too_small",
                    f"Team {tid} ({name!r}) has only {count} rows on the Players sheet -- fewer than "
                    f"{min_roster}, may indicate an incomplete scrape.")
    # teams with zero players at all
    tracked_ids = set(counts)
    for tid, info in teams.items():
        if tid not in tracked_ids:
            f.warn("no_players", f"Team {tid} ({info['name']!r}) has zero rows on the Players sheet.")
 
 
def check_orphaned_team_refs(teams, games, players, pgs_rows, f):
    known = set(teams)
    for g in games:
        for side, tid in (("home", g["home_id"]), ("away", g["away_id"])):
            if tid is not None and tid not in known:
                f.error("orphaned_team_ref",
                         f"Game {g['game_id']} ({g['season']}) references {side} Team ID {tid}, "
                         f"which isn't on the Teams sheet.")
    seen_player_refs = set()
    for p in players:
        if p["team_id"] is not None and p["team_id"] not in known and p["team_id"] not in seen_player_refs:
            seen_player_refs.add(p["team_id"])
            f.error("orphaned_team_ref",
                     f"Players sheet has rows with Team ID {p['team_id']}, which isn't on the Teams sheet.")
    seen_pgs_refs = set()
    for r in pgs_rows:
        for label, tid in (("Team ID", r["team_id"]), ("Opponent Team ID", r["opp_team_id"])):
            if tid is not None and tid not in known and (label, tid) not in seen_pgs_refs:
                seen_pgs_refs.add((label, tid))
                f.error("orphaned_team_ref",
                         f"PlayerGameStats has rows with {label} {tid}, which isn't on the Teams sheet.")
 
 
def check_self_play(games, pgs_rows, f):
    for g in games:
        if g["home_id"] is not None and g["home_id"] == g["away_id"]:
            f.error("self_play", f"Game {g['game_id']} ({g['season']}) has the same Team ID "
                                  f"({g['home_id']}) as both home and away.")
    for r in pgs_rows:
        if r["team_id"] is not None and r["team_id"] == r["opp_team_id"]:
            f.error("self_play", f"PlayerGameStats row for Player {r['player_id']}, Game {r['game_id']} "
                                  f"has Team ID == Opponent Team ID ({r['team_id']}).")
 
 
def check_duplicate_pgs_rows(pgs_rows, f):
    """A player with two stat lines for the same game -- never legitimate,
    and exactly what a re-run/double-backfill without proper dedup would
    produce."""
    seen = Counter((r["player_id"], r["game_id"]) for r in pgs_rows if r["game_id"])
    dupes = [(pid, gid, n) for (pid, gid), n in seen.items() if n > 1]
    if dupes:
        sample = ", ".join(f"player {pid} game {gid} ({n}x)" for pid, gid, n in dupes[:10])
        more = f" (+{len(dupes) - 10} more)" if len(dupes) > 10 else ""
        f.error("duplicate_pgs_row", f"{len(dupes)} (player, game) pairs have more than one "
                                      f"PlayerGameStats row: {sample}{more}")
 
 
def check_stat_line_sanity(pgs_rows, f):
    bad_shooting = 0
    outliers = 0
    for r in pgs_rows:
        fgm, fga = r["fgm"] or 0, r["fga"] or 0
        fg3m, fg3a = r["fg3m"] or 0, r["fg3a"] or 0
        ftm, fta = r["ftm"] or 0, r["fta"] or 0
        if fgm > fga or fg3m > fg3a or ftm > fta:
            bad_shooting += 1
        minutes, points, rebounds = r["minutes"] or 0, r["points"] or 0, r["rebounds"] or 0
        if minutes > MAX_PLAUSIBLE_MINUTES or points > MAX_PLAUSIBLE_POINTS or rebounds > MAX_PLAUSIBLE_REBOUNDS:
            outliers += 1
    if bad_shooting:
        f.error("impossible_shooting", f"{bad_shooting} PlayerGameStats rows have makes > attempts "
                                        f"(FG, 3PT, or FT) -- physically impossible, likely a parsing bug.")
    if outliers:
        f.warn("stat_outlier", f"{outliers} PlayerGameStats rows exceed plausible single-game bounds "
                                f"(> {MAX_PLAUSIBLE_MINUTES} min, > {MAX_PLAUSIBLE_POINTS} pts, or "
                                f"> {MAX_PLAUSIBLE_REBOUNDS} reb) -- rare real games do happen, worth "
                                f"a manual glance rather than an automatic red flag.")
 
 
def run_all_checks(path, max_roster, min_roster):
    teams, games, players, pgs_rows = load_sheets(path)
    print(f"  Teams: {len(teams)}, Games: {len(games)}, Players: {len(players)}, "
          f"PlayerGameStats: {len(pgs_rows)}\n")
 
    f = Findings()
    check_duplicate_espn_ids(teams, f)
    check_missing_espn_ids(teams, f)
    check_roster_sizes(teams, players, f, max_roster, min_roster)
    check_orphaned_team_refs(teams, games, players, pgs_rows, f)
    check_self_play(games, pgs_rows, f)
    check_duplicate_pgs_rows(pgs_rows, f)
    check_stat_line_sanity(pgs_rows, f)
    return f
 
 
def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--path", default="WomensSummitTPE.xlsx")
    parser.add_argument("--max-roster", type=int, default=DEFAULT_MAX_ROSTER)
    parser.add_argument("--min-roster", type=int, default=DEFAULT_MIN_ROSTER)
    args = parser.parse_args()
 
    f = run_all_checks(args.path, args.max_roster, args.min_roster)
 
    if not f.items:
        print("All checks passed clean. No errors or warnings.")
        return 0
 
    if f.errors:
        print(f"=== {len(f.errors)} ERROR(S) -- these indicate real corruption, fix before using this data ===")
        for _, category, message in f.errors:
            print(f"  [{category}] {message}\n")
    if f.warnings:
        print(f"=== {len(f.warnings)} WARNING(S) -- worth a look, not necessarily broken ===")
        for _, category, message in f.warnings:
            print(f"  [{category}] {message}\n")
 
    print(f"Summary: {len(f.errors)} error(s), {len(f.warnings)} warning(s).")
    return 1 if f.errors else 0
 
 
if __name__ == "__main__":
    sys.exit(main())