"""Fills SeasonRankings and TeamRatings on WomensSummitTPE.xlsx from real
box-score data, adapting the summit_tpe/ prototype's PDI engine and ratings
network (see project README) to the actual scraped columns.

TransferProjection is intentionally NOT computed here. It turned out to be
the wrong shape for what it's actually for: a coach picks one player and
tries her against many different target schools with an adjustable minutes
value, which is a live per-request calculation, not something you can
precompute into fixed rows. That's now build_cache.py + projection.py +
api.py -- delete the TransferProjection sheet, this script only touches
SeasonRankings and TeamRatings.

REAL-DATA FINDING (as of this run): the Teams and Players sheets currently
hold D1 only -- 362/362 Teams rows and 8,718/8,718 Players rows are
Division="D1". "D2" only shows up as an Opponent Level label on ~2,477
PlayerGameStats rows (early-season buy games against teams that aren't in
the Teams sheet at all). SeasonRankings' "Projected Level" is just each
player's real Division right now -- there's no sub-D1 pool to flag
outliers against yet (see HoopScore.docx's coverage model for what this
should do once D2 data exists). Division is looked up live throughout, so
this degrades gracefully once D2 rows are added.

Adaptations from the synthetic prototype (documented per-topic below):

1. Position weights (summit_tpe/pdi.py) were defined for 5 buckets
   (PG/SG/WING/FORWARD/CENTER) with separate oreb/dreb weights. Real
   PlayerSeasons position data is coarse (mostly G/F/C, a few PG/SG/SF/PF/
   ATH) and PlayerGameStats has one combined "Rebound" column, no oreb/dreb
   split. Collapsed to 3 buckets (GUARD/FORWARD/CENTER), each stat weight
   the average of the two 5-bucket dicts it absorbs, "reb" weight the
   average of that bucket's oreb/dreb weights. G/PG/SG -> GUARD;
   F/SF/PF/ATH -> FORWARD; C -> CENTER; missing/unrecognized position ->
   the games-weighted average of all three (still counts, not thrown out).

2. Opponent strength (summit_tpe/pdi.py's opponent_strength_factor) assumed
   an Elo-like rating centered at 1600 with an ~800-point spread. Summit
   Rat here is an iterative net-efficiency rating (Off - Def, points-per-
   game scale, e.g. Akron -12.0, Alabama +26.3) computed fresh per season
   with the SAME iterative method as scrapers/compute_summit_ratings.py, so
   the reference/scale are calibrated per-season from that season's actual
   Rat distribution (mean / 2*stdev) instead of hardcoded Elo constants.

3. Close-game weighting (summit_tpe/weighting.py) is unchanged (full weight
   through a 12-point margin, decaying to a 0.25 floor by 30), just sourced
   from the real Games sheet's Margin column (already signed Home - Away;
   flipped to the player's own team's perspective via Game ID).

4. Hoop Score scaling (summit_tpe/pdi.py's scale_to_pdi, 30-99 non-linear
   z-score+convexity curve) is applied twice at different granularities to
   match the SeasonRankings columns: once across all qualifying players'
   SEASON-level raw scores (-> Average Hoop Score, used for ranking), and
   once across every individual qualifying GAME's raw value league-wide
   (-> that same per-game scale used to report each player's Highest/
   Lowest Hoop Score and the Standard Deviation of her games).

5. BUG FOUND AND FIXED after initial delivery: a player with 7 games
   totaling 21 minutes (several 1-2 minute cameos) ranked #1 in the
   country. Root cause was per-game per-40 extrapolation with no floor --
   a 1-minute, 5-point cameo scales to a "200 PPG pace" game, and the
   existing reliability weight (0.3 floor) wasn't enough to fully suppress
   that in the season-weighted average. Fixed two ways: the per-game
   extrapolation now floors minutes at PER40_MINUTES_FLOOR (no single game
   can be stretched past a 4x multiplier), and eligibility now requires
   MIN_TOTAL_MINUTES_FOR_RANKING of real season playing time, not just
   MIN_GAMES_FOR_RANKING games appeared in.

Usage:
    python compute_derived_sheets.py --path WomensSummitTPE.xlsx [--dry-run]
"""

import argparse
import datetime
import statistics
import sys
from collections import defaultdict

import openpyxl

MIN_GAMES_FOR_RANKING = 5
# A games-count floor alone isn't enough: a player can clear 5 games while
# barely playing in any of them (see bug note #5 above). Require real
# season-length playing time, not just games appeared in.
MIN_TOTAL_MINUTES_FOR_RANKING = 100
# Per-40 extrapolation blows up for very short stints (1 minute, 5 points
# -> a "200 PPG pace" game). Floor the minutes used in the 40/minutes scale
# factor so no single game extrapolates further than 4x (equivalent to
# treating anything under 10 minutes as if it were exactly 10).
PER40_MINUTES_FLOOR = 10


# ---------- shared helpers (mirrors scrapers/xlsx_io.py conventions) ----------

def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping


def clamp(x, lo, hi):
    return max(lo, min(hi, x))


# ---------- position weight buckets (collapsed from summit_tpe/pdi.py) ----------

_ORIGINAL = {
    "PG":      dict(pts=1.0, fgm=0.4, fga=-0.7, ftmiss=-0.4, oreb=0.4, dreb=0.30, ast=1.10, stl=1.10, blk=0.40, tov=-1.10, pf=-0.30),
    "SG":      dict(pts=1.0, fgm=0.4, fga=-0.7, ftmiss=-0.4, oreb=0.5, dreb=0.35, ast=0.80, stl=0.90, blk=0.40, tov=-1.00, pf=-0.30),
    "WING":    dict(pts=1.0, fgm=0.4, fga=-0.7, ftmiss=-0.4, oreb=0.6, dreb=0.45, ast=0.70, stl=0.85, blk=0.55, tov=-0.90, pf=-0.30),
    "FORWARD": dict(pts=1.0, fgm=0.4, fga=-0.7, ftmiss=-0.4, oreb=0.8, dreb=0.55, ast=0.50, stl=0.70, blk=0.75, tov=-0.80, pf=-0.35),
    "CENTER":  dict(pts=1.0, fgm=0.4, fga=-0.7, ftmiss=-0.4, oreb=1.0, dreb=0.65, ast=0.35, stl=0.55, blk=1.00, tov=-0.70, pf=-0.40),
}


def _collapse(*keys):
    dicts = [_ORIGINAL[k] for k in keys]
    out = {}
    for stat in dicts[0]:
        if stat in ("oreb", "dreb"):
            continue
        out[stat] = sum(d[stat] for d in dicts) / len(dicts)
    out["reb"] = sum((d["oreb"] + d["dreb"]) / 2.0 for d in dicts) / len(dicts)
    return out


BUCKET_WEIGHTS = {
    "GUARD": _collapse("PG", "SG"),
    "FORWARD": _collapse("WING", "FORWARD"),
    "CENTER": _collapse("CENTER"),
}
BUCKET_WEIGHTS["ALL"] = {
    stat: sum(BUCKET_WEIGHTS[b][stat] for b in ("GUARD", "FORWARD", "CENTER")) / 3.0
    for stat in BUCKET_WEIGHTS["GUARD"]
}

POSITION_TO_BUCKET = {
    "G": "GUARD", "PG": "GUARD", "SG": "GUARD",
    "F": "FORWARD", "SF": "FORWARD", "PF": "FORWARD", "ATH": "FORWARD",
    "C": "CENTER",
}

EXPERIENCE_MULT = {"FR": 1.03, "SO": 1.015, "JR": 1.0, "SR": 1.0, "GR": 0.995}


def composite_game_score(bucket, points, fgm, fga, ftmiss, reb, ast, stl, blk, tov, pf):
    w = BUCKET_WEIGHTS[bucket]
    return (
        points * w["pts"] + fgm * w["fgm"] + fga * w["fga"] + ftmiss * w["ftmiss"]
        + reb * w["reb"] + ast * w["ast"] + stl * w["stl"] + blk * w["blk"]
        + tov * w["tov"] + pf * w["pf"]
    )


def close_game_weight(margin):
    m = abs(margin)
    if m <= 12:
        return 1.0
    return clamp(1.0 - (m - 12) / 18.0, 0.25, 1.0)


def scale_to_hoopscore(raw_by_key, mean=None, std=None):
    """Returns (displayed, unclamped, mean, std). `displayed` is the 30-99
    clamped/rounded Hoop Score shown to users. `unclamped` keeps full
    precision above 99 / below 30 -- several distinct elite performances can
    legitimately round to the same displayed 99.0 (that convex top-end
    compression is the intended design, see summit_tpe/pdi.py), but ranking
    output should still order those players correctly rather than falling
    back to arbitrary tie order, so callers sort by `unclamped`, not
    `displayed`.
    """
    values = list(raw_by_key.values())
    if mean is None:
        mean = statistics.mean(values) if values else 0.0
    if std is None:
        std = statistics.pstdev(values) if len(values) > 1 else 1.0
        std = std or 1.0
    displayed = {}
    unclamped = {}
    for key, raw in raw_by_key.items():
        z = (raw - mean) / std
        hs = 60.0 + 12.0 * z + 3.0 * max(0.0, z) ** 2
        unclamped[key] = hs
        displayed[key] = round(clamp(hs, 30.0, 99.0), 1)
    return displayed, unclamped, mean, std


# ---------- team strength: iterative Off/Def/Rat (mirrors compute_summit_ratings.py) ----------

TOL = 1e-4
MAX_ITER = 200


def iterative_off_def(games):
    """games: [(home_id, away_id, home_score, away_score), ...]. Returns (off, def_) dicts."""
    per_team_games = defaultdict(list)
    for h, a, hs, as_ in games:
        per_team_games[h].append((hs, as_, a))
        per_team_games[a].append((as_, hs, h))

    active = list(per_team_games.keys())
    if not active:
        return {}, {}

    off = {t: sum(s for s, _, _ in per_team_games[t]) / len(per_team_games[t]) for t in active}
    def_ = {t: sum(al for _, al, _ in per_team_games[t]) / len(per_team_games[t]) for t in active}

    for _ in range(MAX_ITER):
        league_off = sum(off.values()) / len(off)
        league_def = sum(def_.values()) / len(def_)
        new_off, new_def = {}, {}
        for t in active:
            gms = per_team_games[t]
            new_off[t] = sum((scored - def_[opp]) for scored, _, opp in gms) / len(gms) + league_def
            new_def[t] = sum((allowed - off[opp]) for _, allowed, opp in gms) / len(gms) + league_off
        max_change = max(
            max(abs(new_off[t] - off[t]) for t in active),
            max(abs(new_def[t] - def_[t]) for t in active),
        )
        off, def_ = new_off, new_def
        if max_change < TOL:
            break
    return off, def_


def compute_sos(games, rat):
    opponents = defaultdict(list)
    for h, a, _, _ in games:
        if h in rat and a in rat:
            opponents[h].append(rat[a])
            opponents[a].append(rat[h])
    return {t: sum(vals) / len(vals) for t, vals in opponents.items() if vals}


def compute_avg_opponent_rating(games, rat):
    """Unique-opponent-weighted (not per-game-weighted like SoS)."""
    opponents = defaultdict(set)
    for h, a, _, _ in games:
        if h in rat and a in rat:
            opponents[h].add(a)
            opponents[a].add(h)
    return {t: sum(rat[o] for o in opps) / len(opps) for t, opps in opponents.items() if opps}


def load():
    print("Opening workbook read-only for extraction ...")
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)

    teams_ws = wb["Teams"]
    th = header_map(teams_ws)
    team_name, team_div = {}, {}
    for row in teams_ws.iter_rows(min_row=2, values_only=True):
        tid = row[th["Team ID"] - 1]
        if tid is None:
            continue
        team_name[tid] = row[th["Team"] - 1]
        team_div[tid] = row[th["Division"] - 1]
    print(f"  Teams: {len(team_name)}")

    games_ws = wb["Games"]
    gh = header_map(games_ws)
    games_by_season = defaultdict(list)   # season -> [(home_id, away_id, hs, as_)]
    game_lookup = {}                      # game_id -> (home_id, away_id, margin, season)
    for row in games_ws.iter_rows(min_row=2, values_only=True):
        gid = row[gh["Game ID"] - 1]
        if gid is None:
            continue
        season = row[gh["Season"] - 1]
        home_id = row[gh["Home Team ID"] - 1]
        away_id = row[gh["Away Team ID"] - 1]
        hs = row[gh["Home Score"] - 1]
        as_ = row[gh["Away Score"] - 1]
        margin = row[gh["Margin"] - 1]
        if home_id is not None and away_id is not None and hs is not None and as_ is not None:
            games_by_season[season].append((home_id, away_id, hs, as_))
        game_lookup[gid] = (home_id, away_id, margin, season)
    print(f"  Games: {sum(len(v) for v in games_by_season.values())} scored, by season: "
          f"{[(s, len(v)) for s, v in games_by_season.items()]}")

    ps_ws = wb["PlayerSeasons"]
    ph = header_map(ps_ws)
    player_season = {}   # (player_id, season) -> dict(team_id, division, position, class, games_played)
    for row in ps_ws.iter_rows(min_row=2, values_only=True):
        pid = row[ph["Player ID"] - 1]
        season = row[ph["Season"] - 1]
        if pid is None or season is None:
            continue
        player_season[(pid, season)] = dict(
            team_id=row[ph["Team ID"] - 1],
            division=row[ph["Division"] - 1],
            position=row[ph["Position"] - 1],
            class_year=row[ph["Class"] - 1],
            games_played=row[ph["Games Played"] - 1],
        )
    print(f"  PlayerSeasons: {len(player_season)}")

    players_ws = wb["Players"]
    plh = header_map(players_ws)
    player_name = {}
    for row in players_ws.iter_rows(min_row=2, values_only=True):
        pid = row[plh["Player ID"] - 1]
        if pid is None:
            continue
        first = row[plh["First Name"] - 1] or ""
        last = row[plh["Last Name"] - 1] or ""
        player_name[pid] = f"{first} {last}".strip()
    print(f"  Players: {len(player_name)}")

    pgs_ws = wb["PlayerGameStats"]
    gsh = header_map(pgs_ws)
    game_rows_by_season = defaultdict(list)
    for row in pgs_ws.iter_rows(min_row=2, values_only=True):
        pid = row[gsh["Player ID"] - 1]
        if pid is None:
            continue
        season = row[gsh["Season"] - 1]
        rec = dict(
            player_id=pid,
            team_id=row[gsh["Team ID"] - 1],
            opp_team_id=row[gsh["Opponent Team ID"] - 1],
            minutes=row[gsh["Min"] - 1] or 0,
            fgm=row[gsh["FG Made"] - 1] or 0,
            fga=row[gsh["FG Attempt"] - 1] or 0,
            ftm=row[gsh["FT M"] - 1] or 0,
            fta=row[gsh["FT A"] - 1] or 0,
            reb=row[gsh["Rebound"] - 1] or 0,
            pf=row[gsh["Foul"] - 1] or 0,
            ast=row[gsh["Ast"] - 1] or 0,
            tov=row[gsh["To"] - 1] or 0,
            blk=row[gsh["Blk"] - 1] or 0,
            stl=row[gsh["Stl"] - 1] or 0,
            points=row[gsh["Points"] - 1] or 0,
            game_id=row[gsh["Game ID"] - 1],
        )
        game_rows_by_season[season].append(rec)
    total = sum(len(v) for v in game_rows_by_season.values())
    print(f"  PlayerGameStats: {total}, by season: {[(s, len(v)) for s, v in game_rows_by_season.items()]}")

    wb.close()
    return dict(
        team_name=team_name, team_div=team_div,
        games_by_season=games_by_season, game_lookup=game_lookup,
        player_season=player_season, player_name=player_name,
        game_rows_by_season=game_rows_by_season,
    )


def pick_latest_season(games_by_season):
    # Same convention as compute_summit_ratings.py: most games = most complete season.
    return max(games_by_season, key=lambda s: len(games_by_season[s]))


def team_ratings_for_season(games_by_season, season):
    off, def_ = iterative_off_def(games_by_season.get(season, []))
    rat = {t: off[t] - def_[t] for t in off}
    sos = compute_sos(games_by_season.get(season, []), rat)
    avg_opp = compute_avg_opponent_rating(games_by_season.get(season, []), rat)
    return off, def_, rat, sos, avg_opp


def home_away_split(games, off_c, def_c):
    """Home/Away Rating: NOT a fresh iterative solve (a ~15-game home-only or
    away-only slate per team is too sparse/disconnected a network to solve on
    its own). Instead: single-pass, opponent-strength-adjusted point
    differential, using the ALREADY-CONVERGED season-wide off_c/def_c as the
    fixed opponent reference.
    """
    home_pts = defaultdict(list)
    away_pts = defaultdict(list)
    for h, a, hs, as_ in games:
        home_pts[h].append((hs, as_, a))
        away_pts[a].append((as_, hs, h))

    league_off = sum(off_c.values()) / len(off_c)
    league_def = sum(def_c.values()) / len(def_c)

    def rate(pts_by_team):
        out = {}
        for t, gms in pts_by_team.items():
            usable = [(s, al, opp) for s, al, opp in gms if opp in off_c and opp in def_c]
            if not usable:
                continue
            off_t = sum(s - def_c[opp] for s, al, opp in usable) / len(usable) + league_def
            def_t = sum(al - off_c[opp] for s, al, opp in usable) / len(usable) + league_off
            out[t] = off_t - def_t
        return out

    return rate(home_pts), rate(away_pts)


def opponent_strength_factor(opp_rating, reference, scale):
    scale = scale or 1.0
    return clamp(1.0 + (opp_rating - reference) / scale, 0.55, 1.6)


def main():
    global PATH
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", required=True)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--out-json", default=None, help="Write computed rows to JSON for inspection instead of/in addition to xlsx")
    args = parser.parse_args()
    PATH = args.path

    data = load()
    season = pick_latest_season(data["games_by_season"])
    others = sorted(s for s in data["games_by_season"] if s != season)
    prior_season = others[-1] if others else None

    print(f"\nSeason (current/most recent): {season}")
    print(f"Season (prior, for Initial Rating): {prior_season}")

    # ---- Team ratings, current + prior season ----
    off_c, def_c, rat_c, sos_c, avgopp_c = team_ratings_for_season(data["games_by_season"], season)
    if prior_season:
        _, _, rat_p, _, _ = team_ratings_for_season(data["games_by_season"], prior_season)
    else:
        rat_p = {}
    home_rat, away_rat = home_away_split(data["games_by_season"][season], off_c, def_c)

    rat_values = list(rat_c.values())
    league_mean_rat = statistics.mean(rat_values)
    league_std_rat = statistics.pstdev(rat_values) or 1.0
    print(f"Current-season Rat: n={len(rat_c)} mean={league_mean_rat:.2f} std={league_std_rat:.2f}")

    team_ratings_rows = []
    now_str = datetime.datetime.now().isoformat(timespec="seconds")
    for tid, name in data["team_name"].items():
        if tid not in rat_c:
            continue  # isolated team, no games this season -- leave out rather than guess
        team_ratings_rows.append(dict(
            team=name,
            division=data["team_div"].get(tid),
            initial_rating=round(rat_p[tid], 2) if tid in rat_p else None,
            current_rating=round(rat_c[tid], 2),
            sos=round(sos_c.get(tid), 2) if tid in sos_c else None,
            avg_opponent_rating=round(avgopp_c.get(tid), 2) if tid in avgopp_c else None,
            home_rating=round(home_rat.get(tid), 2) if tid in home_rat else None,
            away_rating=round(away_rat.get(tid), 2) if tid in away_rat else None,
            last_updated=now_str,
        ))
    print(f"TeamRatings rows: {len(team_ratings_rows)}")

    # ---- Hoop Score / SeasonRankings ----
    rows = data["game_rows_by_season"].get(season, [])
    game_lookup = data["game_lookup"]

    per_player_games = defaultdict(list)
    per_player_meta = {}

    for r in rows:
        pid = r["player_id"]
        ps = data["player_season"].get((pid, season))
        if ps is None:
            continue
        position = ps["position"]
        bucket = POSITION_TO_BUCKET.get(position, "ALL")
        minutes = max(r["minutes"], 1)  # actual minutes played, used for totals/weighting below
        scale = 40.0 / max(minutes, PER40_MINUTES_FLOOR)  # FLOORED scale, per-game extrapolation only
        ftmiss = r["fta"] - r["ftm"]

        raw = composite_game_score(bucket, r["points"], r["fgm"], r["fga"], ftmiss,
                                    r["reb"], r["ast"], r["stl"], r["blk"], r["tov"], r["pf"])
        composite_per40 = raw * scale

        opp_rat = rat_c.get(r["opp_team_id"], league_mean_rat)
        opp_factor = opponent_strength_factor(opp_rat, league_mean_rat, 2 * league_std_rat)
        adjusted_value = composite_per40 * opp_factor

        margin = None
        gl = game_lookup.get(r["game_id"])
        if gl:
            home_id, away_id, gmargin, _ = gl
            if gmargin is not None:
                if r["team_id"] == home_id:
                    margin = gmargin
                elif r["team_id"] == away_id:
                    margin = -gmargin
        cgw = close_game_weight(margin) if margin is not None else 1.0
        reliability = clamp(minutes / 15.0, 0.3, 1.0)
        weight = cgw * reliability

        per_player_games[pid].append(dict(value=adjusted_value, weight=weight, minutes=minutes))
        if pid not in per_player_meta:
            per_player_meta[pid] = dict(
                team_id=ps["team_id"], division=ps["division"], position=position,
                class_year=ps["class_year"],
            )

    print(f"Players with >=1 qualifying game in {season}: {len(per_player_games)}")

    # per-game population scaling (for Highest/Lowest/StdDev)
    all_game_values = {}
    i = 0
    for pid, glist in per_player_games.items():
        for g in glist:
            all_game_values[(pid, i)] = g["value"]
            i += 1
    per_game_hs, _per_game_hs_raw, gmean, gstd = scale_to_hoopscore(all_game_values)

    # season-aggregate scaling (for Average Hoop Score / ranking)
    season_raw = {}
    for pid, glist in per_player_games.items():
        wsum = sum(g["weight"] for g in glist)
        if wsum <= 0:
            sraw = sum(g["value"] for g in glist) / len(glist)
        else:
            sraw = sum(g["value"] * g["weight"] for g in glist) / wsum
        mult = EXPERIENCE_MULT.get(per_player_meta[pid]["class_year"], 1.0)
        season_raw[pid] = sraw * mult
    season_hs, season_hs_raw, smean, sstd = scale_to_hoopscore(season_raw)

    season_ranking_rows = []
    skipped_thin_minutes = 0
    idx = 0
    for pid, glist in per_player_games.items():
        n_games = len(glist)
        game_scores = [per_game_hs[(pid, idx + k)] for k in range(n_games)]
        idx += n_games
        if n_games < MIN_GAMES_FOR_RANKING:
            continue
        total_minutes = sum(g["minutes"] for g in glist)
        if total_minutes < MIN_TOTAL_MINUTES_FOR_RANKING:
            skipped_thin_minutes += 1
            continue
        meta = per_player_meta[pid]
        season_ranking_rows.append(dict(
            player_id=pid,
            player=data["player_name"].get(pid, f"Player {pid}"),
            team=data["team_name"].get(meta["team_id"], "?"),
            division=meta["division"],
            position=meta["position"],
            games=n_games,
            average_hoop_score=season_hs[pid],
            _rank_key=season_hs_raw[pid],  # unclamped -- see scale_to_hoopscore docstring
            highest_hoop_score=max(game_scores),
            lowest_hoop_score=min(game_scores),
            stdev=round(statistics.pstdev(game_scores), 1) if len(game_scores) > 1 else 0.0,
            projected_level=meta["division"],
        ))

    season_ranking_rows.sort(key=lambda r: -r["_rank_key"])
    for i, r in enumerate(season_ranking_rows, start=1):
        r["overall_rank"] = i
        del r["_rank_key"]
    by_division_rank = defaultdict(int)
    for r in season_ranking_rows:
        by_division_rank[r["division"]] += 1
        r["division_rank"] = by_division_rank[r["division"]]

    print(f"SeasonRankings rows (>= {MIN_GAMES_FOR_RANKING} games AND >= {MIN_TOTAL_MINUTES_FOR_RANKING} total minutes): "
          f"{len(season_ranking_rows)} ({skipped_thin_minutes} more had enough games but too few total minutes, excluded)")

    print("\n--- anomaly scan (full row sets, not just the JSON sample) ---")
    for label, out_rows, numeric_fields in [
        ("TeamRatings", team_ratings_rows,
         ["initial_rating", "current_rating", "sos", "avg_opponent_rating", "home_rating", "away_rating"]),
        ("SeasonRankings", season_ranking_rows,
         ["overall_rank", "division_rank", "games", "average_hoop_score", "highest_hoop_score",
          "lowest_hoop_score", "stdev"]),
    ]:
        print(f"{label}: {len(out_rows)} rows")
        for field in numeric_fields:
            vals = [r[field] for r in out_rows]
            n_none = sum(1 for v in vals if v is None)
            present = [v for v in vals if v is not None]
            if present:
                print(f"  {field}: none={n_none} min={min(present):.2f} max={max(present):.2f} "
                      f"mean={statistics.mean(present):.2f}")
            else:
                print(f"  {field}: none={n_none} (all missing)")
        if label == "SeasonRankings":
            bad = [r for r in out_rows if r["lowest_hoop_score"] > r["highest_hoop_score"]]
            print(f"  rows where lowest > highest (should be 0): {len(bad)}")

    if args.out_json:
        import json
        with open(args.out_json, "w") as f:
            json.dump(dict(
                season=season, prior_season=prior_season,
                team_ratings=team_ratings_rows,
                season_rankings=season_ranking_rows[:50],
                counts=dict(
                    team_ratings=len(team_ratings_rows),
                    season_rankings=len(season_ranking_rows),
                ),
            ), f, indent=2, default=str)
        print(f"Wrote sample JSON to {args.out_json}")

    if args.dry_run:
        print("\n--dry-run: not writing to workbook.")
        return

    write_to_workbook(team_ratings_rows, season_ranking_rows)


def write_to_workbook(team_ratings_rows, season_ranking_rows):
    print(f"\nOpening {PATH} for write (this can take several minutes on the real file) ...")
    t0 = datetime.datetime.now()
    wb = openpyxl.load_workbook(PATH)
    print(f"  loaded in {(datetime.datetime.now()-t0).total_seconds():.0f}s")

    def fill(sheet_name, rows, field_order):
        if sheet_name not in wb.sheetnames:
            print(f"  [!] Sheet '{sheet_name}' not found, skipping (already removed?)")
            return
        ws = wb[sheet_name]
        hmap = header_map(ws)
        for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, 1)):
            for cell in row:
                cell.value = None
        for i, r in enumerate(rows, start=2):
            for field, header in field_order:
                col = hmap.get(header)
                if col:
                    ws.cell(row=i, column=col, value=r.get(field))

    fill("TeamRatings", team_ratings_rows, [
        ("team", "Team"), ("division", "Division"), ("initial_rating", "Initial Rating"),
        ("current_rating", "Current Rating"), ("sos", "Strength of Schedule"),
        ("avg_opponent_rating", "Average Opponent Rating"), ("home_rating", "Home Rating"),
        ("away_rating", "Away Rating"), ("last_updated", "Last Updated"),
    ])
    fill("SeasonRankings", season_ranking_rows, [
        ("overall_rank", "Overall Rank"), ("division_rank", "Division Rank"), ("player", "Player"),
        ("team", "Team"), ("division", "Division"), ("position", "Position"), ("games", "Games"),
        ("average_hoop_score", "Average Hoop Score"), ("highest_hoop_score", "Highest Hoop Score"),
        ("lowest_hoop_score", "Lowest Hoop Score"), ("stdev", "Standard Deviation"),
        ("projected_level", "Projected Level"),
    ])

    print("Saving (this is the slow part on the real file) ...")
    t1 = datetime.datetime.now()
    tmp_path = f"{PATH}.tmp_saving"
    wb.save(tmp_path)
    import os
    os.replace(tmp_path, PATH)
    print(f"  saved in {(datetime.datetime.now()-t1).total_seconds():.0f}s")
    print("Done.")


if __name__ == "__main__":
    main()
