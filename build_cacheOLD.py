"""Builds summit_tpe_cache.sqlite from WomensSummitTPE.xlsx: one row per
current-season team (with rating + tier) and one row per current-season
player (season stat profile: PPG/RPG/APG/BPG/SPG/TOPG/TS%/FG%/Hoop Score/
per-40 rates for points, rebounds, assists, blocks, steals, turnovers).

This is the STATIC half of the transfer calculator -- a player's own
season stats don't change based on what team you're evaluating her
against, so they're computed once here and cached, instead of being
recomputed on every API request. The FLUID half (target team, minutes
override) is computed live per-request by projection.py, reading from this
cache.

Re-run this once per season (or whenever the underlying workbook's box
scores update) to refresh the cache; the API always reads whatever's
currently in summit_tpe_cache.sqlite.

Usage:
    python build_cache.py --path WomensSummitTPE.xlsx --out summit_tpe_cache.sqlite
"""

import argparse
import sqlite3
import statistics
from collections import defaultdict

import openpyxl

from summit_calc import (
    BUCKET_WEIGHTS, EXPERIENCE_MULT, MIN_GAMES_FOR_PROFILE, MIN_TOTAL_MINUTES_FOR_PROFILE,
    POSITION_TO_BUCKET, classify_tier, close_game_weight, compute_sos, composite_game_score,
    iterative_off_def, opponent_strength_factor, per40_scale, scale_to_hoopscore,
)


def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping


def load(path):
    print(f"Opening {path} read-only for extraction ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    teams_ws = wb["Teams"]
    th = header_map(teams_ws)
    teams = {}   # team_id -> dict(name, division, conference)
    for row in teams_ws.iter_rows(min_row=2, values_only=True):
        tid = row[th["Team ID"] - 1]
        if tid is None:
            continue
        teams[tid] = dict(
            name=row[th["Team"] - 1],
            division=row[th["Division"] - 1],
            conference=row[th["Conference"] - 1],
        )
    print(f"  Teams: {len(teams)}")

    games_ws = wb["Games"]
    gh = header_map(games_ws)
    games_by_season = defaultdict(list)
    game_lookup = {}
    for row in games_ws.iter_rows(min_row=2, values_only=True):
        gid = row[gh["Game ID"] - 1]
        if gid is None:
            continue
        season = row[gh["Season"] - 1]
        home_id, away_id = row[gh["Home Team ID"] - 1], row[gh["Away Team ID"] - 1]
        hs, as_ = row[gh["Home Score"] - 1], row[gh["Away Score"] - 1]
        margin = row[gh["Margin"] - 1]
        if home_id is not None and away_id is not None and hs is not None and as_ is not None:
            games_by_season[season].append((home_id, away_id, hs, as_))
        game_lookup[gid] = (home_id, away_id, margin, season)
    print(f"  Games: by season {[(s, len(v)) for s, v in games_by_season.items()]}")

    ps_ws = wb["PlayerSeasons"]
    ph = header_map(ps_ws)
    player_season = {}
    for row in ps_ws.iter_rows(min_row=2, values_only=True):
        pid, season = row[ph["Player ID"] - 1], row[ph["Season"] - 1]
        if pid is None or season is None:
            continue
        player_season[(pid, season)] = dict(
            team_id=row[ph["Team ID"] - 1], division=row[ph["Division"] - 1],
            position=row[ph["Position"] - 1], class_year=row[ph["Class"] - 1],
        )
    print(f"  PlayerSeasons: {len(player_season)}")

    players_ws = wb["Players"]
    plh = header_map(players_ws)
    player_name = {}
    for row in players_ws.iter_rows(min_row=2, values_only=True):
        pid = row[plh["Player ID"] - 1]
        if pid is None:
            continue
        player_name[pid] = f"{row[plh['First Name'] - 1] or ''} {row[plh['Last Name'] - 1] or ''}".strip()
    print(f"  Players: {len(player_name)}")

    pgs_ws = wb["PlayerGameStats"]
    gsh = header_map(pgs_ws)
    game_rows_by_season = defaultdict(list)
    for row in pgs_ws.iter_rows(min_row=2, values_only=True):
        pid = row[gsh["Player ID"] - 1]
        if pid is None:
            continue
        season = row[gsh["Season"] - 1]
        game_rows_by_season[season].append(dict(
            player_id=pid, team_id=row[gsh["Team ID"] - 1], opp_team_id=row[gsh["Opponent Team ID"] - 1],
            minutes=row[gsh["Min"] - 1] or 0, fgm=row[gsh["FG Made"] - 1] or 0, fga=row[gsh["FG Attempt"] - 1] or 0,
            ftm=row[gsh["FT M"] - 1] or 0, fta=row[gsh["FT A"] - 1] or 0, reb=row[gsh["Rebound"] - 1] or 0,
            pf=row[gsh["Foul"] - 1] or 0, ast=row[gsh["Ast"] - 1] or 0, tov=row[gsh["To"] - 1] or 0,
            blk=row[gsh["Blk"] - 1] or 0, stl=row[gsh["Stl"] - 1] or 0, points=row[gsh["Points"] - 1] or 0,
            game_id=row[gsh["Game ID"] - 1],
        ))
        # blk/stl/tov/fgm/fga above already existed for the Hoop Score
        # composite; block/steal/turnover season averages (added below)
        # reuse these same per-game values rather than re-reading the sheet.
    print(f"  PlayerGameStats: by season {[(s, len(v)) for s, v in game_rows_by_season.items()]}")

    wb.close()
    return dict(teams=teams, games_by_season=games_by_season, game_lookup=game_lookup,
                player_season=player_season, player_name=player_name, game_rows_by_season=game_rows_by_season)


def pick_latest_season(games_by_season):
    return max(games_by_season, key=lambda s: len(games_by_season[s]))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", required=True)
    parser.add_argument("--out", default="summit_tpe_cache.sqlite")
    args = parser.parse_args()

    data = load(args.path)
    season = pick_latest_season(data["games_by_season"])
    print(f"\nSeason: {season}")

    off, def_ = iterative_off_def(data["games_by_season"][season])
    rat = {t: off[t] - def_[t] for t in off}
    sos = compute_sos(data["games_by_season"][season], rat)
    rat_values = list(rat.values())
    league_mean_rat = statistics.mean(rat_values)
    league_std_rat = statistics.pstdev(rat_values) or 1.0
    print(f"Rat: n={len(rat)} mean={league_mean_rat:.2f} std={league_std_rat:.2f}")

    # ---- player season profiles (same math as compute_derived_sheets.py) ----
    rows = data["game_rows_by_season"].get(season, [])
    game_lookup = data["game_lookup"]
    per_player_games = defaultdict(list)
    per_player_meta = {}

    for r in rows:
        pid = r["player_id"]
        ps = data["player_season"].get((pid, season))
        if ps is None:
            continue
        position = ps["position"]
        bucket = POSITION_TO_BUCKET.get(position, "ALL")
        minutes = max(r["minutes"], 1)  # actual minutes played, used for totals/weighting below
        scale = per40_scale(minutes)     # FLOORED scale, used only for per-game extrapolation
        ftmiss = r["fta"] - r["ftm"]
        raw = composite_game_score(bucket, r["points"], r["fgm"], r["fga"], ftmiss,
                                    r["reb"], r["ast"], r["stl"], r["blk"], r["tov"], r["pf"])
        composite_per40 = raw * scale
        opp_rat = rat.get(r["opp_team_id"], league_mean_rat)
        opp_factor = opponent_strength_factor(opp_rat, league_mean_rat, 2 * league_std_rat)
        adjusted_value = composite_per40 * opp_factor

        margin = None
        gl = game_lookup.get(r["game_id"])
        if gl:
            home_id, away_id, gmargin, _ = gl
            if gmargin is not None:
                if r["team_id"] == home_id:
                    margin = gmargin
                elif r["team_id"] == away_id:
                    margin = -gmargin
        cgw = close_game_weight(margin) if margin is not None else 1.0
        reliability = max(0.3, min(1.0, minutes / 15.0))
        weight = cgw * reliability

        per_player_games[pid].append(dict(
            value=adjusted_value, weight=weight,
            minutes=minutes, points=r["points"], reb=r["reb"], ast=r["ast"], fga=r["fga"], fta=r["fta"],
            blk=r["blk"], stl=r["stl"], tov=r["tov"], fgm=r["fgm"],
        ))
        if pid not in per_player_meta:
            per_player_meta[pid] = dict(team_id=ps["team_id"], division=ps["division"],
                                         position=position, class_year=ps["class_year"])

    season_raw = {}
    for pid, glist in per_player_games.items():
        wsum = sum(g["weight"] for g in glist)
        sraw = (sum(g["value"] * g["weight"] for g in glist) / wsum) if wsum > 0 else (sum(g["value"] for g in glist) / len(glist))
        mult = EXPERIENCE_MULT.get(per_player_meta[pid]["class_year"], 1.0)
        season_raw[pid] = sraw * mult
    season_hs, season_hs_raw, _, _ = scale_to_hoopscore(season_raw)

    player_rows = []
    skipped_thin_minutes = 0
    for pid, glist in per_player_games.items():
        n_games = len(glist)
        if n_games < MIN_GAMES_FOR_PROFILE:
            continue
        meta = per_player_meta[pid]
        total_min = sum(g["minutes"] for g in glist)
        if total_min < MIN_TOTAL_MINUTES_FOR_PROFILE:
            skipped_thin_minutes += 1
            continue
        total_pts = sum(g["points"] for g in glist)
        total_reb = sum(g["reb"] for g in glist)
        total_ast = sum(g["ast"] for g in glist)
        total_fga = sum(g["fga"] for g in glist)
        total_fta = sum(g["fta"] for g in glist)
        total_blk = sum(g["blk"] for g in glist)
        total_stl = sum(g["stl"] for g in glist)
        total_tov = sum(g["tov"] for g in glist)
        total_fgm = sum(g["fgm"] for g in glist)
        true_shot_attempts = total_fga + 0.44 * total_fta
        ts_pct = (total_pts / (2 * true_shot_attempts)) if true_shot_attempts >= 15 else None
        fg_pct = (total_fgm / total_fga) if total_fga >= 15 else None
        # Per-40 rates from SEASON TOTALS (total stat / total minutes * 40),
        # not an average of each game's per-game extrapolation -- a real
        # per-40 rate should be implicitly weighted by how many minutes
        # each game actually contributed, which this does automatically; an
        # unweighted average of per-game rates lets a handful of short,
        # extrapolation-prone stints dominate regardless of weight.
        player_rows.append(dict(
            player_id=pid, name=data["player_name"].get(pid, f"Player {pid}"),
            team_id=meta["team_id"], division=meta["division"], position=meta["position"],
            class_year=meta["class_year"], season=season, games=n_games,
            avg_minutes=total_min / n_games, ppg=total_pts / n_games, rpg=total_reb / n_games,
            apg=total_ast / n_games, bpg=total_blk / n_games, spg=total_stl / n_games,
            topg=total_tov / n_games, ts_pct=ts_pct, fg_pct=fg_pct,
            per40_pts=total_pts / total_min * 40.0,
            per40_reb=total_reb / total_min * 40.0,
            per40_ast=total_ast / total_min * 40.0,
            per40_blk=total_blk / total_min * 40.0,
            per40_stl=total_stl / total_min * 40.0,
            per40_tov=total_tov / total_min * 40.0,
            hoop_score=season_hs[pid], hoop_score_raw=season_hs_raw[pid],
        ))
    print(f"Player profiles (>= {MIN_GAMES_FOR_PROFILE} games AND >= {MIN_TOTAL_MINUTES_FOR_PROFILE} total minutes): "
          f"{len(player_rows)} ({skipped_thin_minutes} more had enough games but too few total minutes, excluded)")

    team_rows = []
    for tid, info in data["teams"].items():
        if tid not in rat:
            continue
        team_rows.append(dict(
            team_id=tid, name=info["name"], division=info["division"], conference=info["conference"],
            tier=classify_tier(info["name"], info["conference"]),
            current_rating=rat[tid], sos=sos.get(tid),
        ))
    print(f"Team ratings: {len(team_rows)}")
    tier_counts = defaultdict(int)
    for t in team_rows:
        tier_counts[t["tier"]] += 1
    print(f"  by tier: {dict(tier_counts)}")

    # ---- write sqlite ----
    print(f"\nWriting {args.out} ...")
    conn = sqlite3.connect(args.out)
    conn.execute("DROP TABLE IF EXISTS teams")
    conn.execute("DROP TABLE IF EXISTS players")
    conn.execute("DROP TABLE IF EXISTS meta")
    conn.execute("CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT)")
    conn.executemany("INSERT INTO meta VALUES (?, ?)", [
        ("season", season),
        ("league_mean_rat", str(league_mean_rat)),
        ("league_std_rat", str(league_std_rat)),
    ])
    conn.execute("""
        CREATE TABLE teams (
            team_id INTEGER PRIMARY KEY, name TEXT, division TEXT, conference TEXT,
            tier TEXT, current_rating REAL, sos REAL
        )
    """)
    conn.execute("""
        CREATE TABLE players (
            player_id INTEGER PRIMARY KEY, name TEXT, team_id INTEGER, division TEXT,
            position TEXT, class_year TEXT, season TEXT, games INTEGER,
            avg_minutes REAL, ppg REAL, rpg REAL, apg REAL, bpg REAL, spg REAL, topg REAL,
            ts_pct REAL, fg_pct REAL,
            per40_pts REAL, per40_reb REAL, per40_ast REAL, per40_blk REAL, per40_stl REAL, per40_tov REAL,
            hoop_score REAL, hoop_score_raw REAL
        )
    """)
    conn.executemany(
        "INSERT INTO teams VALUES (:team_id,:name,:division,:conference,:tier,:current_rating,:sos)",
        team_rows,
    )
    conn.executemany(
        """INSERT INTO players VALUES (:player_id,:name,:team_id,:division,:position,:class_year,
           :season,:games,:avg_minutes,:ppg,:rpg,:apg,:bpg,:spg,:topg,:ts_pct,:fg_pct,
           :per40_pts,:per40_reb,:per40_ast,:per40_blk,:per40_stl,:per40_tov,
           :hoop_score,:hoop_score_raw)""",
        player_rows,
    )
    conn.execute("CREATE INDEX idx_players_name ON players(name)")
    conn.execute("CREATE INDEX idx_players_team ON players(team_id)")
    conn.execute("CREATE INDEX idx_teams_tier ON teams(tier)")
    conn.commit()
    conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
