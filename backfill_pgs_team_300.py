"""Backfill PlayerGameStats for games that already have a correct Games-sheet
row for --team-id, but zero PlayerGameStats rows for that team (see
check_pgs_gaps_team_300.py -- run that first to see the gap this fixes).
 
Assumes the Games sheet's Team ID columns are already correct (that's how
"games involving --team-id" is determined here) -- this script only adds
missing PlayerGameStats rows; it does not touch the Games sheet.
 
For each affected game:
  - fetches the box score from ESPN (no custom User-Agent -- see
    espn_client.py's comment on why a custom UA gets 403'd)
  - writes PlayerGameStats rows for --team-id's side
  - opportunistically writes the OPPONENT's side too, but only if the
    opponent also has zero PlayerGameStats rows for that game already
    (skipped otherwise, so nothing already-correct gets duplicated)
  - creates new Players rows for any athlete not already known (matched
    by ESPN athlete ID, same as run_d1_scrape.py)
 
Usage:
    python backfill_pgs_team_300.py --team-id 300                 # dry run, no writes
    python backfill_pgs_team_300.py --team-id 300 --apply          # fetch + write
    python backfill_pgs_team_300.py --team-id 300 --path "C:\\...\\WomensSummitTPE.xlsx"
 
Requires: openpyxl, requests.
"""
 
import argparse
import os
import sys
import time
 
import requests
 
BASE = "https://site.api.espn.com/apis/site/v2/sports/basketball/womens-college-basketball"
BOX_SCORE_STAT_ORDER = [
    "MIN", "PTS", "FG", "3PT", "FT", "REB", "AST", "TO", "STL", "BLK",
    "OREB", "DREB", "PF",
]
 
 
def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping
 
 
def _get(url, params=None, retries=4, backoff=2.0, timeout=20):
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(url, params=params, timeout=timeout)
            if resp.status_code == 429:
                wait = backoff * attempt * 3
                print(f"    [rate limited] sleeping {wait:.0f}s ...")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp.json()
        except (requests.RequestException, ValueError) as exc:
            last_exc = exc
            wait = backoff * attempt
            print(f"    [retry {attempt}/{retries}] {url} -> {exc}; sleeping {wait:.0f}s")
            time.sleep(wait)
    raise RuntimeError(f"Giving up on {url}: {last_exc}")
 
 
def get_boxscore(event_id):
    url = f"{BASE}/summary"
    data = _get(url, params={"event": event_id})
    players_block = (data.get("boxscore") or {}).get("players")
    if not players_block:
        return {}
    out = {}
    for team_block in players_block:
        team_id = (team_block.get("team") or {}).get("id")
        stats_block = (team_block.get("statistics") or [{}])[0]
        names = stats_block.get("names") or BOX_SCORE_STAT_ORDER
        athletes = stats_block.get("athletes") or []
        rows = []
        for a in athletes:
            athlete = a.get("athlete") or {}
            raw_stats = a.get("stats") or []
            stat_map = dict(zip(names, raw_stats))
            rows.append({
                "espn_athlete_id": athlete.get("id"),
                "name": athlete.get("displayName"),
                "position": (athlete.get("position") or {}).get("abbreviation"),
                "starter": bool(a.get("starter")),
                "raw_stats": stat_map,
            })
        out[str(team_id)] = rows
    return out
 
 
def split_made_attempted(value):
    if not value or value == "-":
        return 0, 0
    try:
        made, attempted = value.split("-")
        return int(made), int(attempted)
    except (ValueError, AttributeError):
        return 0, 0
 
 
def parse_minutes(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0
 
 
def parse_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0
 
 
def save_with_retry(wb, path, attempts=5, wait_seconds=10):
    tmp_path = f"{path}.tmp_saving"
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, path)
            return
        except PermissionError as exc:
            last_error = exc
            print(f"  [!] Could not save {path} (attempt {attempt}/{attempts}) -- "
                  f"is it open in Excel? Close it and I'll retry in {wait_seconds}s.")
            time.sleep(wait_seconds)
    raise last_error
 
 
def find_gaps(games_ws, ghmap, pgs_ws, phmap, team_id):
    """Games rows involving team_id whose (team_id, game_id) has zero
    PlayerGameStats rows. Returns list of dicts with everything needed to
    fetch+write, plus which (team_id, game_id) pairs already have stats at
    all (so the opponent-side backfill knows what to skip)."""
    home_id_col = ghmap["Home Team ID"] - 1
    away_id_col = ghmap["Away Team ID"] - 1
    home_name_col = ghmap["Home Team"] - 1
    away_name_col = ghmap["Away Team"] - 1
    game_id_col = ghmap["Game ID"] - 1
    season_col = ghmap["Season"] - 1
    date_col = ghmap["Date"] - 1
 
    games_for_team = {}
    for row in games_ws.iter_rows(min_row=2, values_only=True):
        gid = row[game_id_col]
        if not gid:
            continue
        home_id, away_id = row[home_id_col], row[away_id_col]
        if home_id != team_id and away_id != team_id:
            continue
        games_for_team[str(gid)] = {
            "season": row[season_col], "date": row[date_col],
            "home_id": home_id, "away_id": away_id,
            "home_name": row[home_name_col], "away_name": row[away_name_col],
        }
 
    pgs_team_col = phmap["Team ID"] - 1
    pgs_game_col = phmap["Game ID"] - 1
    existing_pairs = set()
    for row in pgs_ws.iter_rows(min_row=2, values_only=True):
        t, g = row[pgs_team_col], row[pgs_game_col]
        if t is not None and g:
            existing_pairs.add((t, str(g)))
 
    gaps = [
        (gid, info) for gid, info in games_for_team.items()
        if (team_id, gid) not in existing_pairs
    ]
    return gaps, existing_pairs
 
 
def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--path", default="WomensSummitTPE.xlsx")
    parser.add_argument("--team-id", type=int, required=True)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    team_id = args.team_id
 
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required (pip install openpyxl).")
 
    print(f"Scanning {args.path} (read-only) ...")
    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
    ghmap = header_map(wb["Games"])
    phmap = header_map(wb["PlayerGameStats"])
    gaps, _ = find_gaps(wb["Games"], ghmap, wb["PlayerGameStats"], phmap, team_id)
    wb.close()
 
    print(f"Games involving Team {team_id} with zero PlayerGameStats rows: {len(gaps)}")
    if not gaps:
        print("Nothing to fix.")
        return
    if not args.apply:
        print("\nRun again with --apply to fetch box scores and write the missing rows.")
        return
 
    print(f"\nOpening {args.path} (write mode) ...")
    wb = openpyxl.load_workbook(args.path)
    games_ws = wb["Games"]
    ghmap = header_map(games_ws)
    pgs_ws = wb["PlayerGameStats"]
    pghmap = header_map(pgs_ws)
    players_ws = wb["Players"]
    phmap = header_map(players_ws)
    teams_ws = wb["Teams"]
    thmap = header_map(teams_ws)
 
    # team_id -> division, for stamping Division/Opponent Level correctly
    team_id_to_division = {}
    for row in range(2, teams_ws.max_row + 1):
        tid = teams_ws.cell(row=row, column=thmap["Team ID"]).value
        if tid is not None:
            team_id_to_division[tid] = teams_ws.cell(row=row, column=thmap["Division"]).value
 
    # our own ESPN Team ID -- this is what a box score's per-side keys
    # actually are (see get_boxscore()). Comparing this against opp_team_id
    # (an INTERNAL Team ID from the Games sheet) is a different ID space and
    # will essentially never match -- that was the bug. We only need to know
    # OUR OWN espn id; whichever box-score key isn't ours is unambiguously
    # the opponent, and we already know the opponent's correct internal
    # Team ID directly from the Games sheet (opp_team_id below).
    my_espn_id = None
    for row in range(2, teams_ws.max_row + 1):
        if teams_ws.cell(row=row, column=thmap["Team ID"]).value == team_id:
            espn_val = teams_ws.cell(row=row, column=thmap["ESPN Team ID"]).value
            my_espn_id = str(espn_val) if espn_val is not None else None
            break
    if not my_espn_id:
        sys.exit(f"Team {team_id} has no ESPN Team ID set on the Teams sheet -- "
                  f"can't safely tell which box-score side is ours.")
 
    gaps, existing_pairs = find_gaps(games_ws, ghmap, pgs_ws, pghmap, team_id)
 
    espn_to_player_id = {}
    max_pid = 0
    for row in range(2, players_ws.max_row + 1):
        pid = players_ws.cell(row=row, column=phmap["Player ID"]).value
        ext_id = players_ws.cell(row=row, column=phmap["External ID"]).value
        if isinstance(pid, int):
            max_pid = max(max_pid, pid)
        if ext_id:
            espn_to_player_id[str(ext_id)] = pid
    next_player_id = [max_pid + 1]
    next_row_players = [players_ws.max_row + 1]
    next_row_pgs = [pgs_ws.max_row + 1]
 
    def write_side(resolved_team_id, gid, opp_team_id, opp_name, date_val, season, rows):
        division = team_id_to_division.get(resolved_team_id)
        opp_level = team_id_to_division.get(opp_team_id)
        written = 0
        for player_row in rows:
            if not player_row.get("espn_athlete_id"):
                continue
            ext_id = str(player_row["espn_athlete_id"])
            if ext_id in espn_to_player_id:
                player_id = espn_to_player_id[ext_id]
            else:
                player_id = next_player_id[0]
                next_player_id[0] += 1
                espn_to_player_id[ext_id] = player_id
                prow = next_row_players[0]
                next_row_players[0] += 1
                name = player_row["name"] or ""
                first_name = name.split(" ")[0]
                last_name = " ".join(name.split(" ")[1:])
                players_ws.cell(row=prow, column=phmap["Player ID"], value=player_id)
                players_ws.cell(row=prow, column=phmap["First Name"], value=first_name)
                players_ws.cell(row=prow, column=phmap["Last Name"], value=last_name)
                players_ws.cell(row=prow, column=phmap["Team ID"], value=resolved_team_id)
                players_ws.cell(row=prow, column=phmap["Position"], value=player_row["position"])
                players_ws.cell(row=prow, column=phmap["Finished"], value="Yes")
                players_ws.cell(row=prow, column=phmap["External ID"], value=player_row["espn_athlete_id"])
                players_ws.cell(row=prow, column=phmap["Source"], value="ESPN")
                players_ws.cell(row=prow, column=phmap["Division"], value=division)
 
            stats = player_row["raw_stats"]
            fgm, fga = split_made_attempted(stats.get("FG"))
            fg3m, fg3a = split_made_attempted(stats.get("3PT"))
            ftm, fta = split_made_attempted(stats.get("FT"))
            values = {
                "Player ID": player_id, "Team ID": resolved_team_id, "Opponent": opp_name,
                "Opponent Team ID": opp_team_id, "Opponent Level": opp_level,
                "Date": date_val, "GS": "Yes" if player_row["starter"] else "No",
                "Min": parse_minutes(stats.get("MIN")), "FG Made": fgm, "FG Attempt": fga,
                "3FG M": fg3m, "3FG A": fg3a, "FT M": ftm, "FT A": fta,
                "Rebound": parse_int(stats.get("REB")), "Foul": parse_int(stats.get("PF")),
                "Ast": parse_int(stats.get("AST")), "To": parse_int(stats.get("TO")),
                "Blk": parse_int(stats.get("BLK")), "Stl": parse_int(stats.get("STL")),
                "Points": parse_int(stats.get("PTS")), "Game ID": gid, "Season": season,
                "Position": player_row["position"], "Division": division,
            }
            prow2 = next_row_pgs[0]
            next_row_pgs[0] += 1
            for key, val in values.items():
                c = pghmap.get(key)
                if c:
                    pgs_ws.cell(row=prow2, column=c, value=val)
            written += 1
        return written
 
    print(f"Backfilling {len(gaps)} games ...")
    games_touched = 0
    stat_rows_written = 0
    failures = []
    for i, (gid, info) in enumerate(gaps, 1):
        my_side_is_home = info["home_id"] == team_id
        opp_team_id = info["away_id"] if my_side_is_home else info["home_id"]
        opp_name = info["away_name"] if my_side_is_home else info["home_name"]
 
        try:
            box = get_boxscore(gid)
        except RuntimeError as exc:
            print(f"  [!] ({i}/{len(gaps)}) box score failed for game {gid}: {exc}")
            failures.append(gid)
            continue
 
        if not box:
            print(f"  [!] ({i}/{len(gaps)}) no box score data at all for game {gid} "
                  f"(postponed / no data on ESPN?) -- skipping")
            continue
 
        for espn_team_key, rows in box.items():
            # figure out whether this box-score side is us or the opponent.
            # espn_team_key is an ESPN team id (e.g. "245"); compare it
            # against OUR OWN espn id, not against opp_team_id (an internal
            # id -- comparing those two id spaces was the original bug).
            if espn_team_key == my_espn_id:
                resolved_team_id = team_id
            else:
                resolved_team_id = opp_team_id
            if (resolved_team_id, gid) in existing_pairs:
                continue  # already has stats, don't duplicate
            other_team_id = opp_team_id if resolved_team_id == team_id else team_id
            other_name = opp_name if resolved_team_id == team_id else (
                info["home_name"] if my_side_is_home else info["away_name"]
            )
            n = write_side(resolved_team_id, gid, other_team_id, other_name,
                            info["date"], info["season"], rows)
            existing_pairs.add((resolved_team_id, gid))
            stat_rows_written += n
 
        games_touched += 1
        if i % 10 == 0 or i == len(gaps):
            print(f"  ({i}/{len(gaps)}) processed")
 
    print(f"\nBackfilled stats for {games_touched}/{len(gaps)} games, "
          f"wrote {stat_rows_written} PlayerGameStats rows.")
    if failures:
        print(f"  [!] {len(failures)} games failed and were left unfixed: {', '.join(failures)}")
        print("  Re-run with --apply again later to retry just those.")
 
    print("\nSaving ...")
    save_with_retry(wb, args.path)
    print("Done. Next: run run_d1_scrape.py once more (harmless -- it'll just re-run "
          "finalize_players(), rebuilding PlayerSeasons from the now-complete "
          "PlayerGameStats sheet), then re-run compute_derived_sheets.py and build_cache.py.")
 
 
if __name__ == "__main__":
    main()