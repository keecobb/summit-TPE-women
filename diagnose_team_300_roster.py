"""Read-only diagnostic: lists every player currently assigned to
--team-id on the Players sheet (the "current roster") and on PlayerSeasons
(broken out by season), with basic per-season stats, so you can eyeball
exactly who doesn't belong and how badly padded it is.
 
Doesn't touch the workbook. Doesn't guess at a fix -- the goal here is
just to SEE the actual rows so we can figure out the real cause (a real
one-off script bug vs. something upstream like finalize_players()
reassigning players it shouldn't).
 
Usage:
    python diagnose_team_300_roster.py
    python diagnose_team_300_roster.py --team-id 300 --path "C:\\...\\WomensSummitTPE.xlsx"
"""
 
import argparse
from collections import defaultdict
 
 
def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping
 
 
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", default="WomensSummitTPE.xlsx")
    parser.add_argument("--team-id", type=int, default=300)
    args = parser.parse_args()
    team_id = args.team_id
 
    import openpyxl
    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
 
    teams_ws = wb["Teams"]
    thmap = header_map(teams_ws)
    team_name = None
    for row in teams_ws.iter_rows(min_row=2, values_only=True):
        if row[thmap["Team ID"] - 1] == team_id:
            team_name = row[thmap["Team"] - 1]
            break
    print(f"Team {team_id}: {team_name!r}\n")
 
    # ---- Players sheet (current roster) ----
    players_ws = wb["Players"]
    phmap = header_map(players_ws)
    current_roster = []
    for row in players_ws.iter_rows(min_row=2, values_only=True):
        if row[phmap["Team ID"] - 1] == team_id:
            current_roster.append(dict(
                player_id=row[phmap["Player ID"] - 1],
                first=row[phmap["First Name"] - 1], last=row[phmap["Last Name"] - 1],
                position=row[phmap["Position"] - 1], class_year=row[phmap["Class"] - 1],
                external_id=row[phmap["External ID"] - 1],
                transfer_history=row[phmap.get("Transfer History", 0) - 1] if phmap.get("Transfer History") else None,
            ))
    print(f"Players sheet: {len(current_roster)} rows with Team ID = {team_id}")
 
    # ---- PlayerSeasons (per season) ----
    ps_ws = wb["PlayerSeasons"]
    pshmap = header_map(ps_ws)
    by_season = defaultdict(list)
    for row in ps_ws.iter_rows(min_row=2, values_only=True):
        if row[pshmap["Team ID"] - 1] == team_id:
            by_season[row[pshmap["Season"] - 1]].append(dict(
                player_id=row[pshmap["Player ID"] - 1],
                games=row[pshmap.get("Games Played", 0) - 1] if pshmap.get("Games Played") else None,
            ))
    print("PlayerSeasons rows with Team ID = {}: {} total".format(
        team_id, sum(len(v) for v in by_season.values())))
    for season, rows in sorted(by_season.items()):
        print(f"    {season}: {len(rows)} players")
 
    # ---- PlayerGameStats: how many rows per player are tagged this team ----
    pgs_ws = wb["PlayerGameStats"]
    pgh = header_map(pgs_ws)
    pgs_team_col = pgh["Team ID"] - 1
    pgs_pid_col = pgh["Player ID"] - 1
    pgs_season_col = pgh["Season"] - 1
    games_by_player = defaultdict(lambda: defaultdict(int))
    for row in pgs_ws.iter_rows(min_row=2, values_only=True):
        if row[pgs_team_col] == team_id:
            pid = row[pgs_pid_col]
            games_by_player[pid][row[pgs_season_col]] += 1
    wb.close()
 
    print(f"\n{len(current_roster)} players currently on the Players-sheet roster for Team {team_id}:")
    print(f"{'Player ID':>10}  {'Name':<25} {'Pos':<5} {'Class':<6} {'ExtID':<12} {'PGS rows by season'}")
    for p in sorted(current_roster, key=lambda x: (x["last"] or "", x["first"] or "")):
        name = f"{p['first'] or ''} {p['last'] or ''}".strip()
        pgs_summary = ", ".join(f"{s}:{n}" for s, n in sorted(games_by_player.get(p["player_id"], {}).items()))
        print(f"{p['player_id']:>10}  {name:<25} {p['position'] or '':<5} {p['class_year'] or '':<6} "
              f"{str(p['external_id'] or ''):<12} {pgs_summary}")
 
 
if __name__ == "__main__":
    main()