"""Detect box-score data contamination in PlayerGameStats using two
validated signatures (task #63 follow-up -- the first version of this
script, which just thresholded on raw player count, was too noisy: most
"big roster" games turned out to be real deep-bench usage, not
corruption. This version replaces that with two signals confirmed
against real cases before shipping:

  1. MERGED_OPPONENT: a team's total logged Minutes for one game sums to
     roughly DOUBLE a normal game (~200 player-minutes per team per
     regulation game -- 5 players x 40 minutes; OT adds ~25/period).
     Confirmed live: Lamar vs Texas College (Game 401583922) shows 26
     players and exactly 400 total minutes for Lamar's Team ID -- Texas
     College's entire real box score (its own ~13 players, all with real
     minutes/points) got attributed to Lamar's Team ID instead of (or in
     addition to) Lamar's own real 13. Every player in a MERGED_OPPONENT
     game has REAL, nonzero stats, so a zero-stat filter can't separate
     the two teams -- this script just flags the game for manual review
     rather than guessing which half is which.

  2. PHANTOM_CLUSTER: a team's box score has an implausibly high player
     count where the "extra" players (beyond a normal ~13-16) all have
     ~0 minutes AND their ESPN External IDs form a tight, near-
     consecutive block, unlike a real roster's IDs (which span the
     player's whole career and are scattered over hundreds of thousands
     of ID values). Confirmed live twice: Abilene Christian vs Howard
     Payne (Game 401826153) -- 13 real players (real minutes/points) +
     17 zero-stat players with External IDs 5320981-5320997 (a real
     University of Rochester roster, verified by name). VCU vs Davidson
     (Game 401728298) -- 9 real-minute players + 12 zero-stat players
     with External IDs 5257490-5257503. Total minutes for both games
     stayed at a completely normal 200 -- the phantom block contributes
     zero playing time, it's just extra rows glued onto an otherwise-
     correct box score.

This script is READ-ONLY -- it never modifies the workbook. For each
flagged game it prints enough detail (every player, minutes, points,
External ID) for a human to confirm before any cleanup script touches
the data.

Usage:
    python scrapers/detect_roster_contamination.py
    python scrapers/detect_roster_contamination.py --min-players 18 --min-zero 8
"""

import argparse
import collections
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl

from scrapers.xlsx_io import header_map

DEFAULT_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "WomensSummitTPE.xlsx")

# A regulation game is 40 minutes x 5 players on court = 200 team-minutes.
# Overtime adds ~25 team-minutes per 5-minute OT period. Generous margin
# to avoid false positives on multi-OT games.
NORMAL_MINUTES_MAX = 260
MERGED_OPPONENT_MINUTES_MIN = 320  # comfortably above even a 2-OT normal game

DEFAULT_MIN_PLAYERS = 18   # a normal box score (deep bench included) rarely exceeds this
DEFAULT_MIN_ZERO = 8       # how many of the "extra" players must be zero-minute to suspect a phantom block
ID_CLUSTER_GAP = 500       # max gap between consecutive External IDs to still count as "the same bulk-created block"
ID_CLUSTER_MIN_SIZE = 6    # minimum cluster size to flag (avoids 2-3 coincidentally-close real IDs)


def find_id_clusters(pid_ext_pairs, gap, min_size):
    """pid_ext_pairs: [(player_id, external_id)], external_id may be None
    (skipped). Returns list of clusters (each a list of player_ids) where
    consecutive (sorted) external IDs are within `gap` of each other and
    the cluster has >= min_size members."""
    usable = [(int(ext), pid) for pid, ext in pid_ext_pairs if ext not in (None, "")]
    usable.sort()
    clusters = []
    current = []
    for ext, pid in usable:
        if current and ext - current[-1][0] > gap:
            if len(current) >= min_size:
                clusters.append([p for _, p in current])
            current = []
        current.append((ext, pid))
    if len(current) >= min_size:
        clusters.append([p for _, p in current])
    return clusters


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--min-players", type=int, default=DEFAULT_MIN_PLAYERS)
    parser.add_argument("--min-zero", type=int, default=DEFAULT_MIN_ZERO)
    parser.add_argument("path", nargs="?", default=DEFAULT_PATH)
    args = parser.parse_args()

    print(f"Opening {args.path} (read-only) ...")
    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)

    tws = wb["Teams"]
    th = header_map(tws)
    team_name = {}
    for row in tws.iter_rows(min_row=2, values_only=True):
        tid = row[th["Team ID"] - 1]
        if tid is not None:
            team_name[tid] = row[th["Team"] - 1]

    pws = wb["Players"]
    ph = header_map(pws)
    player_name = {}
    player_ext = {}
    for row in pws.iter_rows(min_row=2, values_only=True):
        pid = row[ph["Player ID"] - 1]
        if pid is not None:
            first = row[ph["First Name"] - 1] or ""
            last = row[ph["Last Name"] - 1] or ""
            player_name[pid] = f"{first} {last}".strip()
            if "External ID" in ph:
                player_ext[pid] = row[ph["External ID"] - 1]

    pgws = wb["PlayerGameStats"]
    pgh = header_map(pgws)
    col_pid = pgh["Player ID"] - 1
    col_team = pgh["Team ID"] - 1
    col_game = pgh["Game ID"] - 1
    col_min = pgh["Min"] - 1
    col_pts = pgh["Points"] - 1
    col_opp = pgh.get("Opponent")
    col_opp = col_opp - 1 if col_opp else None
    col_date = pgh.get("Date")
    col_date = col_date - 1 if col_date else None
    col_season = pgh.get("Season")
    col_season = col_season - 1 if col_season else None

    print("Scanning PlayerGameStats (this can take a minute or two on the full sheet) ...")
    groups = collections.defaultdict(lambda: {"rows": [], "opponent": None, "date": None, "season": None})
    n = 0
    for row in pgws.iter_rows(min_row=2, values_only=True):
        n += 1
        tid = row[col_team]
        gid = row[col_game]
        if tid is None or gid is None:
            continue
        key = (tid, gid)
        g = groups[key]
        g["rows"].append((row[col_pid], row[col_min] or 0, row[col_pts] or 0))
        if g["opponent"] is None:
            g["opponent"] = row[col_opp] if col_opp is not None else None
            g["date"] = row[col_date] if col_date is not None else None
            g["season"] = row[col_season] if col_season is not None else None
        if n % 200000 == 0:
            print(f"  ... {n} rows scanned")

    print(f"{n} PlayerGameStats rows scanned, {len(groups)} distinct (Team, Game) box scores.")

    merged_opponent = []
    phantom_cluster = []
    for (tid, gid), g in groups.items():
        rows = g["rows"]
        total_min = sum(r[1] for r in rows)
        player_count = len(rows)
        zero_min_count = sum(1 for r in rows if r[1] == 0)

        if total_min >= MERGED_OPPONENT_MINUTES_MIN:
            merged_opponent.append((tid, gid, g, rows, total_min))
        elif player_count >= args.min_players and zero_min_count >= args.min_zero:
            phantom_cluster.append((tid, gid, g, rows, total_min))

    print(f"\n{len(merged_opponent)} MERGED_OPPONENT candidates (total minutes >= {MERGED_OPPONENT_MINUTES_MIN}, "
          f"suggests two teams' box scores combined under one Team ID):")
    for tid, gid, g, rows, total_min in sorted(merged_opponent, key=lambda x: -x[4]):
        print(f"\n=== {team_name.get(tid, tid)} (Team ID {tid}) vs {g['opponent']} on {g['date']} "
              f"(Season {g['season']}, Game ID {gid}): {len(rows)} players, {total_min} total minutes ===")
        for pid, mins, pts in sorted(rows, key=lambda r: -r[1]):
            ext_display = player_ext.get(pid) or ""
            print(f"    Player ID {pid:>6}  ext {ext_display:<10}  min={mins:<4} pts={pts:<4}  {player_name.get(pid, '?')}")

    print(f"\n\n{len(phantom_cluster)} PHANTOM_CLUSTER candidates (>= {args.min_players} players, "
          f">= {args.min_zero} of them zero-minute, normal total minutes):")
    for tid, gid, g, rows, total_min in sorted(phantom_cluster, key=lambda x: -len(x[3])):
        print(f"\n=== {team_name.get(tid, tid)} (Team ID {tid}) vs {g['opponent']} on {g['date']} "
              f"(Season {g['season']}, Game ID {gid}): {len(rows)} players, {total_min} total minutes ===")
        zero_pairs = [(pid, player_ext.get(pid)) for pid, mins, pts in rows if mins == 0]
        clusters = find_id_clusters(zero_pairs, ID_CLUSTER_GAP, ID_CLUSTER_MIN_SIZE)
        suspect_pids = {pid for c in clusters for pid in c}
        for pid, mins, pts in sorted(rows, key=lambda r: -r[1]):
            flag = " <-- SUSPECT (tight ext-ID cluster)" if pid in suspect_pids else ""
            ext_display = player_ext.get(pid) or ""
            print(f"    Player ID {pid:>6}  ext {ext_display:<10}  min={mins:<4} pts={pts:<4}  {player_name.get(pid, '?')}{flag}")
        if not clusters:
            print("    (no tight External-ID cluster found among the zero-minute players -- "
                  "review manually, may just be a deep real bench)")


if __name__ == "__main__":
    main()
