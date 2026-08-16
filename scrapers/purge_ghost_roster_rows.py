"""Purge confirmed ghost/phantom roster rows from the Players sheet --
the roster_too_large findings from check_data_health.py, triaged down to
the small subset that's real contamination (most of the 82/168 flagged
teams are just the workbook's 3-season roster history, which the health
check's docstring incorrectly assumes is a single-season snapshot).

Detection (same two-stage heuristic used for the manual investigation):
  1. Tight External-ID clustering within one team's Players rows -- a
     real 3-season roster has IDs scattered across a player's whole
     career; a bulk-mis-copied block lands as a tight, near-consecutive
     run.
  2. Cross-team/cross-sport name duplication -- >=50% of a cluster's
     players sharing an exact (First, Last) name with a player on a
     DIFFERENT team (either sport) confirms the cluster is someone
     else's real roster glued onto this team.

Because the two sheets are SEPARATE WORKBOOKS with independently
assigned Player IDs that collide constantly (8375 of 8738 women's IDs
also exist as men's IDs), this script loads BOTH files for detection
but only ever writes back to whichever file(s) actually have rows to
remove, and every "has real data" check is done against that row's OWN
sport's PlayerGameStats/PlayerSeasons sheets -- never a pooled union.

Safety gate, applied per candidate cluster:
  A cluster is only auto-purged if 100% of its members have ZERO
  PlayerGameStats rows AND zero PlayerSeasons rows anywhere in their
  own sport's workbook. If even one member has real stats attached,
  the WHOLE cluster is skipped and printed for manual review -- this
  is what correctly excludes e.g. Georgia Southern's 11-row cluster
  (real players who happen to share a name-duplication signature with
  Florida Gulf Coast's real ghost cluster) and Mississippi Valley
  State's 10-row cluster (real women's players who share names with
  Alabama A&M's ghost men's cluster).

Default is a DRY RUN -- prints exactly what would change, touches
nothing. Pass --write to actually modify and save the workbook(s).

Usage:
    python scrapers/purge_ghost_roster_rows.py                # dry run, both files
    python scrapers/purge_ghost_roster_rows.py --write         # apply
"""

import argparse
import collections
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from scrapers.xlsx_io import header_map, open_workbook, save_with_retry, wipe_data_rows

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_WOMEN_PATH = os.path.join(ROOT, "WomensSummitTPE.xlsx")
DEFAULT_MEN_PATH = os.path.join(ROOT, "MensSummitTPE.xlsx")

ID_CLUSTER_GAP = 500
ID_CLUSTER_MIN_SIZE = 4
NAME_DUP_THRESHOLD = 0.5
MAX_ROSTER = 26


def find_id_clusters(pid_ext_pairs, gap, min_size):
    usable = [(int(ext), pid) for pid, ext in pid_ext_pairs if ext not in (None, "")]
    usable.sort()
    clusters = []
    current = []
    for ext, pid in usable:
        if current and ext - current[-1][0] > gap:
            if len(current) >= min_size:
                clusters.append([p for _, p in current])
            current = []
        current.append((ext, pid))
    if len(current) >= min_size:
        clusters.append([p for _, p in current])
    return clusters


def load_readonly(path, sport):
    """Read-only pass (fast) used for detection only."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    tws = wb["Teams"]
    th = header_map(tws)
    teams = {}
    for row in tws.iter_rows(min_row=2, values_only=True):
        tid = row[th["Team ID"] - 1]
        if tid is not None:
            teams[tid] = row[th["Team"] - 1]

    pws = wb["Players"]
    ph = header_map(pws)
    players = []
    for row in pws.iter_rows(min_row=2, values_only=True):
        pid = row[ph["Player ID"] - 1]
        if pid is None:
            continue
        first = row[ph["First Name"] - 1] or ""
        last = row[ph["Last Name"] - 1] or ""
        players.append(dict(
            sport=sport, player_id=pid, team_id=row[ph["Team ID"] - 1],
            name=f"{first} {last}".strip().lower(),
            display_name=f"{first} {last}".strip(),
            ext=row[ph["External ID"] - 1] if "External ID" in ph else None,
        ))

    pgws = wb["PlayerGameStats"]
    pgh = header_map(pgws)
    pgs_pids = set()
    for row in pgws.iter_rows(min_row=2, values_only=True):
        pid = row[pgh["Player ID"] - 1]
        if pid is not None:
            pgs_pids.add(pid)

    psws = wb["PlayerSeasons"]
    psh = header_map(psws)
    ps_pids = set()
    for row in psws.iter_rows(min_row=2, values_only=True):
        pid = row[psh["Player ID"] - 1]
        if pid is not None:
            ps_pids.add(pid)

    wb.close()
    return teams, players, pgs_pids, ps_pids


def detect(women_path, men_path):
    w_teams, w_players, w_pgs, w_ps = load_readonly(women_path, "WOMEN")
    m_teams, m_players, m_pgs, m_ps = load_readonly(men_path, "MEN")

    all_players = w_players + m_players
    pgs_by_sport = {"WOMEN": w_pgs, "MEN": m_pgs}
    ps_by_sport = {"WOMEN": w_ps, "MEN": m_ps}
    teams_by_sport = {"WOMEN": w_teams, "MEN": m_teams}

    name_teams = collections.defaultdict(set)
    for p in all_players:
        name_teams[p["name"]].add((p["sport"], p["team_id"]))

    by_team = collections.defaultdict(list)
    for p in all_players:
        by_team[(p["sport"], p["team_id"])].append(p)

    confirmed = []    # clusters where 100% of members are zero-impact -- safe to purge
    skipped = []      # clusters that matched the pattern but have real data mixed in -- manual review

    for (sport, tid), roster in by_team.items():
        if len(roster) <= MAX_ROSTER:
            continue
        pairs = [(p["player_id"], p["ext"]) for p in roster]
        clusters = find_id_clusters(pairs, ID_CLUSTER_GAP, ID_CLUSTER_MIN_SIZE)
        by_pid = {p["player_id"]: p for p in roster}
        for cluster_pids in clusters:
            cluster_players = [by_pid[pid] for pid in cluster_pids]
            dup_count = sum(1 for p in cluster_players if name_teams[p["name"]] - {(sport, tid)})
            frac = dup_count / len(cluster_players)
            if frac < NAME_DUP_THRESHOLD:
                continue
            impacted = [p for p in cluster_players if p["player_id"] in pgs_by_sport[sport]
                        or p["player_id"] in ps_by_sport[sport]]
            entry = dict(sport=sport, team_id=tid, team_name=teams_by_sport[sport].get(tid, "?"),
                         players=cluster_players, frac=frac, impacted=impacted)
            if impacted:
                skipped.append(entry)
            else:
                confirmed.append(entry)

    return confirmed, skipped


def apply_purge(path, sport, pids_to_remove, write):
    if not pids_to_remove:
        return
    print(f"\nOpening {path} ...")
    wb = open_workbook(path)
    pws = wb["Players"]
    ph = header_map(pws)

    kept = []
    removed = 0
    for row in range(2, pws.max_row + 1):
        pid = pws.cell(row=row, column=ph["Player ID"]).value
        if pid in pids_to_remove:
            removed += 1
            continue
        vals = [pws.cell(row=row, column=c).value for c in range(1, pws.max_column + 1)]
        if all(v is None for v in vals):
            continue
        kept.append(vals)

    print(f"  Players ({sport}): {len(kept)} rows kept, {removed} removed.")

    if not write:
        return

    new_pws = wipe_data_rows(wb, "Players", keep_header=True)
    for vals in kept:
        new_pws.append(vals)
    print("  Saving ...")
    save_with_retry(wb, path)
    print("  Done.")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--women-path", default=DEFAULT_WOMEN_PATH)
    parser.add_argument("--men-path", default=DEFAULT_MEN_PATH)
    parser.add_argument("--write", action="store_true", help="Actually modify and save the workbook(s) (default: dry run)")
    args = parser.parse_args()

    print(f"Scanning {args.women_path} and {args.men_path} for ghost roster clusters ...")
    confirmed, skipped = detect(args.women_path, args.men_path)

    print(f"\n{len(confirmed)} cluster(s) confirmed safe to purge (100% zero PlayerGameStats/PlayerSeasons impact):")
    to_remove = {"WOMEN": set(), "MEN": set()}
    for c in confirmed:
        print(f"\n=== {c['sport']} Team {c['team_id']} ({c['team_name']}) -- {len(c['players'])} rows, "
              f"{c['frac']:.0%} name-duplicated elsewhere ===")
        for p in c["players"]:
            print(f"    Player {p['player_id']:>6}  ext={p['ext']}  {p['display_name']!r}")
            to_remove[c["sport"]].add(p["player_id"])

    if skipped:
        print(f"\n{len(skipped)} cluster(s) matched the pattern but have real data mixed in -- "
              f"NOT auto-purged, review manually:")
        for c in skipped:
            print(f"\n=== {c['sport']} Team {c['team_id']} ({c['team_name']}) -- {len(c['players'])} rows, "
                  f"{len(c['impacted'])} of them have real PlayerGameStats/PlayerSeasons rows ===")
            for p in c["impacted"]:
                print(f"    Player {p['player_id']:>6}  {p['display_name']!r}  <-- HAS REAL DATA, cluster skipped")

    total = sum(len(s) for s in to_remove.values())
    print(f"\nTotal rows to remove: {total} (Women's: {len(to_remove['WOMEN'])}, Men's: {len(to_remove['MEN'])})")

    if not args.write:
        print("\n--write not passed: dry run only, workbook(s) NOT modified.")
        return

    apply_purge(args.women_path, "WOMEN", to_remove["WOMEN"], write=True)
    apply_purge(args.men_path, "MEN", to_remove["MEN"], write=True)
    print("\nDone.")


if __name__ == "__main__":
    main()
