"""Fix duplicate_pgs_row findings from check_data_health.py -- a player
with more than one PlayerGameStats row for the same Game ID.

Two distinct sub-patterns, handled differently:

  1. LITERAL DUPLICATE: every row in the group has the exact same Team
     ID and the exact same stat line (minutes, points, everything).
     Harmless double-write -- just keep one copy, drop the rest.

  2. CROSS-TEAM DUPLICATE: the rows have the same stat line but
     DIFFERENT Team IDs -- one real box-score line got attributed to
     both sides of the game (confirmed live: Providence/VCU Game
     401713181, four players; UNC Asheville/Southeast Missouri State
     Game 401719238, ten players -- in both cases the "extra" team's
     row for these players carried real minutes/points that also
     belong to the correct team, silently inflating both teams'
     aggregates). This DOES need a real fix, not just a dedup, because
     removing the wrong side is what actually matters -- so this
     script determines which Team ID is correct for that player in
     that season by parsing the Players sheet's "Transfer History"
     column (e.g. "Southeast Missouri State (2024-25 -- 2025-26)" or
     "Providence (2023-24) -> VCU (2024-25)"), and only auto-fixes rows
     where exactly one of the candidate teams matches the player's
     transfer history for that season. Anything else -- unparseable
     history (including the literal "#REF!" marker), or more than one
     team matching -- is left alone and printed for manual review.

     "GS" (Game Started) is deliberately excluded from the required
     stat match: it's been observed to differ between two copies of an
     otherwise byte-identical real stat line (Valentina Ojeda, Game
     401713181 -- 'No' on the Providence-side copy, 'Yes' on the
     VCU-side copy, every other column identical). A GS mismatch alone
     doesn't stop the cross-team match; the kept row's own GS value is
     used regardless.

  Any group that's neither of the above (different stats, more than
  two rows with a mix of matches) is left alone and printed for manual
  review -- this script never guesses.

Detection runs against a fast READ-ONLY pass. Default is a DRY RUN --
prints exactly what would change, touches nothing. Pass --write to
reopen the workbook in write mode and actually apply + save.

Usage:
    python scrapers/fix_duplicate_pgs_rows.py WomensSummitTPE.xlsx                # dry run
    python scrapers/fix_duplicate_pgs_rows.py WomensSummitTPE.xlsx --write         # apply
    python scrapers/fix_duplicate_pgs_rows.py MensSummitTPE.xlsx --write
"""

import argparse
import collections
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl

from scrapers.xlsx_io import header_map, open_workbook, save_with_retry, wipe_data_rows

# "GS" excluded on purpose -- see module docstring.
STAT_COLUMNS = [
    "Min", "FG Made", "FG Attempt", "3FG M", "3FG A", "FT M", "FT A",
    "Rebound", "Foul", "Ast", "To", "Blk", "Stl", "Points",
]

TH_SEGMENT_RE = re.compile(r"^(.*?)\s*\((\d{4}-\d{2})(?:\s*--\s*(\d{4}-\d{2}))?\)\s*$")


def parse_transfer_history(th):
    """'VCU (2023-24 -- 2025-26)' or 'Providence (2023-24) -> VCU (2024-25)'
    -> [(team_name, season_label), ...] covering every season in each
    segment's range. Returns None if the string doesn't cleanly parse
    (including the literal '#REF!' corruption marker) -- callers must
    treat None as "don't know, don't guess."""
    if not th or not str(th).strip():
        return None
    segments = [s.strip() for s in str(th).split("->")]
    result = []
    for seg in segments:
        m = TH_SEGMENT_RE.match(seg)
        if not m:
            return None
        team_name, start, end = m.group(1).strip(), m.group(2), m.group(3) or m.group(2)
        start_year = int(start.split("-")[0])
        end_year = int(end.split("-")[0])
        if end_year < start_year or end_year - start_year > 6:
            return None
        for y in range(start_year, end_year + 1):
            result.append((team_name, f"{y}-{str(y + 1)[-2:]}"))
    return result


def detect(path):
    """Read-only detection pass. Returns (team_name, player_display,
    literal_fixed, cross_team_fixed, manual_review, rows_to_delete)."""
    print(f"Opening {path} (read-only) ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    tws = wb["Teams"]
    th_map = header_map(tws)
    team_name = {}
    for row in tws.iter_rows(min_row=2, values_only=True):
        tid = row[th_map["Team ID"] - 1]
        if tid is not None:
            team_name[tid] = row[th_map["Team"] - 1]

    pws = wb["Players"]
    ph = header_map(pws)
    player_transfer_history = {}
    player_display = {}
    for row in pws.iter_rows(min_row=2, values_only=True):
        pid = row[ph["Player ID"] - 1]
        if pid is None:
            continue
        first = row[ph["First Name"] - 1] or ""
        last = row[ph["Last Name"] - 1] or ""
        player_display[pid] = f"{first} {last}".strip()
        player_transfer_history[pid] = row[ph["Transfer History"] - 1] if "Transfer History" in ph else None

    pgws = wb["PlayerGameStats"]
    pgh = header_map(pgws)
    stat_col_idx = {name: pgh[name] - 1 for name in STAT_COLUMNS if name in pgh}
    col_pid = pgh["Player ID"] - 1
    col_gid = pgh["Game ID"] - 1
    col_team = pgh["Team ID"] - 1
    col_season = pgh["Season"] - 1 if "Season" in pgh else None

    print("Scanning PlayerGameStats ...")
    groups = collections.defaultdict(list)  # (player_id, game_id) -> [(row_idx, team_id, season, stats), ...]
    n = 0
    for row in pgws.iter_rows(min_row=2):
        n += 1
        pid = row[col_pid].value
        gid = row[col_gid].value
        if pid is None or gid is None:
            continue
        row_idx = row[0].row
        tid = row[col_team].value
        season = row[col_season].value if col_season is not None else None
        stats = tuple(row[c].value for c in stat_col_idx.values())
        groups[(pid, gid)].append((row_idx, tid, season, stats))
        if n % 200000 == 0:
            print(f"  ... {n} rows scanned")
    wb.close()

    dupe_groups = {k: v for k, v in groups.items() if len(v) > 1}
    print(f"{n} rows scanned, {len(dupe_groups)} (player, game) pair(s) with more than one row.")

    rows_to_delete = set()
    literal_fixed = []
    cross_team_fixed = []
    manual_review = []

    for (pid, gid), entries in dupe_groups.items():
        row_ids = [e[0] for e in entries]
        team_ids = {e[1] for e in entries}
        stat_sets = {e[3] for e in entries}

        if len(team_ids) == 1 and len(stat_sets) == 1:
            keep, *extra = sorted(row_ids)
            rows_to_delete.update(extra)
            literal_fixed.append((pid, gid, keep, extra))
            continue

        if len(team_ids) == len(entries) and len(stat_sets) == 1:
            season = entries[0][2]
            th = player_transfer_history.get(pid)
            parsed = parse_transfer_history(th)
            if parsed is None:
                manual_review.append((pid, gid, entries, "unparseable Transfer History"))
                continue
            season_teams = {tname for tname, s in parsed if s == season}
            matching = [e for e in entries if team_name.get(e[1]) in season_teams]
            if len(matching) != 1:
                manual_review.append((pid, gid, entries,
                                       f"{len(matching)} team(s) matched season {season} in Transfer History"))
                continue
            keep_entry = matching[0]
            extra = [e[0] for e in entries if e[0] != keep_entry[0]]
            rows_to_delete.update(extra)
            cross_team_fixed.append((pid, gid, keep_entry, extra, entries))
            continue

        manual_review.append((pid, gid, entries, "rows differ in a way this script doesn't auto-resolve"))

    return team_name, player_display, player_transfer_history, literal_fixed, cross_team_fixed, manual_review, rows_to_delete


def apply_write(path, rows_to_delete):
    # Gather the kept rows via a fast READ-ONLY bulk pass first -- looping
    # with pgws.cell(row=row, column=c).value on a non-read-only sheet for
    # ~350k rows x 28 columns (~9.7M individual cell lookups) is the actual
    # bottleneck here, not the fix logic itself (which resolves in
    # seconds). iter_rows(values_only=True) reads in bulk and is orders of
    # magnitude faster. Only the final wipe+append+save needs the
    # non-read-only workbook (to preserve every other sheet untouched).
    print(f"\nRe-reading {path} (read-only, bulk) to gather kept rows ...")
    ro_wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ro_pgws = ro_wb["PlayerGameStats"]
    kept_rows = []
    removed = 0
    for row_idx, row in enumerate(ro_pgws.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx in rows_to_delete:
            removed += 1
            continue
        if all(v is None for v in row):
            continue
        kept_rows.append(list(row))
    ro_wb.close()
    print(f"  PlayerGameStats: {len(kept_rows)} rows kept, {removed} removed.")

    print(f"Opening {path} (write mode) ...")
    wb = open_workbook(path)
    new_pgws = wipe_data_rows(wb, "PlayerGameStats", keep_header=True)
    for vals in kept_rows:
        new_pgws.append(vals)

    print("Saving ...")
    save_with_retry(wb, path)
    print("Done.")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--write", action="store_true", help="Actually modify and save the workbook (default: dry run)")
    parser.add_argument("path")
    args = parser.parse_args()

    (team_name, player_display, player_transfer_history,
     literal_fixed, cross_team_fixed, manual_review, rows_to_delete) = detect(args.path)

    print(f"\n{len(literal_fixed)} literal-duplicate group(s) -- keeping one copy, dropping the rest:")
    for pid, gid, keep, extra in literal_fixed:
        print(f"    Player {pid} ({player_display.get(pid, '?')}), Game {gid}: "
              f"kept row {keep}, removed row(s) {extra}")

    print(f"\n{len(cross_team_fixed)} cross-team double-count(s) -- removing the wrong-team row per Transfer History:")
    for pid, gid, keep_entry, extra, entries in cross_team_fixed:
        keep_row, keep_tid = keep_entry[0], keep_entry[1]
        for e in entries:
            if e[0] == keep_row:
                continue
            print(f"    Player {pid} ({player_display.get(pid, '?')}), Game {gid}: "
                  f"kept {team_name.get(keep_tid)!r} (row {keep_row}, matches Transfer History for that season), "
                  f"removed {team_name.get(e[1])!r} (row {e[0]})")

    if manual_review:
        print(f"\n{len(manual_review)} group(s) NOT auto-fixed -- need a human look:")
        for pid, gid, entries, reason in manual_review:
            print(f"    Player {pid} ({player_display.get(pid, '?')}), Game {gid}: -- {reason}")
            for row_idx, tid, season, stats in entries:
                print(f"        row {row_idx}: Team {tid} ({team_name.get(tid)}), Season {season}, "
                      f"TH={player_transfer_history.get(pid)!r}")

    print(f"\nTotal rows to remove: {len(rows_to_delete)}")

    if not args.write:
        print("\n--write not passed: dry run only, workbook NOT modified.")
        return

    if not rows_to_delete:
        print("\nNothing to change.")
        return

    apply_write(args.path, rows_to_delete)


if __name__ == "__main__":
    main()
