"""Shared helpers for reading/writing WomensSummitTPE.xlsx.

Every sheet's columns are looked up by header name (not hardcoded
letter/index), so if you insert or reorder a column in Excel these
scripts keep working instead of silently writing into the wrong
place. That's how we caught the earlier bug where find_womens_team_url.py
had URL/SiteType column letters hardcoded and a 'Division' column got
inserted in between, which would have silently corrupted data on the
next run.

All the "add missing columns" calls in this module APPEND new columns
at the end of a sheet -- they never reorder or remove what's already
there, so existing formulas/scripts that reference specific columns
keep working.
"""

import os
import time

import openpyxl


def open_workbook(path):
    """Open for read+write. Keeps formulas in other sheets untouched
    (data_only=False is the default, which is what we want for saving --
    data_only=True would drop formulas when the file is saved back)."""
    return openpyxl.load_workbook(path)


def header_map(ws):
    """Return {header_name: 1-based column index} from row 1.

    Blank/None header cells are skipped (PlayerGameStats has a couple
    of trailing unnamed columns -- ignored unless you name them).
    """
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping


def ensure_columns(ws, new_headers):
    """Append any header names not already present. Returns the
    refreshed header_map (existing columns untouched, order preserved)."""
    hmap = header_map(ws)
    next_col = ws.max_column + 1
    for name in new_headers:
        if name in hmap:
            continue
        ws.cell(row=1, column=next_col, value=name)
        hmap[name] = next_col
        next_col += 1
    return hmap


def wipe_data_rows(wb, sheet_name, keep_header=True):
    """Delete every data row, leaving (by default) just the header.

    Implemented as delete-sheet-and-recreate rather than ws.delete_rows(),
    because delete_rows() on a sheet with hundreds of thousands of rows
    is extremely slow (openpyxl shifts every remaining row one at a
    time) -- recreating the sheet is effectively instant regardless of
    how large it was.
    """
    ws = wb[sheet_name]
    idx = wb.sheetnames.index(sheet_name)
    header = None
    if keep_header and ws.max_row >= 1:
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    del wb[sheet_name]
    new_ws = wb.create_sheet(sheet_name, idx)
    if header:
        new_ws.append(header)
    return new_ws


def append_row(ws, hmap, row_dict, default=None):
    """Append one row, placing each value in the column matching its
    key in hmap. Keys not present in hmap are silently ignored (so
    callers can pass a superset dict safely); columns not present in
    row_dict are left as `default`."""
    width = ws.max_column
    out = [default] * width
    for key, value in row_dict.items():
        col = hmap.get(key)
        if col:
            out[col - 1] = value
    ws.append(out)


def append_rows(ws, hmap, row_dicts, default=None):
    for row_dict in row_dicts:
        append_row(ws, hmap, row_dict, default=default)


def iter_data_rows(ws, hmap):
    """Yield (row_index, {header_name: value}) for every data row."""
    reverse = {col: name for name, col in hmap.items()}
    for row in ws.iter_rows(min_row=2):
        row_idx = row[0].row
        values = {}
        for cell in row:
            name = reverse.get(cell.column)
            if name:
                values[name] = cell.value
        if any(v is not None for v in values.values()):
            yield row_idx, values


def save_with_retry(wb, path, attempts=5, wait_seconds=10):
    """Save, retrying if the file is open/locked in Excel.

    Windows will refuse to save over a file that's open in Excel.
    Rather than crash a multi-hour scrape over that, wait and retry --
    close the workbook in Excel and the next attempt will succeed.

    BUG FIXED HERE: this used to call wb.save(path) directly -- writing
    straight over the live, shared workbook in place. openpyxl writes a
    zip archive incrementally (small metadata parts like docProps/*.xml
    first, the big worksheet XML after), so anything that kills the
    process mid-write -- a Ctrl+C after a silently-buffered run made it
    LOOK like nothing was happening yet, a OneDrive sync hiccup, a
    crash -- leaves a truncated file behind: valid enough to look like
    an .xlsx (real PK zip header, a couple of tiny entries) but missing
    everything else, "zipfile.BadZipFile: File is not a zip file" the
    next time anything tries to open it. Confirmed live: exactly this
    happened to WomensSummitTPE.xlsx, dropping it from ~46MB to 2.9KB,
    only recoverable because a recent manual .bak backup happened to
    exist -- a checkpoint save mid-run wouldn't have one.
    Fix: save to a throwaway temp file in the same directory first (a
    half-written temp file is harmless -- the real workbook is never
    touched while that's happening), then atomically swap it into place
    with os.replace() only once the write has fully succeeded.
    os.replace() on the same filesystem is a single directory-entry
    update, not a byte-by-byte copy, so there's no window where the
    real path points at a partial file -- a kill at any point either
    leaves the OLD file completely intact or the NEW file fully in
    place, never something in between.
    """
    tmp_path = f"{path}.tmp_saving"
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, path)
            return
        except PermissionError as exc:
            last_error = exc
            print(
                f"  [!] Could not save {path} (attempt {attempt}/{attempts}) -- "
                f"is it open in Excel? Close it and I'll retry in {wait_seconds}s."
            )
            time.sleep(wait_seconds)
    raise last_error
