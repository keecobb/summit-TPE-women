"""One-time cleanup: strip ESPN's "did not play" ghost rows out of
PlayerGameStats and rebuild PlayerSeasons's Games Played / Games Started
from the cleaned data.

WHY THIS EXISTS
----------------
run_d1_scrape.py's (and run_d1_mbb_scrape.py's) box-score loop wrote a
PlayerGameStats row for every athlete ESPN's box score JSON listed for a
team -- including athletes marked `didNotPlay: true` with an entirely
empty stat line. ESPN keeps a player on a team's box score listing (as a
DNP entry) even after they've transferred away -- confirmed live against
a real example: Dallan Coleman transferred Georgia Tech -> UCF for
2024-25, and Georgia Tech's own 2024-25 box scores still list him,
`didNotPlay: true`, every single game. Since Games Played is just "count
of PlayerGameStats rows for this player+season" (see
build_full_season_agg() in run_d1_scrape.py), those ghost rows inflated
Games Played for EVERY player who had one -- not just the 4 transfer
cases that were extreme enough to exceed a real season's game count and
get caught by a sanity check. Measured directly against the real
MensSummitTPE.xlsx: 204,389 of 554,795 PlayerGameStats rows (36.8%) were
all-zero-stat, zero-minute rows.

The scrapers themselves are already fixed (both now skip
`did_not_play` box-score entries before writing a row), so this
problem won't recur on future runs. This script is the one-time fix
for data that was already written before that fix existed.

WHAT IT DOES
------------
1. Streams PlayerGameStats (read-only, low memory) and drops every row
   where Min AND every counting stat (FG/3FG/FT makes+attempts,
   Rebound, Foul, Ast, To, Blk, Stl, Points) are all zero -- the exact
   signature of a `didNotPlay: true` box-score entry. A real (even
   scoreless) appearance always has Min > 0, so this never touches a
   legitimate garbage-time row.
2. Rebuilds PlayerSeasons the same way run_d1_scrape.py's
   finalize_players() already does: Games Played/Games Started
   recomputed from the cleaned PlayerGameStats, Height/Class carried
   over from the existing PlayerSeasons row (this script doesn't call
   ESPN, so it can't refresh those). A (player_id, season) that turns
   out to have had ONLY ghost rows -- the player never actually
   appeared in a real game that season -- correctly disappears from
   PlayerSeasons entirely, rather than showing a fabricated "0 games"
   or a leftover wrong-team row.
3. Copies every other sheet through unchanged, trimming any fully-blank
   trailing rows found past a sheet's real data (pure size cleanup,
   zero data change -- only rows confirmed to have no data anywhere in
   them are ever dropped).
4. Saves atomically via xlsx_io.save_with_retry (temp file + os.replace,
   same safety net run_d1_scrape.py's own checkpoint saves use).

TWO EARLIER VERSIONS OF THIS SCRIPT DIDN'T WORK -- both bugs are worth
knowing about if this one also has trouble on your machine:

  v1 streamed everything through an openpyxl write_only workbook (each
  worksheet buffered into its own temp file as you .append() to it,
  over the whole build). That hit a real Windows/openpyxl bug on a
  live run here: by the time .save() went to archive that temp file,
  Windows reported it gone (FileNotFoundError), before any data was
  written. The temp file sat around, unused, for the entire multi-
  minute build -- a long window for something to interfere with it.

  v2 avoided write_only mode by loading the *entire* existing workbook
  in normal read+write mode (openpyxl.load_workbook without
  read_only/write_only) and editing it in place -- the same way
  run_d1_scrape.py itself opens the file. That's provably reliable on
  this exact machine (the real scrape ran for hours doing exactly
  this) but it's memory-heavy: loading a workbook this size in that
  mode costs several GB just to open, before doing anything else.

  This version (v3) splits the difference: it reads the ORIGINAL file
  through a read-only (streaming, low-memory) handle -- never loading
  the full thing into memory -- and writes the cleaned result into a
  brand-new, freshly-created (not loaded from disk) normal workbook.
  A fresh normal workbook only touches the temp-file mechanism once,
  briefly, inside the final save() call -- the same call
  run_d1_scrape.py's own checkpoint saves already make successfully,
  many times, on this machine -- rather than holding a temp file open
  across the whole build like write_only mode does. Building up to
  ~350K rows x 29 columns this way peaks at roughly 3GB, not several
  times that.

USAGE
-----
    python -m scrapers.fix_dnp_ghost_rows MensSummitTPE.xlsx
    python -m scrapers.fix_dnp_ghost_rows WomensSummitTPE.xlsx

Run from the project root (same reason the scrapers themselves need
`-m`: this imports scrapers.xlsx_io). Safe to run on either workbook --
nothing here is men's/women's-specific. Close the file in Excel first
if it's open. Takes a couple of minutes; prints progress as it streams
PlayerGameStats.
"""

import sys
import time

import openpyxl

from scrapers.xlsx_io import save_with_retry

STAT_FIELDS = [
    "Min", "FG Made", "FG Attempt", "3FG M", "3FG A", "FT M", "FT A",
    "Rebound", "Foul", "Ast", "To", "Blk", "Stl", "Points",
]

# Only bother trimming a copied sheet if it has more than this many
# fully-blank trailing rows -- avoids a second scan pass on sheets that
# don't have the issue, and never touches real data (only rows after
# the last one with any non-None value, found by an explicit scan, are
# ever dropped).
TRIM_THRESHOLD = 500


def header_map(ws):
    """{header_name: 0-based index} from row 1, for values_only tuples."""
    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header = next(rows)
    return {n: i for i, n in enumerate(header) if n is not None}, header


def last_nonblank_row(ws):
    last = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v is not None for v in row):
            last = i
    return last


def main():
    if len(sys.argv) != 2:
        print("Usage: python -m scrapers.fix_dnp_ghost_rows <path-to-workbook.xlsx>")
        sys.exit(1)
    path = sys.argv[1]

    t0 = time.time()
    src = openpyxl.load_workbook(path, read_only=True, data_only=False)
    print(f"Opened {path} (read-only pass) -- sheets: {src.sheetnames} t={time.time() - t0:.1f}s")

    for required in ("PlayerGameStats", "PlayerSeasons", "Teams"):
        if required not in src.sheetnames:
            print(f"ERROR: no '{required}' sheet in this workbook -- is this the right file?")
            sys.exit(1)

    # --- Small lookups needed for the PlayerSeasons rebuild ---
    seasons_ws = src["PlayerSeasons"]
    sh_old, sheader_old = header_map(seasons_ws)
    old_extra = {}  # (player_id, season) -> (height, class)
    for r in seasons_ws.iter_rows(min_row=2, values_only=True):
        pid = r[sh_old["Player ID"]] if "Player ID" in sh_old else None
        season = r[sh_old["Season"]] if "Season" in sh_old else None
        if pid is None or season is None:
            continue
        old_extra[(pid, season)] = (
            r[sh_old["Height"]] if "Height" in sh_old else None,
            r[sh_old["Class"]] if "Class" in sh_old else None,
        )

    teams_ws = src["Teams"]
    th, _ = header_map(teams_ws)
    team_name_by_id, team_div_by_id = {}, {}
    for r in teams_ws.iter_rows(min_row=2, values_only=True):
        tid = r[th["Team ID"]] if "Team ID" in th else None
        if tid is None:
            continue
        team_name_by_id[tid] = r[th["Team"]] if "Team" in th else None
        team_div_by_id[tid] = r[th["Division"]] if "Division" in th else None

    print(f"Lookups built ({len(team_name_by_id)} teams, {len(old_extra)} existing "
          f"PlayerSeasons rows) t={time.time() - t0:.1f}s")

    # --- Set up the destination: a brand-new, freshly-created workbook
    # (NOT loaded from disk, NOT write_only -- see the module docstring
    # for why that combination matters). ---
    dst = openpyxl.Workbook()
    dst.remove(dst.active)  # drop the default blank sheet
    sheet_order = src.sheetnames
    new_sheets = {name: dst.create_sheet(name) for name in sheet_order}

    pgs_ws = src["PlayerGameStats"]
    gh, gheader = header_map(pgs_ws)
    stat_idx = [gh[f] for f in STAT_FIELDS if f in gh]
    missing_stats = [f for f in STAT_FIELDS if f not in gh]
    if missing_stats:
        print(f"WARNING: PlayerGameStats is missing expected column(s) {missing_stats} -- "
              f"continuing, but the DNP check is weaker without them.")
    col_pid, col_team = gh["Player ID"], gh["Team ID"]
    col_season, col_gs = gh["Season"], gh["GS"]
    col_pos = gh.get("Position")
    col_div = gh.get("Division")

    new_pgs = new_sheets["PlayerGameStats"]
    new_pgs.append(gheader)

    agg = {}
    total = kept = removed = 0
    for row in pgs_ws.iter_rows(min_row=2, values_only=True):
        total += 1
        if stat_idx and all((row[i] or 0) == 0 for i in stat_idx):
            removed += 1
            continue
        kept += 1
        new_pgs.append(row)

        pid, season = row[col_pid], row[col_season]
        if pid and season:
            team_id, gs = row[col_team], row[col_gs]
            position = row[col_pos] if col_pos is not None else None
            division = row[col_div] if col_div is not None else None
            key = (pid, season)
            entry = agg.setdefault(key, {"team_id": team_id, "division": division,
                                           "position": position, "games": 0, "starts": 0})
            entry["team_id"] = team_id
            if division:
                entry["division"] = division
            if position:
                entry["position"] = position
            entry["games"] += 1
            if gs == "Yes":
                entry["starts"] += 1

        if total % 100000 == 0:
            print(f"  ...{total} rows processed, kept={kept}, removed={removed}, "
                  f"t={time.time() - t0:.1f}s")

    pct = (100 * removed / total) if total else 0
    print(f"PlayerGameStats: {total} rows -> kept {kept}, removed {removed} DNP ghost "
          f"rows ({pct:.2f}%) t={time.time() - t0:.1f}s")

    # --- Rebuild PlayerSeasons from the cleaned aggregation ---
    new_seasons = new_sheets["PlayerSeasons"]
    new_seasons.append(sheader_old)
    col_map = {n: i for i, n in enumerate(sheader_old) if n is not None}
    season_rows = 0
    for (player_id, season), a in sorted(agg.items(), key=lambda x: (str(x[0][0]), str(x[0][1]))):
        old_height, old_class = old_extra.get((player_id, season), (None, None))
        row_values = [None] * len(sheader_old)

        def set_col(name, val):
            if name in col_map:
                row_values[col_map[name]] = val

        set_col("Player ID", player_id)
        set_col("Season", season)
        set_col("Team ID", a["team_id"])
        set_col("Team Name", team_name_by_id.get(a["team_id"]))
        set_col("Division", a["division"] or team_div_by_id.get(a["team_id"]))
        set_col("Position", a["position"])
        set_col("Height", old_height)
        set_col("Class", old_class)
        set_col("Games Played", a["games"])
        set_col("Games Started", a["starts"])
        set_col("Finished", "Yes")
        new_seasons.append(row_values)
        season_rows += 1

    dropped_seasons = len(old_extra) - season_rows
    print(f"PlayerSeasons rebuilt: {season_rows} rows (was {len(old_extra)}; {dropped_seasons} "
          f"player-seasons had ONLY ghost rows and are correctly gone now) "
          f"t={time.time() - t0:.1f}s")

    # --- Copy every other sheet through unchanged, trimming dead
    # trailing rows if a sheet has a lot of them. ---
    for name in sheet_order:
        if name in ("PlayerGameStats", "PlayerSeasons"):
            continue
        src_ws = src[name]
        dst_ws = new_sheets[name]
        last = last_nonblank_row(src_ws)
        total_rows = getattr(src_ws, "max_row", last) or last
        max_row = last if last and (total_rows - last) > TRIM_THRESHOLD else None
        iterator = (src_ws.iter_rows(values_only=True, max_row=max_row) if max_row
                    else src_ws.iter_rows(values_only=True))
        n = 0
        for row in iterator:
            dst_ws.append(row)
            n += 1
        note = f" (trimmed {total_rows - last} blank trailing rows)" if max_row else ""
        print(f"Copied '{name}': {n} rows{note} t={time.time() - t0:.1f}s")

    src.close()
    save_with_retry(dst, path)
    print(f"Saved {path} in place t={time.time() - t0:.1f}s")


if __name__ == "__main__":
    main()
