"""Automated backfill for missing Height/Class, using ESPN data instead
of hand-entering it.

SCOPE: by default this only touches PlayerSeasons rows in the CURRENT
season (latest season found, or --season to override) -- for both
Height and Class. Pass --height-scope all-seasons if you ever want
Height backfilled for older seasons too (Height doesn't change person
to person, so this is safe to widen; Class can never be widened past
the current season, there's no source for it -- see below).

WHY TWO DIFFERENT STRATEGIES FOR HEIGHT VS. CLASS
---------------------------------------------------
Height is a fixed attribute of the person -- it doesn't change season to
season -- so it CAN be backfilled for any season (via --height-scope
all-seasons), from three sources, cheapest first:
  1. Already present on the Players sheet for that Player ID (free, no
     network call -- also fixes the fact that Players' Height and
     PlayerSeasons' Height don't currently sync with each other; this
     pass syncs them once).
  2. That player's CURRENT team roster (espn.get_roster) -- only works
     for players still on a current D1 roster.
  3. ESPN's per-athlete profile endpoint (espn.get_athlete_height) --
     works for ANY athlete ID regardless of current roster status
     (confirmed in espn_client.py's own docstring: it still returns a
     departed/graduated player's height when the roster endpoint has
     no record of them at all). This is the one that can fill in
     Height for players from past seasons who've since transferred
     out or graduated -- only relevant with --height-scope all-seasons,
     since the default (current-season-only) scope never needs it for
     anyone not already reachable via source 1 or 2.

Class (Freshman/Sophomore/etc.) is NOT a fixed attribute -- it's tied to
a specific season -- so it can ONLY ever be backfilled for the CURRENT
season, and only from that team's current roster (ESPN has no season-
scoped or per-athlete class endpoint; a past season's Class can't be
inferred from a player's current standing). Past-season Class gaps
always need a human/other source -- this script won't touch them, and
won't guess, regardless of any flag.

Even after this runs, some players will still be missing Height and/or
Class -- ESPN's own roster/profile data is genuinely incomplete for
some players (walk-ons, incomplete profiles). Whatever's still blank
after this needs a different source or manual entry.

USAGE
-----
    python -m scrapers.backfill_height_class MensSummitTPE.xlsx --sport men
    python -m scrapers.backfill_height_class WomensSummitTPE.xlsx --sport women

Optional: --season "2025-26" to override which season counts as
"current" (default: the latest season found on PlayerSeasons).
Optional: --height-scope all-seasons to widen Height backfilling to
every season instead of just the current one (default: current-only).

Makes real ESPN calls -- one per team whose current-season roster is
needed, plus one per player still missing Height after the free
lookups. Expect it to take a few minutes depending on how much is
missing; it prints progress and rate-limits itself the same way the
main scrapers do.
"""

import argparse
import sys
import time

import openpyxl

from scrapers.xlsx_io import save_with_retry

TRIM_THRESHOLD = 500  # same convention as fix_dnp_ghost_rows.py


def header_map(ws):
    """{header_name: 0-based index} from row 1, for values_only tuples."""
    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header = next(rows)
    return {n: i for i, n in enumerate(header) if n is not None}, header


def last_nonblank_row(ws):
    last = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v is not None for v in row):
            last = i
    return last


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    parser.add_argument("--sport", choices=["men", "women"], required=True)
    parser.add_argument("--season", default=None,
                         help="Which season is 'current' (e.g. 2025-26). Default: "
                              "the latest season found on PlayerSeasons.")
    parser.add_argument("--height-scope", choices=["current-only", "all-seasons"],
                         default="current-only",
                         help="'current-only' (default): only fill Height on "
                              "current-season PlayerSeasons rows (and Players). "
                              "'all-seasons': also backfill Height for older "
                              "seasons via the per-athlete ESPN lookup -- slower, "
                              "since that means one extra ESPN call per player "
                              "not already covered by a current roster.")
    args = parser.parse_args()

    if args.sport == "men":
        import scrapers.espn_client_mbb as espn
    else:
        import scrapers.espn_client as espn

    t0 = time.time()
    src = openpyxl.load_workbook(args.path, read_only=True, data_only=False)
    print(f"Opened {args.path} -- sheets: {src.sheetnames} t={time.time() - t0:.1f}s")

    for required in ("Teams", "Players", "PlayerSeasons"):
        if required not in src.sheetnames:
            print(f"ERROR: no '{required}' sheet -- is this the right file?")
            sys.exit(1)

    # --- Teams: Team ID -> ESPN Team ID ---
    teams_ws = src["Teams"]
    th, _ = header_map(teams_ws)
    espn_id_by_team = {}
    for r in teams_ws.iter_rows(min_row=2, values_only=True):
        tid = r[th["Team ID"]]
        if tid is not None:
            espn_id_by_team[tid] = r[th.get("ESPN Team ID")] if "ESPN Team ID" in th else None

    # --- Players: Player ID -> {external_id, height, class} ---
    players_ws = src["Players"]
    ph, pheader = header_map(players_ws)
    players_by_id = {}
    for r in players_ws.iter_rows(min_row=2, values_only=True):
        pid = r[ph["Player ID"]]
        if pid is not None:
            players_by_id[pid] = {
                "external_id": r[ph["External ID"]],
                "height": r[ph["Height"]],
                "class": r[ph["Class"]],
            }

    # --- PlayerSeasons: read fully (small sheet) ---
    seasons_ws = src["PlayerSeasons"]
    sh, sheader = header_map(seasons_ws)
    season_rows = list(seasons_ws.iter_rows(min_row=2, values_only=True))
    all_seasons = sorted({r[sh["Season"]] for r in season_rows if r[sh["Season"]]})
    current_season = args.season or (all_seasons[-1] if all_seasons else None)
    print(f"Seasons on file: {all_seasons}")
    print(f"Treating '{current_season}' as current -- Class is only ever backfilled "
          f"for that season. Height scope: {args.height_scope}"
          f"{' (that season only)' if args.height_scope == 'current-only' else ' (all seasons)'}.")

    # --- Which teams' current rosters do we actually need? ---
    teams_needed = set()
    for r in season_rows:
        if r[sh["Season"]] != current_season:
            continue
        missing_height = not r[sh["Height"]]
        missing_class = not r[sh["Class"]]
        if missing_height or missing_class:
            tid = r[sh["Team ID"]]
            if tid in espn_id_by_team and espn_id_by_team[tid]:
                teams_needed.add(tid)

    print(f"Fetching current rosters for {len(teams_needed)} teams ...")
    roster_by_team = {}
    for i, tid in enumerate(sorted(teams_needed), 1):
        espn_id = espn_id_by_team[tid]
        try:
            roster = espn.get_roster(espn_id)
            roster_by_team[tid] = {
                str(p["espn_athlete_id"]): p for p in roster if p.get("espn_athlete_id")
            }
        except Exception as exc:
            print(f"  [!] roster fetch failed for team {tid} (ESPN {espn_id}): {exc}")
            roster_by_team[tid] = {}
        if i % 25 == 0:
            print(f"  ...{i}/{len(teams_needed)} rosters fetched, t={time.time() - t0:.1f}s")
        time.sleep(0.3)
    print(f"Rosters fetched t={time.time() - t0:.1f}s")

    # A flat index across every fetched roster, keyed by external ID --
    # good enough here since ESPN athlete IDs are globally unique (not
    # scoped to a team), so we don't need to know which team a player's
    # entry came from to look them up.
    roster_flat = {}
    for roster in roster_by_team.values():
        roster_flat.update(roster)

    # --- Figure out what's missing ---
    missing_height_players = set()
    for r in season_rows:
        if args.height_scope == "current-only" and r[sh["Season"]] != current_season:
            continue
        pid = r[sh["Player ID"]]
        if pid is not None and not r[sh["Height"]]:
            missing_height_players.add(pid)
    if args.height_scope == "all-seasons":
        for pid, info in players_by_id.items():
            if not info["height"]:
                missing_height_players.add(pid)

    missing_class_players = set()
    for r in season_rows:
        if r[sh["Season"]] == current_season:
            pid = r[sh["Player ID"]]
            if pid is not None and not r[sh["Class"]]:
                missing_class_players.add(pid)

    height_scope_note = current_season if args.height_scope == "current-only" else "any season"
    print(f"{len(missing_height_players)} distinct players missing Height ({height_scope_note}); "
          f"{len(missing_class_players)} distinct players missing Class in {current_season}.")

    # --- Resolve Height ---
    height_updates = {}
    height_sources = {"players_sheet": 0, "roster": 0, "athlete_endpoint": 0, "still_missing": 0}
    athlete_calls = 0
    for n, pid in enumerate(sorted(missing_height_players), 1):
        info = players_by_id.get(pid, {})
        height = info.get("height")
        if height:
            height_updates[pid] = height
            height_sources["players_sheet"] += 1
            continue

        ext_id = info.get("external_id")
        roster_entry = roster_flat.get(str(ext_id)) if ext_id else None
        if roster_entry and roster_entry.get("height"):
            height_updates[pid] = roster_entry["height"]
            height_sources["roster"] += 1
            continue

        if ext_id:
            try:
                h = espn.get_athlete_height(ext_id)
            except Exception:
                h = None
            athlete_calls += 1
            time.sleep(0.25)
            if athlete_calls % 50 == 0:
                print(f"  ...{athlete_calls} per-athlete height calls made, "
                      f"t={time.time() - t0:.1f}s")
            if h:
                height_updates[pid] = h
                height_sources["athlete_endpoint"] += 1
                continue

        height_sources["still_missing"] += 1

    print(f"Height sources: {height_sources} t={time.time() - t0:.1f}s")

    # --- Resolve Class (current season only, roster-sourced only) ---
    class_updates = {}
    class_sources = {"roster": 0, "still_missing": 0}
    for pid in missing_class_players:
        info = players_by_id.get(pid, {})
        ext_id = info.get("external_id")
        roster_entry = roster_flat.get(str(ext_id)) if ext_id else None
        if roster_entry and roster_entry.get("class_year"):
            class_updates[pid] = roster_entry["class_year"]
            class_sources["roster"] += 1
        else:
            class_sources["still_missing"] += 1

    print(f"Class sources ({current_season} only): {class_sources} t={time.time() - t0:.1f}s")

    if not height_updates and not class_updates:
        print("Nothing found to fill in -- exiting without writing anything.")
        return

    # --- Write: fresh normal workbook, copy everything through, patch
    # Players + PlayerSeasons only (values already loaded, no need to
    # touch the big PlayerGameStats sheet's data at all here). ---
    dst = openpyxl.Workbook()
    dst.remove(dst.active)
    sheet_order = src.sheetnames
    new_sheets = {name: dst.create_sheet(name) for name in sheet_order}

    new_players = new_sheets["Players"]
    new_players.append(pheader)
    p_height_filled = p_class_filled = 0
    for r in players_ws.iter_rows(min_row=2, values_only=True):
        pid = r[ph["Player ID"]]
        row = list(r)
        if pid in height_updates and not row[ph["Height"]]:
            row[ph["Height"]] = height_updates[pid]
            p_height_filled += 1
        if pid in class_updates and not row[ph["Class"]]:
            row[ph["Class"]] = class_updates[pid]
            p_class_filled += 1
        new_players.append(row)
    print(f"Players: filled Height on {p_height_filled} rows, Class on {p_class_filled} rows.")

    new_seasons = new_sheets["PlayerSeasons"]
    new_seasons.append(sheader)
    s_height_filled = s_class_filled = 0
    for r in season_rows:
        pid = r[sh["Player ID"]]
        season = r[sh["Season"]]
        row = list(r)
        height_in_scope = season == current_season or args.height_scope == "all-seasons"
        if height_in_scope and not row[sh["Height"]] and pid in height_updates:
            row[sh["Height"]] = height_updates[pid]
            s_height_filled += 1
        if season == current_season and not row[sh["Class"]] and pid in class_updates:
            row[sh["Class"]] = class_updates[pid]
            s_class_filled += 1
        new_seasons.append(row)
    print(f"PlayerSeasons: filled Height on {s_height_filled} rows, "
          f"Class on {s_class_filled} rows.")

    for name in sheet_order:
        if name in ("Players", "PlayerSeasons"):
            continue
        src_ws = src[name]
        dst_ws = new_sheets[name]
        last = last_nonblank_row(src_ws)
        total_rows = getattr(src_ws, "max_row", last) or last
        max_row = last if last and (total_rows - last) > TRIM_THRESHOLD else None
        iterator = (src_ws.iter_rows(values_only=True, max_row=max_row) if max_row
                    else src_ws.iter_rows(values_only=True))
        n = 0
        for row in iterator:
            dst_ws.append(row)
            n += 1
        note = f" (trimmed {total_rows - last} blank trailing rows)" if max_row else ""
        print(f"Copied '{name}': {n} rows{note} t={time.time() - t0:.1f}s")

    src.close()
    save_with_retry(dst, args.path)
    print(f"Saved {args.path} in place t={time.time() - t0:.1f}s")


if __name__ == "__main__":
    main()
