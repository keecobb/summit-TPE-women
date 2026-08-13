"""Repairs the corruption caused by backfill_pgs_team_300.py's team-ID bug.
 
WHAT HAPPENED: backfill_pgs_team_300.py resolved which side of an ESPN box
score was "us" vs "the opponent" by comparing an ESPN team id (e.g. "245")
against opp_team_id, which is actually an INTERNAL Team ID (e.g. 299) --
two different id spaces that essentially never match. So for every one of
the games it backfilled, BOTH box-score sides (our real players AND every
opponent's players) got written to PlayerGameStats -- and, for any
opponent athlete not already known to the workbook, to Players too --
under --team-id instead of the opponent's real team.  That's why the
Players sheet ended up with ~360 rows tagged Team 300 when the real roster
is ~15-20 players: roughly 340 of those are actually opposing players from
84 different games, all misfiled under Corpus Christi.
 
backfill_pgs_team_300.py itself has now been fixed (it used to compare
espn_team_key against opp_team_id; it now compares espn_team_key against
our own ESPN Team ID, which is the only comparison that was ever going to
work). This script cleans up the damage the buggy version already wrote.
 
HOW THE FIX WORKS: for every game --team-id played, this re-fetches the
real ESPN box score and uses OUR OWN ESPN Team ID (unambiguous) to figure
out which box-score side is genuinely ours. Any existing PlayerGameStats
row tagged --team-id whose player doesn't appear on our real side gets:
  1. deleted from under --team-id (it never belonged there), and
  2. (unless already present) re-written under the real opponent's
     internal Team ID, using the same box-score data -- so the opponent's
     own record for that game gets completed instead of just leaving a
     second gap.
Then any Players-sheet row left with zero real PlayerGameStats under
--team-id has its Team ID corrected to the real team the data shows it
belongs to.
 
This is READ-ONLY (network calls only) until you pass --apply. Run once
without --apply first and read the report before applying anything.
 
Usage:
    python fix_team_300_backfill_bug.py --team-id 300                 # dry run
    python fix_team_300_backfill_bug.py --team-id 300 --apply          # fix it
    python fix_team_300_backfill_bug.py --team-id 300 --apply --no-reattribute
        # (delete the bad rows but don't bother re-filing them under the
        #  real opponent team -- use this if you'd rather backfill the
        #  opponents separately/later)
 
Requires: openpyxl, requests.
"""
 
import argparse
import os
import sys
import time
from collections import defaultdict
 
import requests
 
from backfill_pgs_team_300 import get_boxscore, header_map, save_with_retry
 
 
def load_games_for_team(games_ws, ghmap, team_id):
    home_col, away_col = ghmap["Home Team ID"] - 1, ghmap["Away Team ID"] - 1
    home_name_col, away_name_col = ghmap["Home Team"] - 1, ghmap["Away Team"] - 1
    gid_col, season_col, date_col = ghmap["Game ID"] - 1, ghmap["Season"] - 1, ghmap["Date"] - 1
    games = {}
    for row in games_ws.iter_rows(min_row=2, values_only=True):
        gid = row[gid_col]
        if not gid:
            continue
        home_id, away_id = row[home_col], row[away_col]
        if home_id != team_id and away_id != team_id:
            continue
        if home_id == team_id:
            opp_id, opp_name = away_id, row[away_name_col]
        else:
            opp_id, opp_name = home_id, row[home_name_col]
        games[str(gid)] = dict(
            season=row[season_col], date=row[date_col],
            opp_id=opp_id, opp_name=opp_name,
        )
    return games
 
 
def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--path", default="WomensSummitTPE.xlsx")
    parser.add_argument("--team-id", type=int, required=True)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--no-reattribute", action="store_true",
                         help="Delete bad rows but don't re-file them under the real opponent team.")
    args = parser.parse_args()
    team_id = args.team_id
    reattribute = not args.no_reattribute
 
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required (pip install openpyxl).")
 
    print(f"Scanning {args.path} (read-only) ...")
    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
 
    teams_ws = wb["Teams"]
    thmap = header_map(teams_ws)
    my_espn_id = None
    team_name = None
    for row in teams_ws.iter_rows(min_row=2, values_only=True):
        if row[thmap["Team ID"] - 1] == team_id:
            team_name = row[thmap["Team"] - 1]
            espn_val = row[thmap["ESPN Team ID"] - 1]
            my_espn_id = str(espn_val) if espn_val is not None else None
            break
    if not my_espn_id:
        wb.close()
        sys.exit(f"Team {team_id} has no ESPN Team ID on the Teams sheet -- can't safely verify.")
    print(f"Team {team_id}: {team_name!r} (ESPN Team ID {my_espn_id})")
 
    games_ws = wb["Games"]
    ghmap = header_map(games_ws)
    games_for_team = load_games_for_team(games_ws, ghmap, team_id)
 
    pgs_ws = wb["PlayerGameStats"]
    pghmap = header_map(pgs_ws)
    pid_col, pgs_team_col, pgs_game_col = (
        pghmap["Player ID"] - 1, pghmap["Team ID"] - 1, pghmap["Game ID"] - 1,
    )
    # every existing PGS row for our team, grouped by game -- these are the
    # rows we need to re-verify against a fresh box score
    rows_by_game = defaultdict(list)   # gid -> [(row_index, player_id), ...]
    # every existing PGS row for ANY team+game, so we can tell whether the
    # real opponent already has a row for a given player+game (avoid dupes)
    existing_by_team_game_player = set()   # (team_id, gid, player_id)
    for i, row in enumerate(pgs_ws.iter_rows(min_row=2, values_only=True), start=2):
        t, g, pid = row[pgs_team_col], row[pgs_game_col], row[pid_col]
        if not g:
            continue
        g = str(g)
        existing_by_team_game_player.add((t, g, pid))
        if t == team_id and g in games_for_team:
            rows_by_game[g].append((i, pid))
 
    players_ws = wb["Players"]
    phmap = header_map(players_ws)
    pid_to_ext, ext_to_pid, pid_to_current_team = {}, {}, {}
    max_pid = 0
    for row in players_ws.iter_rows(min_row=2, values_only=True):
        pid = row[phmap["Player ID"] - 1]
        ext = row[phmap["External ID"] - 1]
        pid_to_current_team[pid] = row[phmap["Team ID"] - 1]
        if isinstance(pid, int):
            max_pid = max(max_pid, pid)
        if ext:
            pid_to_ext[pid] = str(ext)
            ext_to_pid[str(ext)] = pid
    wb.close()
 
    suspect_games = sorted(g for g in rows_by_game if rows_by_game[g])
    print(f"\n{len(suspect_games)} games currently have PlayerGameStats rows tagged Team {team_id}.")
    print("Re-fetching each box score from ESPN to verify which players actually belong to us ...")
 
    bad_rows = []          # (row_index, gid, player_id, opp_id, opp_name)
    reattribute_writes = []  # (gid, opp_id, opp_name, date, season, box_row_for_this_player)
    good_count = 0
    unverifiable = []
    for n, gid in enumerate(suspect_games, 1):
        info = games_for_team[gid]
        try:
            box = get_boxscore(gid)
        except RuntimeError as exc:
            print(f"  [!] game {gid}: box score fetch failed ({exc}) -- skipping, can't verify")
            unverifiable.append(gid)
            continue
        if my_espn_id not in box:
            print(f"  [!] game {gid}: box score doesn't contain our own ESPN id {my_espn_id} -- skipping")
            unverifiable.append(gid)
            continue
 
        real_cc_ext_ids = {str(r["espn_athlete_id"]) for r in box[my_espn_id] if r.get("espn_athlete_id")}
        # the opponent side is whichever other key is present (there should
        # be exactly one other key in a normal 2-team box score)
        opp_rows_by_ext = {}
        for key, rows in box.items():
            if key == my_espn_id:
                continue
            for r in rows:
                if r.get("espn_athlete_id"):
                    opp_rows_by_ext[str(r["espn_athlete_id"])] = r
 
        for row_idx, pid in rows_by_game[gid]:
            ext_id = pid_to_ext.get(pid)
            if ext_id in real_cc_ext_ids:
                good_count += 1
                continue
            bad_rows.append((row_idx, gid, pid, info["opp_id"], info["opp_name"]))
            if reattribute and ext_id in opp_rows_by_ext:
                key = (info["opp_id"], gid, pid)
                if key not in existing_by_team_game_player:
                    reattribute_writes.append((gid, info, opp_rows_by_ext[ext_id]))
                    existing_by_team_game_player.add(key)  # avoid re-adding twice in this pass
 
        if n % 10 == 0 or n == len(suspect_games):
            print(f"  ({n}/{len(suspect_games)}) checked")
        time.sleep(0.15)
 
    print(f"\n{good_count} PlayerGameStats rows confirmed as real Team {team_id} players.")
    print(f"{len(bad_rows)} PlayerGameStats rows are mis-attributed opponent players "
          f"(written under Team {team_id} by the backfill bug).")
    if unverifiable:
        print(f"{len(unverifiable)} games could not be verified (box score unavailable) -- left untouched: "
              f"{', '.join(unverifiable)}")
    if reattribute:
        print(f"{len(reattribute_writes)} of those will be re-filed under the real opponent team "
              f"(the rest already have a row there, or the opponent-side box data wasn't found).")
 
    if not bad_rows:
        print("\nNothing to fix.")
        return
 
    by_opp = defaultdict(int)
    for _, _, _, opp_id, opp_name in bad_rows:
        by_opp[(opp_id, opp_name)] += 1
    print("\nBad rows by real opponent team:")
    for (opp_id, opp_name), n in sorted(by_opp.items(), key=lambda x: -x[1]):
        print(f"    {opp_name} (Team {opp_id}): {n} rows")
 
    if not args.apply:
        print(f"\nDry run only -- re-run with --apply to delete these {len(bad_rows)} rows"
              + (f" and re-file {len(reattribute_writes)} of them under the real opponent." if reattribute else ".")
              + "\n(Close the workbook in Excel first.)")
        return
 
    print(f"\nOpening {args.path} (write mode) ...")
    wb = openpyxl.load_workbook(args.path)
    pgs_ws = wb["PlayerGameStats"]
    pghmap = header_map(pgs_ws)
    players_ws = wb["Players"]
    phmap = header_map(players_ws)
    teams_ws = wb["Teams"]
    thmap = header_map(teams_ws)
 
    team_id_to_division = {}
    for row in range(2, teams_ws.max_row + 1):
        tid = teams_ws.cell(row=row, column=thmap["Team ID"]).value
        if tid is not None:
            team_id_to_division[tid] = teams_ws.cell(row=row, column=thmap["Division"]).value
 
    # --- 1. delete bad rows ---
    # NOTE: calling ws.delete_rows() once per row is O(sheet size) EACH call
    # (it has to shift every row below up by one), so doing it hundreds of
    # times over a PlayerGameStats sheet with many thousands of rows across
    # every tracked team can take a very long time. Instead: read every
    # surviving data row once, wipe all data rows in a single delete_rows
    # call, then re-append just the survivors -- O(sheet size) total instead
    # of O(sheet size x rows deleted).
    bad_row_index_set = {idx for idx, *_ in bad_rows}
    print(f"Rewriting PlayerGameStats: dropping {len(bad_row_index_set)} mis-attributed rows ...")
    max_row_before = pgs_ws.max_row
    survivors = [
        row for i, row in enumerate(pgs_ws.iter_rows(min_row=2, values_only=True), start=2)
        if i not in bad_row_index_set
    ]
    if max_row_before >= 2:
        pgs_ws.delete_rows(2, max_row_before - 1)  # single pass, wipes all data rows at once
    for row in survivors:
        pgs_ws.append(row)
 
    # --- 2. re-file under the real opponent, if requested ---
    if reattribute and reattribute_writes:
        print(f"Re-filing {len(reattribute_writes)} rows under their real opponent team ...")
 
        def split_made_attempted(value):
            if not value or value == "-":
                return 0, 0
            try:
                made, attempted = value.split("-")
                return int(made), int(attempted)
            except (ValueError, AttributeError):
                return 0, 0
 
        def parse_minutes(value):
            try:
                return float(value)
            except (TypeError, ValueError):
                return 0.0
 
        def parse_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0
 
        next_row_pgs = pgs_ws.max_row + 1
        for gid, info, player_row in reattribute_writes:
            ext_id = str(player_row["espn_athlete_id"])
            pid = ext_to_pid.get(ext_id)
            if pid is None:
                print(f"  [!] game {gid}: no known Player ID for external id {ext_id} "
                      f"({player_row.get('name')}) -- skipping re-file (their PGS row was still deleted "
                      f"from Team {team_id}); create/backfill this player separately if needed.")
                continue
            opp_id = info["opp_id"]
            stats = player_row["raw_stats"]
            fgm, fga = split_made_attempted(stats.get("FG"))
            fg3m, fg3a = split_made_attempted(stats.get("3PT"))
            ftm, fta = split_made_attempted(stats.get("FT"))
            values = {
                "Player ID": pid, "Team ID": opp_id, "Opponent": team_name,
                "Opponent Team ID": team_id, "Opponent Level": team_id_to_division.get(team_id),
                "Date": info["date"], "GS": "Yes" if player_row["starter"] else "No",
                "Min": parse_minutes(stats.get("MIN")), "FG Made": fgm, "FG Attempt": fga,
                "3FG M": fg3m, "3FG A": fg3a, "FT M": ftm, "FT A": fta,
                "Rebound": parse_int(stats.get("REB")), "Foul": parse_int(stats.get("PF")),
                "Ast": parse_int(stats.get("AST")), "To": parse_int(stats.get("TO")),
                "Blk": parse_int(stats.get("BLK")), "Stl": parse_int(stats.get("STL")),
                "Points": parse_int(stats.get("PTS")), "Game ID": gid, "Season": info["season"],
                "Position": player_row["position"], "Division": team_id_to_division.get(opp_id),
            }
            for key, val in values.items():
                c = pghmap.get(key)
                if c:
                    pgs_ws.cell(row=next_row_pgs, column=c, value=val)
            next_row_pgs += 1
 
    # --- 3. fix Players-sheet Team ID for anyone left with zero real rows under team_id ---
    bad_pids_touched = {pid for _, _, pid, _, _ in bad_rows}
    # recompute which of those pids still have ANY surviving PGS row under team_id
    surviving_team_pids = set()
    for row in pgs_ws.iter_rows(min_row=2, values_only=True):
        if row[pghmap["Team ID"] - 1] == team_id:
            surviving_team_pids.add(row[pghmap["Player ID"] - 1])
 
    opp_by_pid = {}
    for _, _, pid, opp_id, _ in bad_rows:
        opp_by_pid.setdefault(pid, opp_id)  # first opponent seen; a real player only has one
 
    fixed_players = 0
    for row in range(2, players_ws.max_row + 1):
        pid = players_ws.cell(row=row, column=phmap["Player ID"]).value
        cur_team = players_ws.cell(row=row, column=phmap["Team ID"]).value
        if pid in bad_pids_touched and cur_team == team_id and pid not in surviving_team_pids:
            correct_team = opp_by_pid.get(pid)
            if correct_team is not None:
                players_ws.cell(row=row, column=phmap["Team ID"], value=correct_team)
                division = team_id_to_division.get(correct_team)
                if division is not None and phmap.get("Division"):
                    players_ws.cell(row=row, column=phmap["Division"], value=division)
                fixed_players += 1
    print(f"Corrected Team ID on {fixed_players} Players-sheet rows that had zero real games for Team {team_id}.")
 
    print("\nSaving ...")
    save_with_retry(wb, args.path)
    print(f"Done. Deleted {len(bad_row_index_set)} PlayerGameStats rows, "
          + (f"re-filed {len(reattribute_writes)} of them under the real opponent, " if reattribute else "")
          + f"corrected {fixed_players} Players rows.\n"
          "Next: run run_d1_scrape.py once more (rebuilds PlayerSeasons from the corrected "
          "PlayerGameStats), then re-run diagnose_team_300_roster.py to confirm the roster "
          "is back to ~15-20 real players, then build_cache.py.")
 
 
if __name__ == "__main__":
    main()