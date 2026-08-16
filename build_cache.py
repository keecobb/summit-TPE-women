"""Builds a summit_tpe_cache_<sport>.sqlite from a Summit TPE workbook
(WomensSummitTPE.xlsx or MensSummitTPE.xlsx -- same column shape, see
--sport below):
  - teams: one row per current-season team (rating + tier)
  - players: one row per current-season player (season stat profile)
  - team_profile: one row per current-season team, roster-derived category
    rates (rebounding/assists/blocks/steals/turnovers/shooting), used by
    /teams/{id}/needs and /teams/{id}/fits to find a team's statistical
    weaknesses and rank transfer targets against them.
  - player_history: one row per (player, season) for EVERY season in the
    workbook (not just the current one), used by /players/{id}/trajectory
    to show whether a player's production is trending up or down.
  - meta: season label + league mean/std for Rat and for each team_profile
    category (needed to turn a team's raw category rate into "how weak is
    this, relative to the rest of the league").

This is the STATIC half of the transfer calculator -- a player's own
season stats and a team's own category profile don't change based on what
matchup you're evaluating, so they're computed once here and cached,
instead of being recomputed on every API request. The FLUID half (target
team, minutes, which player(s) you're comparing) is computed live per-
request by projection.py, reading from this cache.

Re-run this once per season (or whenever the underlying workbook's box
scores update) to refresh the cache; the API always reads whatever's
currently in the matching summit_tpe_cache_<sport>.sqlite for that sport
(see api.py's get_conn(sport)).

--sport is required and picks BOTH the tier-classification fallback (see
classify_tier() in summit_calc.py -- men's and women's conferences aren't
scored the same way) and the default --out filename, so a plain
`--path MensSummitTPE.xlsx --sport men` can't accidentally overwrite the
women's cache (or vice versa) by forgetting --out.

Usage:
    python build_cache.py --path WomensSummitTPE.xlsx --sport women
    python build_cache.py --path MensSummitTPE.xlsx --sport men
    python build_cache.py --path WomensSummitTPE.xlsx --sport women --out summit_tpe_cache_women.sqlite
"""
 
import argparse
import re
import sqlite3
import statistics
from collections import defaultdict

import openpyxl
 
from summit_calc import (
    BUCKET_WEIGHTS, EXPERIENCE_MULT, MIN_GAMES_FOR_PROFILE, MIN_TOTAL_MINUTES_FOR_PROFILE,
    POSITION_TO_BUCKET, classify_tier, close_game_weight, compute_sos, composite_game_score,
    iterative_off_def, opponent_strength_factor, per40_scale, scale_to_hoopscore,
)
 
# Category rates a team's roster is profiled on, for /teams/{id}/needs and
# /teams/{id}/fits. per40_tov is included but flagged LOWER_IS_BETTER in
# projection.py -- a team with a HIGH turnover rate has a weakness there,
# not a strength, so its z-score gets flipped at read time.
TEAM_PROFILE_STATS = ["per40_pts", "per40_reb", "per40_ast", "per40_blk", "per40_stl", "per40_tov"]
TEAM_PROFILE_PCT_STATS = ["ts_pct", "fg_pct"]
 
 
_HEIGHT_RE = re.compile(r"^(\d+)'\s*(\d+)?")


def parse_height_inches(height_text):
    """Free-text height like 6' 2\" or 5'9 -> total inches (int), for the
    Players page's height filter -- the Height column itself stays free
    text for display (not every source formats it identically), but a
    numeric form is needed to filter/sort by height at all. Returns None
    for anything that doesn't match the expected feet'inches\" shape rather
    than guessing."""
    if not height_text:
        return None
    m = _HEIGHT_RE.match(height_text.strip())
    if not m:
        return None
    feet = int(m.group(1))
    inches = int(m.group(2)) if m.group(2) else 0
    return feet * 12 + inches


def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping
 
 
# Column name(s) to look for on the Players sheet for transfer-portal
# status. Optional -- most workbooks won't have this yet (see load()'s
# "no Transfer Portal column found" note). Checked in order; first match
# wins.
TRANSFER_PORTAL_COLUMNS = ("Transfer Portal", "In Transfer Portal")
_TRUE_VALUES = {"yes", "y", "true", "t", "1"}
_FALSE_VALUES = {"no", "n", "false", "f", "0"}
 
 
def _parse_tri_bool(value):
    """Yes/No/TRUE/FALSE/1/0 (any case, or an actual bool/int) -> 1/0.
    Blank or unrecognized -> None (means "unknown", NOT "confirmed not in
    the portal" -- an important distinction for a column that mostly isn't
    populated yet)."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 1 if value else 0
    s = str(value).strip().lower()
    if s in _TRUE_VALUES:
        return 1
    if s in _FALSE_VALUES:
        return 0
    return None
 
 
def load(path):
    print(f"Opening {path} read-only for extraction ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
 
    teams_ws = wb["Teams"]
    th = header_map(teams_ws)
    teams = {}   # team_id -> dict(name, division, conference)
    for row in teams_ws.iter_rows(min_row=2, values_only=True):
        tid = row[th["Team ID"] - 1]
        if tid is None:
            continue
        teams[tid] = dict(
            name=row[th["Team"] - 1],
            division=row[th["Division"] - 1],
            conference=row[th["Conference"] - 1],
            school_level=row[th["School Level"] - 1] if "School Level" in th else None,
        )
    print(f"  Teams: {len(teams)}")
 
    games_ws = wb["Games"]
    gh = header_map(games_ws)
    games_by_season = defaultdict(list)
    game_lookup = {}
    schedule_rows = []
    for row in games_ws.iter_rows(min_row=2, values_only=True):
        gid = row[gh["Game ID"] - 1]
        if gid is None:
            continue
        season = row[gh["Season"] - 1]
        home_id, away_id = row[gh["Home Team ID"] - 1], row[gh["Away Team ID"] - 1]
        hs, as_ = row[gh["Home Score"] - 1], row[gh["Away Score"] - 1]
        margin = row[gh["Margin"] - 1]
        if home_id is not None and away_id is not None and hs is not None and as_ is not None:
            games_by_season[season].append((home_id, away_id, hs, as_))
        game_lookup[gid] = (home_id, away_id, margin, season)
        # Persisted for the team Schedule tab -- every game regardless of
        # whether it had usable scores for the rating computation above.
        winner_id = row[gh["Winner"] - 1] if "Winner" in gh else None
        # The sheet's own Winner column is blank for a small number of rows
        # that otherwise have complete scores (28 of 6,102 games checked in
        # the current-season workbook) -- fall back to deriving it from the
        # scores directly rather than leaving a real, decided game out of
        # win/loss records (conference standings, in particular).
        if winner_id is None and hs is not None and as_ is not None and hs != as_:
            winner_id = home_id if hs > as_ else away_id
        schedule_rows.append(dict(
            game_id=gid, season=season, date=row[gh["Date"] - 1],
            home_team_id=home_id, home_team_name=row[gh["Home Team"] - 1],
            away_team_id=away_id, away_team_name=row[gh["Away Team"] - 1],
            home_score=hs, away_score=as_, winner_team_id=winner_id, margin=margin,
            neutral_site=1 if row[gh["Neutral Site"] - 1] else 0,
            overtime=1 if row[gh["Overtime"] - 1] else 0,
            conference_game=1 if row[gh["conference Game"] - 1] else 0,
        ))
    print(f"  Games: by season {[(s, len(v)) for s, v in games_by_season.items()]}")
 
    ps_ws = wb["PlayerSeasons"]
    ph = header_map(ps_ws)
    player_season = {}
    players_with_any_season = set()
    for row in ps_ws.iter_rows(min_row=2, values_only=True):
        pid, season = row[ph["Player ID"] - 1], row[ph["Season"] - 1]
        if pid is None or season is None:
            continue
        player_season[(pid, season)] = dict(
            team_id=row[ph["Team ID"] - 1], division=row[ph["Division"] - 1],
            position=row[ph["Position"] - 1], class_year=row[ph["Class"] - 1],
        )
        players_with_any_season.add(pid)
    print(f"  PlayerSeasons: {len(player_season)}")
 
    players_ws = wb["Players"]
    plh = header_map(players_ws)
    portal_col_name = next((c for c in TRANSFER_PORTAL_COLUMNS if c in plh), None)
    has_height_col = "Height" in plh
    player_name = {}
    player_height = {}
    player_transfer_portal = {}
    # Team ID/Position/Class/Division straight off the Players sheet (the
    # roster/biographical sheet, one row per player) -- the LAST-RESORT
    # fallback for a current-season roster player who has NO PlayerSeasons
    # row and NO PlayerGameStats rows at all this season (a real example:
    # Sydney Fenn, Indiana F, present on the Players sheet with a current
    # Team ID but zero season/game rows anywhere in the workbook). This
    # sheet has no season column, so it's only trustworthy as "current"
    # team info -- used in main() to backfill the CURRENT season's roster
    # only, never for historical seasons (see main()'s sheet_meta param to
    # compute_season_profiles).
    player_sheet_meta = {}
    has_ext_id_col = "External ID" in plh
    player_external_id = {}
    for row in players_ws.iter_rows(min_row=2, values_only=True):
        pid = row[plh["Player ID"] - 1]
        if pid is None:
            continue
        player_name[pid] = f"{row[plh['First Name'] - 1] or ''} {row[plh['Last Name'] - 1] or ''}".strip()
        if has_height_col:
            h = row[plh["Height"] - 1]
            # Sheet stores this as free text like 6' 2" -- normalize stray
            # whitespace but otherwise pass it through as-is for display;
            # not every player has one on record, so this is often None.
            player_height[pid] = h.strip() if isinstance(h, str) and h.strip() else None
        if has_ext_id_col:
            player_external_id[pid] = row[plh["External ID"] - 1]
        if portal_col_name:
            player_transfer_portal[pid] = _parse_tri_bool(row[plh[portal_col_name] - 1])
        player_sheet_meta[pid] = dict(
            team_id=row[plh["Team ID"] - 1] if "Team ID" in plh else None,
            position=row[plh["Position"] - 1] if "Position" in plh else None,
            class_year=row[plh["Class"] - 1] if "Class" in plh else None,
            division=row[plh["Division"] - 1] if "Division" in plh else None,
        )

    # Duplicate-source cleanup: some players get scraped into the Players
    # sheet twice under two different Player IDs -- once from ESPN (a real
    # External ID, richer bio data, a full season of box scores) and once
    # from a secondary athletics-site source (Sidearm or Presto -- no
    # External ID, almost always just 1-2 box score rows). Most often seen
    # right after a transfer: a real example, Alexis Black shows up as both
    # pid 5140 (ESPN, Fordham, External ID 5107953, 22 games played) and pid
    # 8707 (Sidearm, Fordham, no External ID, 1 game played) after her
    # 2025-26 transfer there -- the new school's own site got scraped as a
    # "new" roster entry instead of being matched to her existing ESPN
    # record, so both sides of her season sat on the site as two separate
    # players. Checked against the full current workbook: every one of 17
    # same-name/same-team pairs splits cleanly into exactly one ESPN-sourced
    # row (has an External ID) and one non-ESPN row (no External ID) -- no
    # case of two real, distinct players sharing a name and team, and no
    # case where both sides of a pair have an External ID -- so this is a
    # safe, non-arbitrary signal for which side is the scrape artifact.
    # The non-ESPN side is dropped entirely, every season (not just the
    # current one), since it's a duplicate of a real person already
    # represented elsewhere, not a second player; the 1-2 box score rows it
    # carries (never overlapping the ESPN side's own games, checked directly)
    # are a small, acceptable loss against no longer showing a ghost player
    # on the site.
    name_team_groups = defaultdict(list)
    for pid, nm in player_name.items():
        tid = player_sheet_meta.get(pid, {}).get("team_id")
        name_team_groups[(nm.strip().lower(), tid)].append((pid, bool(player_external_id.get(pid))))
    duplicate_exclude_pids = set()
    for key, entries in name_team_groups.items():
        if len(entries) < 2:
            continue
        has_ext_pids = [pid for pid, has_ext in entries if has_ext]
        no_ext_pids = [pid for pid, has_ext in entries if not has_ext]
        if has_ext_pids and no_ext_pids:
            duplicate_exclude_pids.update(no_ext_pids)
    if duplicate_exclude_pids:
        for pid in duplicate_exclude_pids:
            player_name.pop(pid, None)
            player_height.pop(pid, None)
            player_sheet_meta.pop(pid, None)
            player_transfer_portal.pop(pid, None)
        for k in [k for k in player_season if k[0] in duplicate_exclude_pids]:
            player_season.pop(k, None)
        players_with_any_season.difference_update(duplicate_exclude_pids)
        print(f"  Duplicate-source players excluded (non-ESPN copy of a player who also has an ESPN "
              f"record): {len(duplicate_exclude_pids)}")

    if portal_col_name:
        n_flagged = sum(1 for v in player_transfer_portal.values() if v == 1)
        print(f"  Players: {len(player_name)} (Transfer Portal column found: {n_flagged} flagged 'Yes')")
    else:
        print(f"  Players: {len(player_name)} (no Transfer Portal column on the Players sheet yet -- "
              f"in_transfer_portal will be NULL/unknown for everyone until one is added)")
 
    pgs_ws = wb["PlayerGameStats"]
    gsh = header_map(pgs_ws)
    game_rows_by_season = defaultdict(list)
    game_log_rows = []
    for row in pgs_ws.iter_rows(min_row=2, values_only=True):
        pid = row[gsh["Player ID"] - 1]
        if pid is None or pid in duplicate_exclude_pids:
            continue
        season = row[gsh["Season"] - 1]
        game_rows_by_season[season].append(dict(
            player_id=pid, team_id=row[gsh["Team ID"] - 1], opp_team_id=row[gsh["Opponent Team ID"] - 1],
            minutes=row[gsh["Min"] - 1] or 0, fgm=row[gsh["FG Made"] - 1] or 0, fga=row[gsh["FG Attempt"] - 1] or 0,
            ftm=row[gsh["FT M"] - 1] or 0, fta=row[gsh["FT A"] - 1] or 0, reb=row[gsh["Rebound"] - 1] or 0,
            pf=row[gsh["Foul"] - 1] or 0, ast=row[gsh["Ast"] - 1] or 0, tov=row[gsh["To"] - 1] or 0,
            blk=row[gsh["Blk"] - 1] or 0, stl=row[gsh["Stl"] - 1] or 0, points=row[gsh["Points"] - 1] or 0,
            game_id=row[gsh["Game ID"] - 1],
        ))
        # Persisted verbatim (not just used transiently for the season
        # aggregate above) for player-profile game logs and opponent-level
        # split leaderboards -- Opponent Level is already denormalized onto
        # this sheet per game, so no join against teams is needed to know
        # what tier a given game's opponent was at.
        gs_val = row[gsh["GS"] - 1] if "GS" in gsh else None
        started = 1 if (isinstance(gs_val, str) and gs_val.strip().lower() in ("yes", "y", "true")) else (
            1 if gs_val is True else 0
        )
        game_log_rows.append(dict(
            player_id=pid, season=season, game_id=row[gsh["Game ID"] - 1], date=row[gsh["Date"] - 1],
            team_id=row[gsh["Team ID"] - 1], opponent_team_id=row[gsh["Opponent Team ID"] - 1],
            opponent_name=row[gsh["Opponent"] - 1] if "Opponent" in gsh else None,
            # NOTE: the sheet's "Opponent Level" column is D1/D2 (division), NOT
            # the High-Major/Mid-Major/Low-Major tier -- that would need a join against
            # teams.tier via opponent_team_id, done at query time in
            # projection.py rather than stored here (avoids a second source of
            # truth that could drift from teams.tier).
            opponent_division=row[gsh["Opponent Level"] - 1] if "Opponent Level" in gsh else None,
            started=started, minutes=row[gsh["Min"] - 1] or 0,
            points=row[gsh["Points"] - 1] or 0, rebounds=row[gsh["Rebound"] - 1] or 0,
            assists=row[gsh["Ast"] - 1] or 0, steals=row[gsh["Stl"] - 1] or 0,
            blocks=row[gsh["Blk"] - 1] or 0, turnovers=row[gsh["To"] - 1] or 0,
            fouls=row[gsh["Foul"] - 1] or 0,
            fgm=row[gsh["FG Made"] - 1] or 0, fga=row[gsh["FG Attempt"] - 1] or 0,
            tfgm=row[gsh["3FG M"] - 1] or 0 if "3FG M" in gsh else 0,
            tfga=row[gsh["3FG A"] - 1] or 0 if "3FG A" in gsh else 0,
            ftm=row[gsh["FT M"] - 1] or 0, fta=row[gsh["FT A"] - 1] or 0,
        ))
    print(f"  PlayerGameStats: by season {[(s, len(v)) for s, v in game_rows_by_season.items()]}")
    print(f"  Game logs persisted: {len(game_log_rows)}")
 
    wb.close()
    return dict(teams=teams, games_by_season=games_by_season, game_lookup=game_lookup,
                player_season=player_season, player_name=player_name, player_height=player_height,
                game_rows_by_season=game_rows_by_season,
                player_transfer_portal=player_transfer_portal, game_log_rows=game_log_rows,
                schedule_rows=schedule_rows, player_sheet_meta=player_sheet_meta,
                players_with_any_season=players_with_any_season)
 
 
def pick_latest_season(games_by_season):
    return max(games_by_season, key=lambda s: len(games_by_season[s]))
 
 
def compute_season_profiles(data, season, sheet_meta=None, sport="women"):
    """Everything derived from ONE season's games + box scores: team Off/
    Def/Rat/SoS and each player's season stat profile (same math as
    compute_derived_sheets.py). Factored out so it can be run once for the
    current season (teams/players tables) and again for every season in
    the workbook (player_history, for trajectory).

    sheet_meta: optional (pass data["player_sheet_meta"]) -- ONLY pass this
    for the current season. Backfills a third, last-resort tier of players:
    those with a Team ID on the Players sheet pointing at an active team
    this season, but NO PlayerSeasons row and NO PlayerGameStats rows at
    all (a real example: Sydney Fenn, Indiana F -- on the roster sheet,
    zero season/game rows anywhere in the workbook). The Players sheet has
    no season column, so this can only be trusted as "current roster" --
    passing it for a historical season would incorrectly place a player on
    a team she may not have been on that year.
    """
    games = data["games_by_season"].get(season, [])
    off, def_ = iterative_off_def(games)
    rat = {t: off[t] - def_[t] for t in off}
    sos = compute_sos(games, rat)
    rat_values = list(rat.values())
    league_mean_rat = statistics.mean(rat_values) if rat_values else 0.0
    league_std_rat = (statistics.pstdev(rat_values) if len(rat_values) > 1 else 1.0) or 1.0

    rows = data["game_rows_by_season"].get(season, [])
    game_lookup = data["game_lookup"]
    per_player_games = defaultdict(list)
    per_player_meta = {}

    # Games each team has REAL BOX-SCORE DATA for this season -- counted
    # from the same PlayerGameStats rows used to build player totals below,
    # not from the Games sheet's schedule/score list. The two counts aren't
    # always equal: 54 of 406 team-seasons in the current workbook have a
    # final score on the Games sheet for more games than they have box-score
    # rows for (a real example: Texas A&M-Corpus Christi shows 29 scored
    # games on its schedule but only 16 games' worth of player box scores).
    # compute_team_profiles() needs to divide summed roster totals by the
    # SAME number of games those totals actually cover -- dividing by the
    # schedule's (larger) game count silently understates every per-game
    # team rate whenever box-score coverage is incomplete, which is exactly
    # the bug a user caught by comparing the Stats Breakdown tab to what a
    # team actually did on the court.
    team_boxscore_games = defaultdict(set)
    for r in rows:
        team_boxscore_games[r["team_id"]].add(r["game_id"])
    team_boxscore_games = {tid: len(gids) for tid, gids in team_boxscore_games.items()}
 
    for r in rows:
        pid = r["player_id"]
        ps = data["player_season"].get((pid, season))
        if ps is None:
            continue
        position = ps["position"]
        bucket = POSITION_TO_BUCKET.get(position, "ALL")
        minutes = max(r["minutes"], 1)
        scale = per40_scale(minutes)
        ftmiss = r["fta"] - r["ftm"]
        raw = composite_game_score(bucket, r["points"], r["fgm"], r["fga"], ftmiss,
                                    r["reb"], r["ast"], r["stl"], r["blk"], r["tov"], r["pf"])
        composite_per40 = raw * scale
        opp_rat = rat.get(r["opp_team_id"], league_mean_rat)
        opp_factor = opponent_strength_factor(opp_rat, league_mean_rat, 2 * league_std_rat)
        adjusted_value = composite_per40 * opp_factor
 
        margin = None
        gl = game_lookup.get(r["game_id"])
        if gl:
            home_id, away_id, gmargin, _ = gl
            if gmargin is not None:
                if r["team_id"] == home_id:
                    margin = gmargin
                elif r["team_id"] == away_id:
                    margin = -gmargin
        cgw = close_game_weight(margin) if margin is not None else 1.0
        reliability = max(0.3, min(1.0, minutes / 15.0))
        weight = cgw * reliability
 
        per_player_games[pid].append(dict(
            value=adjusted_value, weight=weight,
            minutes=minutes, points=r["points"], reb=r["reb"], ast=r["ast"], fga=r["fga"], fta=r["fta"],
            blk=r["blk"], stl=r["stl"], tov=r["tov"], fgm=r["fgm"],
        ))
        if pid not in per_player_meta:
            per_player_meta[pid] = dict(team_id=ps["team_id"], division=ps["division"],
                                         position=position, class_year=ps["class_year"],
                                         in_transfer_portal=data["player_transfer_portal"].get(pid))
 
    season_raw = {}
    for pid, glist in per_player_games.items():
        wsum = sum(g["weight"] for g in glist)
        sraw = (sum(g["value"] * g["weight"] for g in glist) / wsum) if wsum > 0 else (sum(g["value"] for g in glist) / len(glist))
        mult = EXPERIENCE_MULT.get(per_player_meta[pid]["class_year"], 1.0)
        season_raw[pid] = sraw * mult
    season_hs, season_hs_raw, _, _ = scale_to_hoopscore(season_raw)
 
    # Every player who appears in a game log this season gets a row --
    # INCLUDING those below the MIN_GAMES_FOR_PROFILE / MIN_TOTAL_MINUTES_
    # FOR_PROFILE thresholds (bench players, players who only got run late
    # in the season, etc). These used to be dropped entirely from the
    # players/player_history tables, which meant a real rostered player
    # (e.g. someone averaging a few minutes a game) simply didn't exist
    # anywhere on the site -- not searchable, no profile page, 404 even by
    # direct ID (found via a real example: Sarah Miller, Pennsylvania G,
    # 14 games / 49 total minutes in 2025-26 -- well short of the 100-
    # minute threshold, and completely absent from the site as a result).
    # They're now included with a `thin_sample` flag so the frontend can
    # caveat the numbers (small-sample noise) instead of hiding the player.
    player_rows = []
    thin_count = 0
    covered_pids = set()
    for pid, glist in per_player_games.items():
        meta = per_player_meta[pid]
        n_games = len(glist)
        total_min = sum(g["minutes"] for g in glist)
        thin = n_games < MIN_GAMES_FOR_PROFILE or total_min < MIN_TOTAL_MINUTES_FOR_PROFILE
        if thin:
            thin_count += 1
        total_pts = sum(g["points"] for g in glist)
        total_reb = sum(g["reb"] for g in glist)
        total_ast = sum(g["ast"] for g in glist)
        total_fga = sum(g["fga"] for g in glist)
        total_fta = sum(g["fta"] for g in glist)
        total_blk = sum(g["blk"] for g in glist)
        total_stl = sum(g["stl"] for g in glist)
        total_tov = sum(g["tov"] for g in glist)
        total_fgm = sum(g["fgm"] for g in glist)
        true_shot_attempts = total_fga + 0.44 * total_fta
        ts_pct = (total_pts / (2 * true_shot_attempts)) if true_shot_attempts >= 15 else None
        fg_pct = (total_fgm / total_fga) if total_fga >= 15 else None
        player_rows.append(dict(
            player_id=pid, name=data["player_name"].get(pid, f"Player {pid}"),
            height=data["player_height"].get(pid),
            height_in=parse_height_inches(data["player_height"].get(pid)),
            team_id=meta["team_id"], division=meta["division"], position=meta["position"],
            class_year=meta["class_year"], in_transfer_portal=meta["in_transfer_portal"],
            season=season, games=n_games,
            total_minutes=total_min,
            avg_minutes=total_min / n_games, ppg=total_pts / n_games, rpg=total_reb / n_games,
            apg=total_ast / n_games, bpg=total_blk / n_games, spg=total_stl / n_games,
            topg=total_tov / n_games, ts_pct=ts_pct, fg_pct=fg_pct,
            # Raw season totals, kept on the row (not exposed via the API --
            # players/player_history only insert the named columns below)
            # purely so compute_team_profiles() can sum real team production
            # instead of averaging individual per-40 rates. See that
            # function's docstring for why the distinction matters.
            _total_pts=total_pts, _total_reb=total_reb, _total_ast=total_ast,
            _total_blk=total_blk, _total_stl=total_stl, _total_tov=total_tov,
            _total_fgm=total_fgm, _total_fga=total_fga, _true_shot_attempts=true_shot_attempts,
            per40_pts=total_pts / total_min * 40.0,
            per40_reb=total_reb / total_min * 40.0,
            per40_ast=total_ast / total_min * 40.0,
            per40_blk=total_blk / total_min * 40.0,
            per40_stl=total_stl / total_min * 40.0,
            per40_tov=total_tov / total_min * 40.0,
            hoop_score=season_hs[pid], hoop_score_raw=season_hs_raw[pid],
            thin_sample=1 if thin else 0,
        ))
        covered_pids.add(pid)

    # Players officially on this season's roster (present on the
    # PlayerSeasons sheet) but with NO game-log rows at all this season --
    # redshirts, season-long injuries, players who never got off the bench.
    # Still worth a placeholder row (name/team/position, no computed stats)
    # so they're at least findable by name and show up on their team's
    # roster, rather than being invisible.
    for (pid, s), ps in data["player_season"].items():
        if s != season or pid in covered_pids:
            continue
        player_rows.append(dict(
            player_id=pid, name=data["player_name"].get(pid, f"Player {pid}"),
            height=data["player_height"].get(pid),
            height_in=parse_height_inches(data["player_height"].get(pid)),
            team_id=ps["team_id"], division=ps["division"], position=ps["position"],
            class_year=ps["class_year"], in_transfer_portal=data["player_transfer_portal"].get(pid),
            season=season, games=0, total_minutes=0,
            avg_minutes=None, ppg=None, rpg=None, apg=None, bpg=None, spg=None, topg=None,
            ts_pct=None, fg_pct=None,
            per40_pts=None, per40_reb=None, per40_ast=None, per40_blk=None, per40_stl=None, per40_tov=None,
            hoop_score=None, hoop_score_raw=None,
            thin_sample=1,
        ))
        covered_pids.add(pid)

    # Last resort: a player with a Team ID on the Players sheet pointing at
    # a team that's active this season (has real games in `rat`), but no
    # PlayerSeasons row for ANY season and no PlayerGameStats rows anywhere
    # -- the workbook simply never got season/game data for her at all.
    # Still added as a bare placeholder so she's findable/on her roster
    # instead of invisible (see sheet_meta docstring above).
    #
    # Critically restricted to players with ZERO PlayerSeasons rows across
    # the WHOLE workbook (not just this season) -- Team ID on the Players
    # sheet is not season-scoped, so a player with PlayerSeasons rows only
    # for 2023-24/2024-25 (graduated/departed, not on the roster this
    # season) still has that same stale Team ID sitting on her Players-
    # sheet row. Checked against real data: 3,580 of 8,740 Players-sheet
    # rows are exactly this case -- without this guard, every one of them
    # would incorrectly show up as a "current" 2025-26 roster player.
    #
    # A second guard is needed beyond that: even among the players with
    # zero PlayerSeasons rows ever, some Players-sheet rows are not real
    # current roster entries at all -- a real example found in the data,
    # "Sarah Bandoma" (Team ID pointing at VCU) is not actually on VCU's
    # 2025-26 roster, unlike a real example that IS meant to show up here,
    # Sydney Fenn (Indiana). The reliable signal that separates them: a
    # genuine roster-page scrape hit always populated BOTH Height and Class
    # together (Fenn has both); a stub/erroneous row has BOTH null (Bandoma
    # has neither). Checked against the full workbook: of 435 players with
    # zero PlayerSeasons rows, exactly 170 have both fields and 265 have
    # neither -- no partial cases -- so requiring both is a clean filter,
    # not a guess.
    if sheet_meta:
        already_seasoned = data.get("players_with_any_season", set())
        for pid, meta in sheet_meta.items():
            if pid in covered_pids or pid in already_seasoned:
                continue
            tid = meta.get("team_id")
            if tid not in rat:
                continue  # not on a team with games this season -- likely a stale/graduated record
            if not meta.get("class_year") or not data["player_height"].get(pid):
                continue  # no real roster bio data scraped -- likely a stub/erroneous entry, not a real roster player
            player_rows.append(dict(
                player_id=pid, name=data["player_name"].get(pid, f"Player {pid}"),
                height=data["player_height"].get(pid),
                height_in=parse_height_inches(data["player_height"].get(pid)),
                team_id=tid, division=meta.get("division"), position=meta.get("position"),
                class_year=meta.get("class_year"), in_transfer_portal=data["player_transfer_portal"].get(pid),
                season=season, games=0, total_minutes=0,
                avg_minutes=None, ppg=None, rpg=None, apg=None, bpg=None, spg=None, topg=None,
                ts_pct=None, fg_pct=None,
                per40_pts=None, per40_reb=None, per40_ast=None, per40_blk=None, per40_stl=None, per40_tov=None,
                hoop_score=None, hoop_score_raw=None,
                thin_sample=1,
            ))
            covered_pids.add(pid)

    team_rows = []
    school_level_fallback_count = 0
    for tid, info in data["teams"].items():
        if tid not in rat:
            continue
        # Prefer the explicit "School Level" column from the Teams sheet
        # (user-curated, confirmed to match classify_tier()'s heuristic for
        # every team in the current women's workbook) -- fall back to the
        # heuristic only if a workbook is missing the column, or (as is
        # currently true for every men's team -- see classify_tier()'s
        # docstring) has it entirely blank.
        school_level = info.get("school_level")
        if school_level:
            tier = school_level
        else:
            tier = classify_tier(info["name"], info["conference"], sport=sport)
            school_level_fallback_count += 1
        team_rows.append(dict(
            team_id=tid, name=info["name"], division=info["division"], conference=info["conference"],
            tier=tier,
            current_rating=rat[tid], sos=sos.get(tid),
        ))
 
    return dict(
        team_rows=team_rows, player_rows=player_rows,
        league_mean_rat=league_mean_rat, league_std_rat=league_std_rat,
        thin_count=thin_count, team_boxscore_games=team_boxscore_games,
        school_level_fallback_count=school_level_fallback_count,
    )
 
 
def compute_team_profiles(player_rows, team_boxscore_games):
    """Real team-level category rates -- each team's roster totals for the
    season (points/rebounds/assists/blocks/steals/turnovers/makes/attempts),
    summed across every player, divided by the number of games the team has
    REAL BOX-SCORE DATA for (team_boxscore_games -- see compute_season_
    profiles' docstring for why this must be the box-score count, not the
    schedule's scored-game count: the two differ for any team with partial
    box-score coverage, and dividing by the larger schedule count would
    silently understate that team's real per-game rate).

    This used to divide by the sum of individual players' per40 rates
    weighted by their own minutes, which is a materially different (and
    wrong) number: a "per-40" rate is meant to extrapolate ONE player's own
    production to a full 40-minute game. A team fields 5 players
    simultaneously, so its roster's total minutes played across a season is
    roughly 5x the team's own game-minutes (200 team-minutes per 40-minute
    game) -- averaging individual per-40 rates weighted by those roster
    minutes converges to roughly team-production-divided-by-5, i.e.
    something close to a single average player's rate, not the team's own
    output. That's exactly the bug a user reported after comparing the
    Stats Breakdown tab's numbers to what the team actually did on the
    court. The fix: sum real totals across the roster, divide by the team's
    own games played, so per40_pts is a team's actual points production
    normalized to one 40-minute game -- effectively the team's real PPG
    (barring OT games, a small effect at this scale).

    ts_pct/fg_pct are likewise now real team shooting splits (total makes /
    total attempts across every player, weighted naturally by how much each
    player actually shot) instead of a per-player-average percentage."""
    acc = defaultdict(lambda: defaultdict(float))
    roster_size = defaultdict(int)
    teams_seen = set()

    for p in player_rows:
        tid = p["team_id"]
        if not p["total_minutes"]:
            continue
        teams_seen.add(tid)
        roster_size[tid] += 1
        acc[tid]["_total_pts"] += p["_total_pts"]
        acc[tid]["_total_reb"] += p["_total_reb"]
        acc[tid]["_total_ast"] += p["_total_ast"]
        acc[tid]["_total_blk"] += p["_total_blk"]
        acc[tid]["_total_stl"] += p["_total_stl"]
        acc[tid]["_total_tov"] += p["_total_tov"]
        acc[tid]["_total_fgm"] += p["_total_fgm"]
        acc[tid]["_total_fga"] += p["_total_fga"]
        acc[tid]["_true_shot_attempts"] += p["_true_shot_attempts"]

    rows = []
    for tid in teams_seen:
        games = team_boxscore_games.get(tid, 0)
        if games <= 0:
            continue  # can't normalize to a per-game/per-40 rate with no box-score games on record
        a = acc[tid]
        row = dict(
            team_id=tid, roster_size=roster_size[tid],
            per40_pts=a["_total_pts"] / games,
            per40_reb=a["_total_reb"] / games,
            per40_ast=a["_total_ast"] / games,
            per40_blk=a["_total_blk"] / games,
            per40_stl=a["_total_stl"] / games,
            per40_tov=a["_total_tov"] / games,
            ts_pct=(a["_total_pts"] / (2 * a["_true_shot_attempts"])) if a["_true_shot_attempts"] > 0 else None,
            fg_pct=(a["_total_fgm"] / a["_total_fga"]) if a["_total_fga"] > 0 else None,
        )
        rows.append(row)
    return rows
 
 
def league_profile_stats(team_profile_rows):
    """League-wide mean/std for each team_profile category, so a single
    team's rate can be turned into a z-score (how unusual is this,
    relative to every other team's roster this season)."""
    stats = {}
    for stat in TEAM_PROFILE_STATS + TEAM_PROFILE_PCT_STATS:
        vals = [r[stat] for r in team_profile_rows if r.get(stat) is not None]
        if vals:
            mean = statistics.mean(vals)
            std = (statistics.pstdev(vals) if len(vals) > 1 else 1.0) or 1.0
        else:
            mean, std = 0.0, 1.0
        stats[stat] = dict(mean=mean, std=std)
    return stats
 
 
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", required=True, help="Path to the sport's workbook, e.g. "
                                                        "WomensSummitTPE.xlsx or MensSummitTPE.xlsx.")
    parser.add_argument("--sport", required=True, choices=["women", "men"],
                         help="Which sport this workbook covers. Picks the tier-classification "
                              "fallback (see classify_tier() in summit_calc.py) and the default "
                              "--out filename -- women's and men's data must never be pooled into "
                              "the same cache file (see ARCHITECTURE_HOSTING_PLAN.md's z-score "
                              "population constraint).")
    parser.add_argument("--out", default=None, help="Defaults to summit_tpe_cache_<sport>.sqlite "
                                                      "if not given.")
    args = parser.parse_args()
    out_path = args.out or f"summit_tpe_cache_{args.sport}.sqlite"

    data = load(args.path)
    season = pick_latest_season(data["games_by_season"])
    print(f"\nSport: {args.sport}")
    print(f"Current season: {season}")

    current = compute_season_profiles(data, season, sheet_meta=data["player_sheet_meta"], sport=args.sport)
    team_rows, player_rows = current["team_rows"], current["player_rows"]
    league_mean_rat, league_std_rat = current["league_mean_rat"], current["league_std_rat"]
    print(f"Rat: n={len(team_rows)} mean={league_mean_rat:.2f} std={league_std_rat:.2f}")
    print(f"Players this season: {len(player_rows)} total "
          f"({current['thin_count']} flagged thin_sample -- below {MIN_GAMES_FOR_PROFILE} games / "
          f"{MIN_TOTAL_MINUTES_FOR_PROFILE} total minutes; still included, just caveated on the site)")
    tier_counts = defaultdict(int)
    for t in team_rows:
        tier_counts[t["tier"]] += 1
    print(f"Team ratings: {len(team_rows)} (by tier: {dict(tier_counts)})")
    if current["school_level_fallback_count"]:
        pct = 100 * current["school_level_fallback_count"] / len(team_rows) if team_rows else 0
        print(f"  NOTE: {current['school_level_fallback_count']}/{len(team_rows)} teams "
              f"({pct:.0f}%) had no 'School Level' value on the Teams sheet -- tier for those came "
              f"from classify_tier()'s conference-based fallback instead (sport={args.sport!r}). "
              f"See that function's docstring, especially for sport='men' (currently 100% "
              f"fallback -- not yet confirmed with the user).")

    # ---- team category profiles (for /teams/{id}/needs, /teams/{id}/fits) ----
    team_profile_rows = compute_team_profiles(player_rows, current["team_boxscore_games"])
    profile_stats = league_profile_stats(team_profile_rows)
    print(f"Team category profiles: {len(team_profile_rows)}")

    # ---- player history across EVERY season (for /players/{id}/trajectory) ----
    print("\nComputing player history across all seasons ...")
    player_history_rows = []
    for hseason in data["games_by_season"]:
        if hseason == season:
            hres = current
        else:
            hres = compute_season_profiles(data, hseason, sport=args.sport)
        for p in hres["player_rows"]:
            team_name = data["teams"].get(p["team_id"], {}).get("name")
            player_history_rows.append(dict(p, team_name=team_name))
        print(f"  {hseason}: {len(hres['player_rows'])} player-season rows")
    print(f"player_history total: {len(player_history_rows)} rows")
 
    # ---- write sqlite ----
    print(f"\nWriting {out_path} ...")
    conn = sqlite3.connect(out_path)
    for tbl in ("teams", "players", "meta", "team_profile", "player_history", "player_game_logs", "games"):
        conn.execute(f"DROP TABLE IF EXISTS {tbl}")
 
    conn.execute("CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT)")
    meta_rows = [
        ("sport", args.sport),
        ("season", season),
        ("league_mean_rat", str(league_mean_rat)),
        ("league_std_rat", str(league_std_rat)),
    ]
    for stat, s in profile_stats.items():
        meta_rows.append((f"league_mean_{stat}", str(s["mean"])))
        meta_rows.append((f"league_std_{stat}", str(s["std"])))
    conn.executemany("INSERT INTO meta VALUES (?, ?)", meta_rows)
 
    conn.execute("""
        CREATE TABLE teams (
            team_id INTEGER PRIMARY KEY, name TEXT, division TEXT, conference TEXT,
            tier TEXT, current_rating REAL, sos REAL
        )
    """)
    conn.execute("""
        CREATE TABLE players (
            player_id INTEGER PRIMARY KEY, name TEXT, height TEXT, height_in INTEGER, team_id INTEGER, division TEXT,
            position TEXT, class_year TEXT, season TEXT, games INTEGER, total_minutes REAL,
            avg_minutes REAL, ppg REAL, rpg REAL, apg REAL, bpg REAL, spg REAL, topg REAL,
            ts_pct REAL, fg_pct REAL,
            per40_pts REAL, per40_reb REAL, per40_ast REAL, per40_blk REAL, per40_stl REAL, per40_tov REAL,
            hoop_score REAL, hoop_score_raw REAL,
            in_transfer_portal INTEGER, thin_sample INTEGER DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE team_profile (
            team_id INTEGER PRIMARY KEY, roster_size INTEGER,
            per40_pts REAL, per40_reb REAL, per40_ast REAL, per40_blk REAL, per40_stl REAL, per40_tov REAL,
            ts_pct REAL, fg_pct REAL
        )
    """)
    conn.execute("""
        CREATE TABLE player_history (
            player_id INTEGER, name TEXT, season TEXT, team_id INTEGER, team_name TEXT,
            division TEXT, position TEXT, class_year TEXT, games INTEGER, total_minutes REAL,
            avg_minutes REAL, ppg REAL, rpg REAL, apg REAL, bpg REAL, spg REAL, topg REAL,
            ts_pct REAL, fg_pct REAL,
            per40_pts REAL, per40_reb REAL, per40_ast REAL, per40_blk REAL, per40_stl REAL, per40_tov REAL,
            hoop_score REAL, hoop_score_raw REAL, thin_sample INTEGER DEFAULT 0,
            PRIMARY KEY (player_id, season)
        )
    """)
 
    conn.executemany(
        "INSERT INTO teams VALUES (:team_id,:name,:division,:conference,:tier,:current_rating,:sos)",
        team_rows,
    )
    conn.executemany(
        """INSERT INTO players VALUES (:player_id,:name,:height,:height_in,:team_id,:division,:position,:class_year,
           :season,:games,:total_minutes,:avg_minutes,:ppg,:rpg,:apg,:bpg,:spg,:topg,:ts_pct,:fg_pct,
           :per40_pts,:per40_reb,:per40_ast,:per40_blk,:per40_stl,:per40_tov,
           :hoop_score,:hoop_score_raw,:in_transfer_portal,:thin_sample)""",
        player_rows,
    )
    conn.executemany(
        """INSERT INTO team_profile VALUES (:team_id,:roster_size,:per40_pts,:per40_reb,:per40_ast,
           :per40_blk,:per40_stl,:per40_tov,:ts_pct,:fg_pct)""",
        team_profile_rows,
    )
    conn.executemany(
        """INSERT INTO player_history VALUES (:player_id,:name,:season,:team_id,:team_name,:division,
           :position,:class_year,:games,:total_minutes,:avg_minutes,:ppg,:rpg,:apg,:bpg,:spg,:topg,:ts_pct,:fg_pct,
           :per40_pts,:per40_reb,:per40_ast,:per40_blk,:per40_stl,:per40_tov,:hoop_score,:hoop_score_raw,:thin_sample)""",
        player_history_rows,
    )
    # ---- player_game_logs: every individual game a player appeared in,
    # across every season in the workbook -- powers per-season game logs on
    # the player profile page and opponent-level split leaderboards (a game
    # log's opponent_level is denormalized straight from the source sheet,
    # so those leaderboards are a plain filter/group-by, no join needed). ----
    conn.execute("""
        CREATE TABLE player_game_logs (
            player_id INTEGER, season TEXT, game_id TEXT, date TEXT,
            team_id INTEGER, opponent_team_id INTEGER, opponent_name TEXT, opponent_division TEXT,
            started INTEGER, minutes REAL, points INTEGER, rebounds INTEGER, assists INTEGER,
            steals INTEGER, blocks INTEGER, turnovers INTEGER, fouls INTEGER,
            fgm INTEGER, fga INTEGER, tfgm INTEGER, tfga INTEGER, ftm INTEGER, fta INTEGER
        )
    """)
    conn.executemany(
        """INSERT INTO player_game_logs VALUES (:player_id,:season,:game_id,:date,:team_id,
           :opponent_team_id,:opponent_name,:opponent_division,:started,:minutes,:points,:rebounds,
           :assists,:steals,:blocks,:turnovers,:fouls,:fgm,:fga,:tfgm,:tfga,:ftm,:fta)""",
        data["game_log_rows"],
    )

    # ---- games: full schedule/results, every season -- powers the team
    # Schedule tab. ----
    conn.execute("""
        CREATE TABLE games (
            game_id TEXT PRIMARY KEY, season TEXT, date TEXT,
            home_team_id INTEGER, home_team_name TEXT, away_team_id INTEGER, away_team_name TEXT,
            home_score INTEGER, away_score INTEGER, winner_team_id INTEGER, margin INTEGER,
            neutral_site INTEGER, overtime INTEGER, conference_game INTEGER
        )
    """)
    conn.executemany(
        """INSERT OR IGNORE INTO games VALUES (:game_id,:season,:date,:home_team_id,:home_team_name,
           :away_team_id,:away_team_name,:home_score,:away_score,:winner_team_id,:margin,
           :neutral_site,:overtime,:conference_game)""",
        data["schedule_rows"],
    )

    conn.execute("CREATE INDEX idx_players_name ON players(name)")
    conn.execute("CREATE INDEX idx_players_team ON players(team_id)")
    conn.execute("CREATE INDEX idx_teams_tier ON teams(tier)")
    conn.execute("CREATE INDEX idx_history_player ON player_history(player_id)")
    conn.execute("CREATE INDEX idx_gamelogs_player_season ON player_game_logs(player_id, season)")
    conn.execute("CREATE INDEX idx_gamelogs_oppteam ON player_game_logs(opponent_team_id)")
    conn.execute("CREATE INDEX idx_gamelogs_team ON player_game_logs(team_id)")
    conn.execute("CREATE INDEX idx_games_season ON games(season)")
    conn.execute("CREATE INDEX idx_games_home ON games(home_team_id)")
    conn.execute("CREATE INDEX idx_games_away ON games(away_team_id)")
    conn.commit()
    conn.close()
    print("Done.")
 
 
if __name__ == "__main__":
    main()