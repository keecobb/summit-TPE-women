"""One-off data fix: Texas A&M's real 3-season roster/schedule/box-score data
is split across two Team IDs due to a scraping bug -- Team ID 299 and Team ID
300 both carry the same ESPN Team ID (245) in the Teams sheet, and whatever
pulled game/roster data via that ESPN ID wrote most of it under 300 instead
of 299. Verified directly against the real data before writing this script:
 
  - 72 games (all 3 seasons) tagged Team ID 300 are text-labeled "Texas A&M
    Aggies" against an unmistakable SEC schedule (Alabama, LSU, Tennessee,
    Georgia, Oklahoma, South Carolina, Florida, Arkansas, ...).
  - 16 games tagged Team ID 299 are text-labeled "Texas A&M" -- the same
    real program, just the smaller slice that happened to get the correct ID.
  - ZERO games anywhere in the workbook are genuinely Texas A&M-Corpus
    Christi (Southland Conference) -- that program's real data was never
    actually collected under either ID.
  - No other Team ID pair in the whole 362-team Teams sheet shares an ESPN
    Team ID -- this is an isolated, one-off error, not a systemic problem.
  - No game row has {Home, Away} == {299, 300} (i.e. no case would become a
    team-vs-itself game after merging), so a full blanket reassignment of
    300 -> 299 is safe with no special-casing needed.
 
What this script does, in order:
  1. Read-only pre-pass (cheap, safe on a bloated workbook -- see
     apply_results.py's find_real_last_row for why this pattern is used)
     to count every cell that will change, so --dry-run can report exact
     impact without the expensive write-mode load.
  2. Reassigns Team ID 300 -> 299 everywhere it represents the real Texas
     A&M data:
       - Players sheet: Team ID column
       - PlayerSeasons sheet: Team ID column
       - PlayerGameStats sheet: Team ID column (whose stat line this is)
         AND Opponent Team ID column (other teams' box scores from games
         played against Texas A&M, which also currently point at 300)
       - Games sheet: Home Team ID and Away Team ID columns
  3. Resets Team ID 300's OWN row on the Teams sheet (Record/Wins/Losses/
     Rat/Off/Def/SoS/Summit Rat/Summit Off/Summit Def/Summit SoS/Summit
     Season/Summit Last Updated/Last Scraped Season) to blank, since those
     numbers were actually computed from Texas A&M's real games under the
     wrong ID -- leaving them in place would make Team 300 (the real
     Texas A&M-Corpus Christi record) look like it has real performance
     data when it doesn't. Also clears the duplicated ESPN Team ID and
     marks Scrape Status = "Needs Rescrape" with an explanatory note in
     Scrape Error, so it's visibly flagged in the sheet itself rather than
     silently blank.
  4. Saves atomically (temp file + os.replace), with a retry loop if the
     file is open in Excel -- same pattern as your other scripts.
 
IMPORTANT: after running this, Texas A&M's team rating and every derived
number that depends on it (TeamRatings, SeasonRankings, the calculator
cache) needs to be recomputed -- re-run compute_derived_sheets.py and
build_cache.py once this fix is applied. Team 300 will then legitimately
show zero games/players until Texas A&M-Corpus Christi's real data is
scraped in with a corrected ESPN Team ID (see the accompanying note on
run_d1_scrape.py for whether your existing scraper can do that).
 
Usage:
    python fix_team_id_299_300.py --dry-run     # report what would change
    python fix_team_id_299_300.py               # apply the fix
    python fix_team_id_299_300.py --path "C:\\...\\WomensSummitTPE.xlsx"
 
Requires: openpyxl.
"""
 
import argparse
import os
import sys
import time
 
WRONG_TEAM_ID = 300
CORRECT_TEAM_ID = 299
 
# (sheet name, column header) pairs to reassign WRONG_TEAM_ID -> CORRECT_TEAM_ID
FIX_TARGETS = [
    ("Players", "Team ID"),
    ("PlayerSeasons", "Team ID"),
    ("PlayerGameStats", "Team ID"),
    ("PlayerGameStats", "Opponent Team ID"),
    ("Games", "Home Team ID"),
    ("Games", "Away Team ID"),
]
 
# Columns on Team 300's OWN Teams-sheet row to blank out, since they were
# actually computed from Texas A&M's real games under the wrong ID.
TEAMS_ROW_RESET_COLUMNS = [
    "Record", "Wins", "Losses", "Rat", "Off", "Def", "SoS",
    "Summit Rat", "Summit Off", "Summit Def", "Summit SoS",
    "Summit Season", "Summit Last Updated", "Last Scraped Season", "Last Updated",
]
 
 
def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping
 
 
def scan_counts(path):
    """Read-only pre-pass: count how many cells will change per (sheet, column).
    Groups targets by sheet so each sheet is scanned exactly once, even when
    multiple columns on it need checking (PlayerGameStats and Games each
    have 2 target columns) -- scanning PlayerGameStats twice separately
    took long enough to time out during testing."""
    import openpyxl
    from collections import defaultdict
    targets_by_sheet = defaultdict(list)
    for sheet_name, col_name in FIX_TARGETS:
        targets_by_sheet[sheet_name].append(col_name)
 
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    counts = {}
    for sheet_name, col_names in targets_by_sheet.items():
        print(f"  scanning {sheet_name} ...")
        ws = wb[sheet_name]
        hmap = header_map(ws)
        col_indices = {}
        for col_name in col_names:
            if col_name not in hmap:
                print(f"  [!] {sheet_name}.{col_name} not found -- skipping")
                counts[(sheet_name, col_name)] = None
            else:
                col_indices[col_name] = hmap[col_name] - 1
        if not col_indices:
            continue
        running = {col_name: 0 for col_name in col_indices}
        for row in ws.iter_rows(min_row=2, values_only=True):
            for col_name, idx in col_indices.items():
                if row[idx] == WRONG_TEAM_ID:
                    running[col_name] += 1
        for col_name, n in running.items():
            counts[(sheet_name, col_name)] = n
    wb.close()
    return counts
 
 
def apply_fix(ws, col_name, dry_run):
    hmap = header_map(ws)
    if col_name not in hmap:
        print(f"  [!] {ws.title}.{col_name} not found -- skipping")
        return 0
    col = hmap[col_name]
    if dry_run:
        return None  # counts already reported from the pre-pass
    n = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[col - 1]
        if cell.value == WRONG_TEAM_ID:
            cell.value = CORRECT_TEAM_ID
            n += 1
    return n
 
 
def reset_wrong_team_row(ws, dry_run):
    hmap = header_map(ws)
    team_id_col = hmap.get("Team ID")
    if team_id_col is None:
        print("  [!] Teams sheet has no 'Team ID' column -- can't locate the row to reset")
        return
    target_row = None
    for row in ws.iter_rows(min_row=2):
        if row[team_id_col - 1].value == WRONG_TEAM_ID:
            target_row = row
            break
    if target_row is None:
        print(f"  [!] Could not find Team ID {WRONG_TEAM_ID}'s row on the Teams sheet")
        return
    if dry_run:
        print(f"  Would reset {len(TEAMS_ROW_RESET_COLUMNS)} stale stat columns on Team ID "
              f"{WRONG_TEAM_ID}'s row, clear its duplicated ESPN Team ID, and flag it "
              f"Scrape Status='Needs Rescrape'")
        return
    for col_name in TEAMS_ROW_RESET_COLUMNS:
        col = hmap.get(col_name)
        if col:
            target_row[col - 1].value = None
    espn_col = hmap.get("ESPN Team ID")
    if espn_col:
        target_row[espn_col - 1].value = None
    status_col = hmap.get("Scrape Status")
    if status_col:
        target_row[status_col - 1].value = "Needs Rescrape"
    error_col = hmap.get("Scrape Error")
    if error_col:
        target_row[error_col - 1].value = (
            f"Was sharing ESPN Team ID with Team ID {CORRECT_TEAM_ID} (Texas A&M) -- that team's "
            f"real games/roster were mistakenly scraped under this ID and have been moved to "
            f"{CORRECT_TEAM_ID}. This team's own real data (Texas A&M-Corpus Christi) still needs "
            f"the correct ESPN Team ID looked up and a fresh scrape."
        )
    print(f"  Reset Team ID {WRONG_TEAM_ID}'s stale stat columns, cleared its ESPN Team ID, "
          f"flagged Scrape Status='Needs Rescrape'")
 
 
def save_with_retry(wb, path, attempts=5, wait_seconds=10):
    tmp_path = f"{path}.tmp_saving"
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, path)
            return
        except PermissionError as exc:
            last_error = exc
            print(f"  [!] Could not save {path} (attempt {attempt}/{attempts}) -- "
                  f"is it open in Excel? Close it and I'll retry in {wait_seconds}s.")
            time.sleep(wait_seconds)
    raise last_error
 
 
def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--path", default="WomensSummitTPE.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
 
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required (pip install openpyxl).")
 
    print(f"Scanning {args.path} for Team ID {WRONG_TEAM_ID} cells (read-only pre-pass) ...")
    counts = scan_counts(args.path)
    total = sum(n for n in counts.values() if n)
    for (sheet_name, col_name), n in counts.items():
        if n is not None:
            print(f"  {sheet_name}.{col_name}: {n} cells currently = {WRONG_TEAM_ID}")
    print(f"  Total: {total} cells will be reassigned to {CORRECT_TEAM_ID}")
 
    if total == 0:
        print("\nNothing to fix -- has this already been applied?")
        return
 
    if args.dry_run:
        print(f"\n--dry-run: would also reset Team ID {WRONG_TEAM_ID}'s stale stat columns on the "
              f"Teams sheet (see script docstring). Not saving.")
        return
 
    print(f"\nOpening {args.path} ...")
    wb = openpyxl.load_workbook(args.path)
 
    print("\nReassigning Team ID cells ...")
    applied_total = 0
    for sheet_name, col_name in FIX_TARGETS:
        n = apply_fix(wb[sheet_name], col_name, dry_run=False)
        print(f"  {sheet_name}.{col_name}: reassigned {n} cells")
        applied_total += n
    if applied_total != total:
        print(f"  [!] Warning: pre-pass counted {total} cells but the write pass changed "
              f"{applied_total} -- the file may have changed between scans. Review before trusting this run.")
 
    print("\nResetting Team ID 300's stale Teams-sheet row ...")
    reset_wrong_team_row(wb["Teams"], dry_run=False)
 
    print("\nSaving ...")
    save_with_retry(wb, args.path)
    print("Done. Next: re-run compute_derived_sheets.py and build_cache.py to pick up the "
          "corrected Texas A&M data in TeamRatings/SeasonRankings/the calculator cache.")
 
 
if __name__ == "__main__":
    main()