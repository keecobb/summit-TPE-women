"""Read-only check: now that you've fixed the Games sheet's Team ID columns
for Team 300 by hand, this checks whether PlayerGameStats actually has a
stat line for Team 300 on every one of those games -- or whether some are
still missing because run_d1_scrape.py's Game ID dedup skipped fetching the
box score for that event entirely (it dedups on Game ID alone, so a game
that already existed in the Games sheet -- even with the wrong/blank Team ID
-- never got its box score (re)fetched, regardless of what the Team ID
column says now).
 
Does NOT touch the workbook. Just reports the gap.
 
Usage:
    python check_pgs_gaps_team_300.py
    python check_pgs_gaps_team_300.py --team-id 300 --path "C:\\...\\WomensSummitTPE.xlsx"
"""
 
import argparse
 
 
def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping
 
 
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", default="WomensSummitTPE.xlsx")
    parser.add_argument("--team-id", type=int, default=300)
    args = parser.parse_args()
 
    import openpyxl
    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
    team_id = args.team_id
 
    games_ws = wb["Games"]
    ghmap = header_map(games_ws)
    home_id_col = ghmap["Home Team ID"] - 1
    away_id_col = ghmap["Away Team ID"] - 1
    game_id_col = ghmap["Game ID"] - 1
    season_col = ghmap["Season"] - 1
    date_col = ghmap["Date"] - 1
    home_name_col = ghmap["Home Team"] - 1
    away_name_col = ghmap["Away Team"] - 1
 
    games_for_team = {}   # game_id -> {season, date, opponent}
    for row in games_ws.iter_rows(min_row=2, values_only=True):
        gid = row[game_id_col]
        if not gid:
            continue
        home_id, away_id = row[home_id_col], row[away_id_col]
        if home_id == team_id:
            opp = row[away_name_col]
        elif away_id == team_id:
            opp = row[home_name_col]
        else:
            continue
        games_for_team[str(gid)] = {
            "season": row[season_col], "date": row[date_col], "opponent": opp,
        }
 
    pgs_ws = wb["PlayerGameStats"]
    phmap = header_map(pgs_ws)
    pgs_team_col = phmap["Team ID"] - 1
    pgs_game_col = phmap["Game ID"] - 1
    games_with_stats = set()
    stat_rows_for_team = 0
    for row in pgs_ws.iter_rows(min_row=2, values_only=True):
        if row[pgs_team_col] == team_id:
            gid = row[pgs_game_col]
            if gid:
                games_with_stats.add(str(gid))
            stat_rows_for_team += 1
 
    wb.close()
 
    print(f"Games sheet: {len(games_for_team)} games involving Team {team_id}")
    print(f"PlayerGameStats: {stat_rows_for_team} rows for Team {team_id}, "
          f"covering {len(games_with_stats)} distinct games")
 
    missing = sorted(
        (gid, info) for gid, info in games_for_team.items() if gid not in games_with_stats
    )
    print(f"\nGames with a Games-sheet row for Team {team_id} but ZERO PlayerGameStats rows: "
          f"{len(missing)}")
    if missing:
        by_season = {}
        for gid, info in missing:
            by_season.setdefault(info["season"], 0)
            by_season[info["season"]] += 1
        for season, n in sorted(by_season.items(), key=lambda x: str(x[0])):
            print(f"    {season}: {n} games missing stats")
        print("\nFirst 15 missing (Game ID, Season, Date, Opponent):")
        for gid, info in missing[:15]:
            print(f"    {gid}  {info['season']}  {info['date']}  vs {info['opponent']}")
 
 
if __name__ == "__main__":
    main()
 