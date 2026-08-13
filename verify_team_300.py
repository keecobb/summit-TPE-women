"""Quick read-only sanity check: did the Texas A&M-Corpus Christi (Team ID
300) scrape actually land real data, or did it silently find zero games?
 
Usage:
    python verify_team_300.py
    python verify_team_300.py --path "C:\\...\\WomensSummitTPE.xlsx"
"""
 
import argparse
 
TEAM_ID = 300
 
 
def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping
 
 
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", default="WomensSummitTPE.xlsx")
    args = parser.parse_args()
 
    import openpyxl
    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
 
    # Teams row for 300
    ws = wb["Teams"]
    hmap = header_map(ws)
    tid_col = hmap["Team ID"] - 1
    name_col = hmap.get("Name", hmap.get("Team Name"))
    espn_col = hmap.get("ESPN Team ID")
    status_col = hmap.get("Scrape Status")
    last_season_col = hmap.get("Last Scraped Season")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[tid_col] == TEAM_ID:
            print("Teams row for 300:")
            if name_col:
                print(f"  Name: {row[name_col-1]}")
            if espn_col:
                print(f"  ESPN Team ID: {row[espn_col-1]}")
            if status_col:
                print(f"  Scrape Status: {row[status_col-1]}")
            if last_season_col:
                print(f"  Last Scraped Season: {row[last_season_col-1]}")
            break
 
    # Games count
    ws = wb["Games"]
    hmap = header_map(ws)
    home_col = hmap["Home Team ID"] - 1
    away_col = hmap["Away Team ID"] - 1
    games = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[home_col] == TEAM_ID or row[away_col] == TEAM_ID:
            games += 1
    print(f"\nGames involving Team 300: {games}")
 
    # PlayerSeasons count
    ws = wb["PlayerSeasons"]
    hmap = header_map(ws)
    pt_col = hmap["Team ID"] - 1
    seasons = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[pt_col] == TEAM_ID:
            seasons += 1
    print(f"PlayerSeasons rows for Team 300: {seasons}")
 
    # PlayerGameStats count
    ws = wb["PlayerGameStats"]
    hmap = header_map(ws)
    pg_col = hmap["Team ID"] - 1
    stats = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[pg_col] == TEAM_ID:
            stats += 1
    print(f"PlayerGameStats rows for Team 300: {stats}")
 
    wb.close()
 
 
if __name__ == "__main__":
    main()