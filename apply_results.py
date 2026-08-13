"""Applies the computed SeasonRankings / TeamRatings data (season_rankings.csv,
team_ratings.csv -- sit these next to this script, or pass --data-dir) into
your live WomensSummitTPE.xlsx. TransferProjection is deliberately not part
of this anymore -- it's a live per-request calculator now (see
summit_tpe_calculator.zip), not something to precompute into fixed rows, so
delete that scaffold sheet rather than expecting this script to fill it.

Why this exists as a separate local step instead of just handing back the
whole updated workbook: the real file is ~40-58MB (PlayerGameStats alone is
~9.4M cells), which is over both the chat upload limit and the file-transfer
limit for writing back to your machine through this session's device bridge.
The actual DELTA -- the two sheets' worth of new data -- is under 1MB, so
that's what got sent. This script does the merge on your end instead.

What it does, in order:
  1. Opens your live WomensSummitTPE.xlsx (default: same folder as this
     script -- pass --path to point elsewhere).
  2. For each of SeasonRankings / TeamRatings / TransferProjection: matches
     columns by HEADER NAME (not position), clears existing data rows, and
     writes in the new rows. If you've reordered or added columns on your
     end since the scaffold was defined, this still lines up correctly by
     name; any of your columns not present in the CSVs are simply left
     alone.
  3. Optionally (--trim-bloat) also trims the Teams and Players sheets down
     to their real data extent. As of this run those two sheets carried
     ~2.1 million empty-but-styled trailing rows (Teams' used range ran to
     row 1,047,979, Players' to the sheet's hard limit of 1,048,575) --
     that's why loading/saving this workbook currently takes several
     minutes even for a one-cell change. Trimming is a pure dimension
     shrink: it only removes EMPTY rows past your real last row, and never
     touches Players' live Transfer History / Schools Attended formulas or
     any other sheet. Off by default since it's a bigger structural change
     than the data merge itself -- run once with --trim-bloat when you're
     ready, or apply it separately later.
  4. Saves atomically (write to a temp file, then swap into place) with a
     retry loop if the file is open in Excel, same pattern as
     scrapers/xlsx_io.py's save_with_retry.

Usage:
    python apply_results.py                     # merge only
    python apply_results.py --trim-bloat         # merge + fix the bloat
    python apply_results.py --path "C:\\...\\WomensSummitTPE.xlsx"
    python apply_results.py --dry-run            # report what would happen, don't save

Requires: openpyxl (already a dependency of your other scrapers/ scripts).
"""

import argparse
import csv
import os
import sys
import time


def header_map(ws):
    mapping = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None and str(cell.value).strip():
            mapping[str(cell.value).strip()] = cell.column
    return mapping


def read_csv_rows(path):
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = []
        for raw in reader:
            row = {}
            for h, v in zip(header, raw):
                if v == "":
                    row[h] = None
                else:
                    try:
                        row[h] = int(v)
                    except ValueError:
                        try:
                            row[h] = float(v)
                        except ValueError:
                            row[h] = v
            rows.append(row)
        return header, rows


def fill_sheet(ws, header, rows, dry_run):
    hmap = header_map(ws)
    missing = [h for h in header if h not in hmap]
    if missing:
        print(f"  [!] {ws.title}: CSV has columns not found on this sheet, skipping them: {missing}")
    if dry_run:
        print(f"  {ws.title}: would clear existing data rows and write {len(rows)} new rows")
        return
    # Clear existing data rows (values only -- leaves any header formatting alone)
    if ws.max_row > 1:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
    for i, r in enumerate(rows, start=2):
        for h in header:
            col = hmap.get(h)
            if col:
                ws.cell(row=i, column=col, value=r.get(h))
    print(f"  {ws.title}: wrote {len(rows)} rows")


def find_real_last_row(path, sheet_name):
    """Cheap, low-memory PRE-PASS using read_only mode to find a sheet's
    true last non-empty row, BEFORE the expensive normal (write-mode) load.
    Doing this scan on an already-write-mode-loaded sheet (which holds every
    one of the ~1M+ bloated rows as live cell objects) risks OOM on a large
    workbook -- read_only mode streams instead, so this pass stays cheap no
    matter how bloated the sheet is."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    last = 1
    row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if any(v is not None for v in row):
            last = row_num
    wb.close()
    return last


def trim_trailing_bloat(ws, real_last, dry_run):
    """Delete everything after `real_last` (computed by find_real_last_row
    in a prior read_only pass). Safe because delete_rows on a pure trailing
    range never shifts/touches any row that has real data."""
    n_to_delete = ws.max_row - real_last
    if n_to_delete <= 0:
        print(f"  {ws.title}: no trailing bloat found (max_row={ws.max_row})")
        return
    if dry_run:
        print(f"  {ws.title}: would delete {n_to_delete} trailing empty rows "
              f"(max_row {ws.max_row} -> {real_last})")
        return
    ws.delete_rows(real_last + 1, n_to_delete)
    print(f"  {ws.title}: trimmed {n_to_delete} trailing empty rows "
          f"(max_row {ws.max_row + n_to_delete} -> {ws.max_row})")


def save_with_retry(wb, path, attempts=5, wait_seconds=10):
    tmp_path = f"{path}.tmp_saving"
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, path)
            return
        except PermissionError as exc:
            last_error = exc
            print(f"  [!] Could not save {path} (attempt {attempt}/{attempts}) -- "
                  f"is it open in Excel? Close it and I'll retry in {wait_seconds}s.")
            time.sleep(wait_seconds)
    raise last_error


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--path", default="WomensSummitTPE.xlsx")
    parser.add_argument("--data-dir", default=".", help="Folder containing the 3 CSVs (default: current folder)")
    parser.add_argument("--trim-bloat", action="store_true",
                         help="Also trim Teams/Players trailing empty rows (see module docstring)")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required (pip install openpyxl) -- it's already used by your other scrapers/ scripts.")

    csv_files = {
        "SeasonRankings": "season_rankings.csv",
        "TeamRatings": "team_ratings.csv",
        # TransferProjection is intentionally not here -- that's now a live
        # calculator (see the separate summit_tpe_calculator.zip), not a
        # precomputed sheet. Delete the TransferProjection tab if you
        # haven't already.
    }
    data = {}
    for sheet_name, fname in csv_files.items():
        path = os.path.join(args.data_dir, fname)
        if not os.path.exists(path):
            sys.exit(f"Missing {path} -- keep the 3 CSVs next to this script or pass --data-dir")
        header, rows = read_csv_rows(path)
        data[sheet_name] = (header, rows)
        print(f"Loaded {fname}: {len(rows)} rows")

    real_last_rows = {}
    if args.trim_bloat:
        print("\nScanning Teams/Players for their real data extent (read-only pass, low memory) ...")
        for name in ("Teams", "Players"):
            real_last_rows[name] = find_real_last_row(args.path, name)
            print(f"  {name}: real last row = {real_last_rows[name]}")

    print(f"\nOpening {args.path} ...")
    wb = openpyxl.load_workbook(args.path)

    print("\nWriting computed sheets ...")
    for sheet_name, (header, rows) in data.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [!] Sheet '{sheet_name}' not found in workbook, skipping.")
            continue
        fill_sheet(wb[sheet_name], header, rows, args.dry_run)

    if args.trim_bloat:
        print("\nTrimming Teams/Players trailing bloat ...")
        for name in ("Teams", "Players"):
            trim_trailing_bloat(wb[name], real_last_rows[name], args.dry_run)

    if args.dry_run:
        print("\n--dry-run: not saving.")
        return

    print("\nSaving ...")
    save_with_retry(wb, args.path)
    print("Done.")


if __name__ == "__main__":
    main()
